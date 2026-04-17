import os
import re
import warnings
from datetime import datetime, time, timedelta

import matplotlib
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from scipy import stats
from scipy.stats import kstest

matplotlib.use("Agg")
import matplotlib.gridspec as gridspec
import matplotlib.pyplot as plt

warnings.filterwarnings("ignore")

# =============================================================================
# AYARLAR
# =============================================================================

KAYNAK_DOSYA = "DBV2-2.xlsm"
CIKTI_KLASOR = "analiz"
ALPHA = 0.05
GECERLI_MIKRON = {7, 13, 25}
MIN_GOZLEM = 3
BOOTSTRAP_ITER = 80
BOOTSTRAP_ATTEMPT_KATSAYI = 4
TRAIN_ORAN = 0.7
RNG_SEED = 42

HAVUZ_BASLANGIC = "YağAlma"
HAVUZ_BITIS = "Bufferda Bekleme"

DAGITIMLAR = {
    "Normal": stats.norm,
    "Lognormal": stats.lognorm,
    "Exponential": stats.expon,
    "Gamma": stats.gamma,
    "Weibull": stats.weibull_min,
    "LogLogistic": stats.fisk,
    "Cauchy": stats.cauchy,
    "Laplace": stats.laplace,
}

ADAY_DAGILIMLAR = {"Lognormal", "Exponential", "Gamma", "Weibull", "LogLogistic"}
DIAGNOSTIC_DAGILIMLAR = {"Normal", "Cauchy", "Laplace"}

SKOR_KOVALARI = {
    "kucuk": {
        "_AICc_num": 0.18,
        "_BIC_num": 0.07,
        "_KS_boot_stat_num": 0.05,
        "_AD_boot_stat_num": 0.05,
        "_CvM_boot_stat_num": 0.05,
        "_KS_boot_p_num": 0.07,
        "_AD_boot_p_num": 0.07,
        "_CvM_boot_p_num": 0.06,
        "_Wasserstein_num": 0.08,
        "_TT_Wasserstein_num": 0.04,
        "_TT_CvM_num": 0.04,
        "_TT_Quantile_Hata_num": 0.04,
        "_TT_KS_stat_num": 0.03,
        "_TT_KS_p_num": 0.02,
        "_TT_NLL_num": 0.05,
    },
    "orta": {
        "_AICc_num": 0.14,
        "_BIC_num": 0.06,
        "_KS_boot_stat_num": 0.04,
        "_AD_boot_stat_num": 0.04,
        "_CvM_boot_stat_num": 0.04,
        "_KS_boot_p_num": 0.06,
        "_AD_boot_p_num": 0.06,
        "_CvM_boot_p_num": 0.06,
        "_Wasserstein_num": 0.08,
        "_TT_Wasserstein_num": 0.05,
        "_TT_CvM_num": 0.04,
        "_TT_Quantile_Hata_num": 0.03,
        "_TT_KS_stat_num": 0.08,
        "_TT_KS_p_num": 0.08,
        "_TT_NLL_num": 0.14,
    },
    "buyuk": {
        "_AICc_num": 0.10,
        "_BIC_num": 0.05,
        "_KS_boot_stat_num": 0.05,
        "_AD_boot_stat_num": 0.05,
        "_CvM_boot_stat_num": 0.05,
        "_KS_boot_p_num": 0.0167,
        "_AD_boot_p_num": 0.0167,
        "_CvM_boot_p_num": 0.0166,
        "_Wasserstein_num": 0.10,
        "_TT_Wasserstein_num": 0.08,
        "_TT_CvM_num": 0.07,
        "_TT_Quantile_Hata_num": 0.05,
        "_TT_KS_stat_num": 0.10,
        "_TT_KS_p_num": 0.07,
        "_TT_NLL_num": 0.18,
    },
}


# =============================================================================
# YARDIMCI FONKSIYONLAR
# =============================================================================


def slugify(metin):
    donusum = str.maketrans(
        {
            "ç": "c",
            "Ç": "c",
            "ğ": "g",
            "Ğ": "g",
            "ı": "i",
            "İ": "i",
            "ö": "o",
            "Ö": "o",
            "ş": "s",
            "Ş": "s",
            "ü": "u",
            "Ü": "u",
        }
    )
    temiz = str(metin).translate(donusum).strip().lower()
    temiz = re.sub(r"[^a-z0-9]+", "_", temiz)
    return temiz.strip("_") or "veri_seti"


def excel_sheet_adi(metin, kullanilanlar):
    yasak = set(r"[]:*?/\\")
    temel = "".join("_" if ch in yasak else ch for ch in metin).strip() or "Sayfa"
    temel = temel[:31]

    aday = temel
    sayac = 1
    while aday in kullanilanlar:
        ek = f"_{sayac}"
        aday = f"{temel[: 31 - len(ek)]}{ek}"
        sayac += 1

    kullanilanlar.add(aday)
    return aday


def kolon_gorunur_adi(kolon_adi):
    return re.sub(r"\.\d+$", "", str(kolon_adi)).strip()


def benzersiz_havuz_adlari(kolonlar):
    sayac = {}
    sonuc = []
    for kolon in kolonlar:
        gorunen = kolon_gorunur_adi(kolon)
        sayac[gorunen] = sayac.get(gorunen, 0) + 1
        if sayac[gorunen] == 1:
            etiket = gorunen
        else:
            etiket = f"{gorunen} {sayac[gorunen]}"
        sonuc.append((kolon, etiket))
    return sonuc


def sayisal_metni_float(text):
    try:
        return float(str(text).strip().replace(",", "."))
    except Exception:
        return None


def sureyi_dakikaya_cevir(deger):
    if pd.isna(deger):
        return np.nan

    if isinstance(deger, pd.Timedelta):
        return deger.total_seconds() / 60.0

    if isinstance(deger, timedelta):
        return deger.total_seconds() / 60.0

    if isinstance(deger, (pd.Timestamp, datetime)):
        return (
            deger.hour * 60
            + deger.minute
            + deger.second / 60.0
            + deger.microsecond / 60000000.0
        )

    if isinstance(deger, time):
        return (
            deger.hour * 60
            + deger.minute
            + deger.second / 60.0
            + deger.microsecond / 60000000.0
        )

    if isinstance(deger, (np.integer, int, np.floating, float)):
        sayi = float(deger)
        if np.isnan(sayi):
            return np.nan
        if abs(sayi) <= 10:
            return sayi * 24.0 * 60.0
        return sayi

    text = str(deger).strip()
    if not text:
        return np.nan

    sayisal = sayisal_metni_float(text)
    if sayisal is not None:
        if abs(sayisal) <= 10:
            return sayisal * 24.0 * 60.0
        return sayisal

    if ":" in text:
        parcalar = text.split(":")
        try:
            if len(parcalar) == 3:
                saat, dakika, saniye = [float(p) for p in parcalar]
                return saat * 60.0 + dakika + saniye / 60.0
            if len(parcalar) == 2:
                dakika, saniye = [float(p) for p in parcalar]
                return dakika + saniye / 60.0
        except Exception:
            return np.nan

    return np.nan


def zamani_saniyeye_cevir(deger):
    if pd.isna(deger):
        return np.nan

    if isinstance(deger, pd.Timedelta):
        return deger.total_seconds()

    if isinstance(deger, timedelta):
        return deger.total_seconds()

    if isinstance(deger, (pd.Timestamp, datetime)):
        return (
            deger.hour * 3600
            + deger.minute * 60
            + deger.second
            + deger.microsecond / 1000000.0
        )

    if isinstance(deger, time):
        return (
            deger.hour * 3600
            + deger.minute * 60
            + deger.second
            + deger.microsecond / 1000000.0
        )

    if isinstance(deger, (np.integer, int, np.floating, float)):
        sayi = float(deger)
        if np.isnan(sayi):
            return np.nan
        kesir = sayi % 1 if abs(sayi) >= 1 else sayi
        return kesir * 24.0 * 3600.0

    text = str(deger).strip()
    if not text:
        return np.nan

    sayisal = sayisal_metni_float(text)
    if sayisal is not None:
        kesir = sayisal % 1 if abs(sayisal) >= 1 else sayisal
        return kesir * 24.0 * 3600.0

    if ":" in text:
        parcalar = text.split(":")
        try:
            if len(parcalar) == 3:
                saat, dakika, saniye = [float(p) for p in parcalar]
                return saat * 3600.0 + dakika * 60.0 + saniye
            if len(parcalar) == 2:
                dakika, saniye = [float(p) for p in parcalar]
                return dakika * 60.0 + saniye
        except Exception:
            return np.nan

    return np.nan


def pozitif_dizi(seri, cevirici):
    dizi = pd.to_numeric(seri.apply(cevirici), errors="coerce").dropna()
    dizi = dizi[dizi > 0]
    return dizi.astype(float).values


# =============================================================================
# VERI YUKLEME
# =============================================================================


def havuz_verilerini_hazirla(dosya, kullanilan_sayfa_adlari):
    df_hv = pd.read_excel(dosya, sheet_name="Havuz Süreler", engine="openpyxl")
    kolonlar = list(df_hv.columns)

    if HAVUZ_BASLANGIC not in kolonlar or HAVUZ_BITIS not in kolonlar:
        raise KeyError(
            f"'Havuz Süreler' sayfasinda '{HAVUZ_BASLANGIC}' ile "
            f"'{HAVUZ_BITIS}' kolonlari bulunamadi."
        )

    bas_idx = kolonlar.index(HAVUZ_BASLANGIC)
    bit_idx = kolonlar.index(HAVUZ_BITIS)
    havuz_kolonlari = kolonlar[bas_idx : bit_idx + 1]

    veri_setleri = {}
    for kolon, gorunen_ad in benzersiz_havuz_adlari(havuz_kolonlari):
        veri = pozitif_dizi(df_hv[kolon], sureyi_dakikaya_cevir)
        if len(veri) == 0:
            continue

        anahtar = f"havuzsureler_{slugify(gorunen_ad)}"
        veri_setleri[anahtar] = {
            "kategori": "havuzsureler",
            "etiket": f"Havuz Süreleri – {gorunen_ad}",
            "birim": "dakika",
            "veri": veri,
            "grup": gorunen_ad,
            "sayfa_adi": excel_sheet_adi(
                f"Havuz_{gorunen_ad}", kullanilan_sayfa_adlari
            ),
        }

    return veri_setleri


def gelisler_arasi_hazirla(dosya, kullanilan_sayfa_adlari):
    df_ag = pd.read_excel(dosya, sheet_name="Askı Gelis", engine="openpyxl")
    df_ag["sn"] = df_ag["Giriş Saati"].apply(zamani_saniyeye_cevir)
    df_ag = df_ag.dropna(subset=["sn"]).sort_values(["Tarih", "sn"])

    aralar = []
    for _, grup in df_ag.groupby("Tarih"):
        sn_dizi = np.sort(grup["sn"].values)
        if len(sn_dizi) > 1:
            farklar = np.diff(sn_dizi) / 60.0
            aralar.extend(farklar[farklar > 0].tolist())

    return {
        "gelislerarasi": {
            "kategori": "gelislerarasi",
            "etiket": "Gelişler Arası Süre",
            "birim": "dakika",
            "veri": np.array(aralar, dtype=float),
            "grup": "Gelişler Arası",
            "sayfa_adi": excel_sheet_adi("GelislerArasi", kullanilan_sayfa_adlari),
        }
    }


def eloksal_uretim_hazirla(dosya, kullanilan_sayfa_adlari):
    df_el = pd.read_excel(dosya, sheet_name="Eloksal Üretim", engine="openpyxl")
    df_el = df_el.copy()
    df_el["Mikron"] = pd.to_numeric(df_el["Mikron"], errors="coerce")
    df_el = df_el[df_el["Mikron"].isin(GECERLI_MIKRON)].copy()
    df_el = df_el.dropna(subset=["Giriş Saati", "Çıkış Saati"])

    df_el["giris_sn"] = df_el["Giriş Saati"].apply(zamani_saniyeye_cevir)
    df_el["cikis_sn"] = df_el["Çıkış Saati"].apply(zamani_saniyeye_cevir)
    df_el["sure_dk"] = (df_el["cikis_sn"] - df_el["giris_sn"]) / 60.0
    df_el = df_el.dropna(subset=["sure_dk"])
    df_el = df_el[df_el["sure_dk"] > 0].copy()

    veri_setleri = {}
    for mikron in sorted(GECERLI_MIKRON):
        seri = df_el.loc[df_el["Mikron"] == mikron, "sure_dk"].astype(float).values
        if len(seri) == 0:
            continue

        anahtar = f"eloksaluretim_mikron_{int(mikron)}"
        veri_setleri[anahtar] = {
            "kategori": "eloksaluretim",
            "etiket": f"Eloksal Üretim – {int(mikron)} Mikron",
            "birim": "dakika",
            "veri": seri,
            "grup": f"{int(mikron)} Mikron",
            "sayfa_adi": excel_sheet_adi(
                f"Eloksal_{int(mikron)}um", kullanilan_sayfa_adlari
            ),
        }

    return veri_setleri


def veri_yukle(dosya):
    kullanilan_sayfa_adlari = set()
    veri_setleri = {}

    veri_setleri.update(havuz_verilerini_hazirla(dosya, kullanilan_sayfa_adlari))
    veri_setleri.update(gelisler_arasi_hazirla(dosya, kullanilan_sayfa_adlari))
    veri_setleri.update(eloksal_uretim_hazirla(dosya, kullanilan_sayfa_adlari))

    return veri_setleri


# =============================================================================
# ISTATISTIKSEL TESTLER
# =============================================================================


def yuvarla_veya_metin(deger, basamak=4, bos_metin="Tanımsız"):
    if deger is None:
        return bos_metin
    try:
        sayi = float(deger)
    except Exception:
        return bos_metin
    if not np.isfinite(sayi):
        return bos_metin
    return round(sayi, basamak)


def dagilim_statusu(ad):
    if ad in ADAY_DAGILIMLAR:
        return "Candidate"
    if ad in DIAGNOSTIC_DAGILIMLAR:
        return "Diagnostic"
    return "Review"


def skor_agirliklari_sec(n):
    if n < 30:
        return "kucuk", SKOR_KOVALARI["kucuk"]
    if n < 100:
        return "orta", SKOR_KOVALARI["orta"]
    return "buyuk", SKOR_KOVALARI["buyuk"]


def aic_bic(dagilim, parametreler, veri):
    try:
        ll = np.sum(dagilim.logpdf(veri, *parametreler))
        if not np.isfinite(ll):
            return None, None
        k = len(parametreler)
        n = len(veri)
        aic = 2 * k - 2 * ll
        bic = k * np.log(n) - 2 * ll
        if not np.isfinite(aic) or not np.isfinite(bic):
            return None, None
        return round(aic, 2), round(bic, 2)
    except Exception:
        return None, None


def aicc_hesapla(aic, parametre_sayisi, n):
    try:
        if aic is None:
            return None
        payda = n - parametre_sayisi - 1
        if payda <= 0:
            return None
        aicc = float(aic) + (2 * parametre_sayisi * (parametre_sayisi + 1)) / payda
        if not np.isfinite(aicc):
            return None
        return round(aicc, 2)
    except Exception:
        return None


def olasilik_izgarasi(n):
    eps = 1e-6
    probs = (np.arange(1, n + 1) - 0.5) / n
    return np.clip(probs, eps, 1 - eps)


def dagilim_cdf_degerleri(veri, dagilim, params):
    sirali = np.sort(np.asarray(veri, dtype=float))
    u = dagilim.cdf(sirali, *params)
    u = np.asarray(u, dtype=float)
    u = np.clip(u, 1e-12, 1 - 1e-12)
    if not np.all(np.isfinite(u)):
        raise ValueError("CDF degerleri sonlu degil.")
    return sirali, u


def ks_istatistigi(veri, dagilim, params):
    sirali, u = dagilim_cdf_degerleri(veri, dagilim, params)
    n = len(sirali)
    i = np.arange(1, n + 1)
    d_plus = np.max(i / n - u)
    d_minus = np.max(u - (i - 1) / n)
    return float(max(d_plus, d_minus))


def ad_istatistigi(veri, dagilim, params):
    sirali, u = dagilim_cdf_degerleri(veri, dagilim, params)
    n = len(sirali)
    i = np.arange(1, n + 1)
    toplam = np.sum((2 * i - 1) * (np.log(u) + np.log(1 - u[::-1])))
    return float(-n - toplam / n)


def cvm_istatistigi(veri, dagilim, params):
    sirali, u = dagilim_cdf_degerleri(veri, dagilim, params)
    n = len(sirali)
    i = np.arange(1, n + 1)
    return float(1.0 / (12 * n) + np.sum((u - (2 * i - 1) / (2 * n)) ** 2))


def teorik_kantiller(dagilim, params, n):
    q = dagilim.ppf(olasilik_izgarasi(n), *params)
    q = np.asarray(q, dtype=float)
    if not np.all(np.isfinite(q)):
        raise ValueError("PPF degerleri sonlu degil.")
    return q


def quantile_wasserstein(veri, dagilim, params):
    sirali = np.sort(np.asarray(veri, dtype=float))
    teorik = teorik_kantiller(dagilim, params, len(sirali))
    return float(np.mean(np.abs(sirali - teorik)))


def quantile_rmse(veri, dagilim, params):
    sirali = np.sort(np.asarray(veri, dtype=float))
    teorik = teorik_kantiller(dagilim, params, len(sirali))
    return float(np.sqrt(np.mean((sirali - teorik) ** 2)))


def gof_istatistikleri(veri, dagilim, params):
    return {
        "KS": ks_istatistigi(veri, dagilim, params),
        "AD": ad_istatistigi(veri, dagilim, params),
        "CvM": cvm_istatistigi(veri, dagilim, params),
    }


def bootstrap_gof(
    dagilim, veri, fit_params, iterasyon=BOOTSTRAP_ITER, alpha=ALPHA, seed=RNG_SEED
):

    gozlenen = gof_istatistikleri(veri, dagilim, fit_params)
    boot_degerleri = {ad: [] for ad in gozlenen}
    hedef = max(int(iterasyon), 20)
    max_deneme = hedef * BOOTSTRAP_ATTEMPT_KATSAYI
    rng = np.random.default_rng(seed)

    deneme = 0
    while len(boot_degerleri["KS"]) < hedef and deneme < max_deneme:
        deneme += 1
        try:
            rastgele_durum = int(rng.integers(0, 2**32 - 1))
            sentetik = dagilim.rvs(
                *fit_params, size=len(veri), random_state=rastgele_durum
            )
            sentetik = np.asarray(sentetik, dtype=float)
            if len(sentetik) != len(veri) or not np.all(np.isfinite(sentetik)):
                continue
            yeniden_params = dagilim.fit(sentetik)
            sentetik_istat = gof_istatistikleri(sentetik, dagilim, yeniden_params)
            for anahtar in boot_degerleri:
                if np.isfinite(sentetik_istat[anahtar]):
                    boot_degerleri[anahtar].append(float(sentetik_istat[anahtar]))
        except Exception:
            continue

    sonuc = {"_bootstrap_n": len(boot_degerleri["KS"])}
    yeterli = sonuc["_bootstrap_n"] >= max(20, hedef // 2)

    for ad in ("KS", "AD", "CvM"):
        stat = gozlenen[ad]
        sonuc[f"{ad} İstatistik"] = yuvarla_veya_metin(stat, 4)
        sonuc[f"_{ad}_boot_stat_num"] = stat

        if yeterli:
            boot = np.asarray(boot_degerleri[ad], dtype=float)
            p_deg = (np.sum(boot >= stat) + 1.0) / (len(boot) + 1.0)
            sonuc[f"{ad} Bootstrap p-değeri"] = yuvarla_veya_metin(p_deg, 4)
            sonuc[f"{ad} Karar"] = "Kabul" if p_deg >= alpha else "Red"
            sonuc[f"_{ad}_boot_p_num"] = float(p_deg)
        else:
            sonuc[f"{ad} Bootstrap p-değeri"] = "Hesaplanamadı"
            sonuc[f"{ad} Karar"] = "Hesaplanamadı"
            sonuc[f"_{ad}_boot_p_num"] = None

    return sonuc


def test_set_performansi(dagilim, params_egitim, test_veri, alpha=ALPHA):
    sonuc = {}
    test_veri = np.asarray(test_veri, dtype=float)

    if len(test_veri) < 2:
        for anahtar in [
            "TT KS İstatistik",
            "TT KS p-değeri",
            "TT KS Karar",
            "TT LogLik",
            "TT Ortalama NLL",
            "TT Wasserstein",
            "TT CvM",
            "TT Quantile Hata",
        ]:
            sonuc[anahtar] = "Yetersiz veri"
        sonuc["_TT_LogLik_num"] = None
        sonuc["_TT_NLL_num"] = None
        sonuc["_TT_Wasserstein_num"] = None
        sonuc["_TT_CvM_num"] = None
        sonuc["_TT_Quantile_Hata_num"] = None
        sonuc["_TT_KS_stat_num"] = None
        sonuc["_TT_KS_p_num"] = None
        return sonuc

    try:
        tt_ks_stat = ks_istatistigi(test_veri, dagilim, params_egitim)
        _, tt_ks_p = kstest(test_veri, lambda x: dagilim.cdf(x, *params_egitim))
        sonuc["_TT_KS_stat_num"] = float(tt_ks_stat)
        sonuc["_TT_KS_p_num"] = float(tt_ks_p)
        sonuc["TT KS İstatistik"] = yuvarla_veya_metin(tt_ks_stat, 4)
        sonuc["TT KS p-değeri"] = yuvarla_veya_metin(tt_ks_p, 4)
        sonuc["TT KS Karar"] = "Kabul" if tt_ks_p >= alpha else "Red"
    except Exception:
        sonuc["_TT_KS_stat_num"] = None
        sonuc["_TT_KS_p_num"] = None
        sonuc["TT KS İstatistik"] = "Hesaplanamadı"
        sonuc["TT KS p-değeri"] = "Hesaplanamadı"
        sonuc["TT KS Karar"] = "Hesaplanamadı"

    try:
        logpdf = np.asarray(dagilim.logpdf(test_veri, *params_egitim), dtype=float)
        if not np.all(np.isfinite(logpdf)):
            raise ValueError("Log-likelihood sonlu degil.")
        loglik = float(np.sum(logpdf))
        ort_nll = float(-np.mean(logpdf))
        sonuc["_TT_LogLik_num"] = loglik
        sonuc["_TT_NLL_num"] = ort_nll
        sonuc["TT LogLik"] = yuvarla_veya_metin(loglik, 2)
        sonuc["TT Ortalama NLL"] = yuvarla_veya_metin(ort_nll, 4)
    except Exception:
        sonuc["_TT_LogLik_num"] = None
        sonuc["_TT_NLL_num"] = None
        sonuc["TT LogLik"] = "Tanımsız"
        sonuc["TT Ortalama NLL"] = "Tanımsız"

    try:
        tt_w = quantile_wasserstein(test_veri, dagilim, params_egitim)
        sonuc["_TT_Wasserstein_num"] = tt_w
        sonuc["TT Wasserstein"] = yuvarla_veya_metin(tt_w, 4)
    except Exception:
        sonuc["_TT_Wasserstein_num"] = None
        sonuc["TT Wasserstein"] = "Tanımsız"

    try:
        tt_cvm = cvm_istatistigi(test_veri, dagilim, params_egitim)
        sonuc["_TT_CvM_num"] = tt_cvm
        sonuc["TT CvM"] = yuvarla_veya_metin(tt_cvm, 4)
    except Exception:
        sonuc["_TT_CvM_num"] = None
        sonuc["TT CvM"] = "Tanımsız"

    try:
        tt_q = quantile_rmse(test_veri, dagilim, params_egitim)
        sonuc["_TT_Quantile_Hata_num"] = tt_q
        sonuc["TT Quantile Hata"] = yuvarla_veya_metin(tt_q, 4)
    except Exception:
        sonuc["_TT_Quantile_Hata_num"] = None
        sonuc["TT Quantile Hata"] = "Tanımsız"

    return sonuc


def sonuclari_skorla(sonuclar):
    if not sonuclar:
        return

    n = sonuclar[0].get("_n", 0)
    kova_adi, agirliklar = skor_agirliklari_sec(n)
    aday_indeksleri = [
        i for i, s in enumerate(sonuclar) if s.get("Statü") == "Candidate"
    ]
    yonler = {
        "_AICc_num": True,
        "_BIC_num": True,
        "_KS_boot_stat_num": True,
        "_AD_boot_stat_num": True,
        "_CvM_boot_stat_num": True,
        "_KS_boot_p_num": False,
        "_AD_boot_p_num": False,
        "_CvM_boot_p_num": False,
        "_Wasserstein_num": True,
        "_TT_Wasserstein_num": True,
        "_TT_CvM_num": True,
        "_TT_Quantile_Hata_num": True,
        "_TT_KS_stat_num": True,
        "_TT_KS_p_num": False,
        "_TT_NLL_num": True,
    }

    ham_toplam = {i: 0.0 for i in aday_indeksleri}
    agirlikli_toplam = {i: 0.0 for i in aday_indeksleri}

    for alan, agirlik in agirliklar.items():
        aday_degerleri = {i: sonuclar[i].get(alan) for i in aday_indeksleri}
        gecerli = {
            i: v for i, v in aday_degerleri.items() if v is not None and np.isfinite(v)
        }
        if not aday_indeksleri:
            continue
        if gecerli:
            seri = pd.Series(gecerli, dtype=float).rank(
                method="average", ascending=yonler[alan]
            )
            ceza = float(seri.max()) + 1.0
        else:
            seri = pd.Series(dtype=float)
            ceza = float(len(aday_indeksleri) + 1)

        for idx in aday_indeksleri:
            sira = float(seri.get(idx, ceza))
            ham_toplam[idx] += sira
            agirlikli_toplam[idx] += sira * agirlik

    for idx, sonuc in enumerate(sonuclar):
        sonuc["Skor Kovası"] = kova_adi
        if idx in aday_indeksleri:
            ham = ham_toplam[idx]
            agirlikli = agirlikli_toplam[idx]
            sonuc["_ham_skor_num"] = float(ham)
            sonuc["Ham Skor"] = round(float(ham), 2)
            sonuc["_score_num"] = float(agirlikli)
            sonuc["Ağırlıklı Skor"] = round(float(agirlikli), 4)
        else:
            sonuc["_ham_skor_num"] = None
            sonuc["Ham Skor"] = "Skor Dışı"
            sonuc["_score_num"] = None
            sonuc["Ağırlıklı Skor"] = "Skor Dışı"


def guven_seviyesi_hesapla(sonuc):
    izlenen = [
        "_AICc_num",
        "_BIC_num",
        "_KS_boot_stat_num",
        "_KS_boot_p_num",
        "_AD_boot_stat_num",
        "_AD_boot_p_num",
        "_CvM_boot_stat_num",
        "_CvM_boot_p_num",
        "_TT_KS_stat_num",
        "_TT_KS_p_num",
        "_TT_NLL_num",
        "_TT_Wasserstein_num",
        "_TT_CvM_num",
        "_TT_Quantile_Hata_num",
    ]
    gecerli_sayi = sum(
        1
        for alan in izlenen
        if sonuc.get(alan) is not None and np.isfinite(sonuc.get(alan))
    )
    oran = gecerli_sayi / len(izlenen) if izlenen else 0.0
    kabul_sayisi = sum(
        1
        for alan in ("KS Karar", "AD Karar", "CvM Karar")
        if sonuc.get(alan) == "Kabul"
    )

    sonuc["_valid_metric_ratio"] = round(oran, 4)
    sonuc["_gof_accept_count"] = kabul_sayisi

    if oran >= 0.85 and kabul_sayisi >= 2:
        sonuc["Güven Seviyesi"] = "Yüksek"
    elif oran >= 0.6 and kabul_sayisi >= 1:
        sonuc["Güven Seviyesi"] = "Orta"
    else:
        sonuc["Güven Seviyesi"] = "Düşük"


def tum_testler(ad, dagilim, veri, alpha=ALPHA):
    sonuc = {
        "Dağılım": ad,
        "_basarili": False,
        "_n": len(veri),
        "Statü": dagilim_statusu(ad),
    }

    if len(veri) < MIN_GOZLEM:
        sonuc["_hata"] = f"En az {MIN_GOZLEM} gözlem gerekli."
        return sonuc

    try:
        params = dagilim.fit(veri)
        sonuc["_params"] = params

        aic, bic = aic_bic(dagilim, params, veri)
        sonuc["_AIC_num"] = aic
        sonuc["_BIC_num"] = bic
        sonuc["_AICc_num"] = aicc_hesapla(aic, len(params), len(veri))
        if aic is None or bic is None:
            sonuc["AIC"] = "Tanımsız"
            sonuc["BIC"] = "Tanımsız"
        else:
            sonuc["AIC"] = aic
            sonuc["BIC"] = bic
        sonuc["AICc"] = (
            sonuc["_AICc_num"] if sonuc.get("_AICc_num") is not None else "Tanımsız"
        )

        sonuc["_Wasserstein_num"] = quantile_wasserstein(veri, dagilim, params)
        sonuc["Wasserstein"] = yuvarla_veya_metin(sonuc["_Wasserstein_num"], 4)

        sonuc.update(bootstrap_gof(dagilim, veri, params, alpha=alpha, seed=RNG_SEED))

        rng = np.random.default_rng(RNG_SEED)
        idx = rng.permutation(len(veri))
        bolum = max(int(len(veri) * TRAIN_ORAN), 1)
        egitim = veri[idx[:bolum]]
        test = veri[idx[bolum:]]

        if len(test) >= 1 and len(egitim) >= MIN_GOZLEM:
            params_eg = dagilim.fit(egitim)
            sonuc.update(test_set_performansi(dagilim, params_eg, test, alpha=alpha))
        else:
            sonuc.update(test_set_performansi(dagilim, (), np.array([]), alpha=alpha))

        guven_seviyesi_hesapla(sonuc)
        sonuc["_basarili"] = True

    except Exception as hata:
        sonuc["_hata"] = str(hata)

    return sonuc


# =============================================================================
# GORSELLESTIRME
# =============================================================================


def histogram_bin_ayari(n):
    if n < 30:
        return min(8, max(5, int(np.ceil(np.sqrt(n)))))
    if n < 80:
        return min(12, max(8, int(np.ceil(np.sqrt(n)))))
    return "fd"


def gorsel_olustur(dagilim_adi, dagilim, params, veri, etiket, birim, cikti_yolu):
    fig = plt.figure(figsize=(16, 5), facecolor="white")
    fig.suptitle(
        f"{etiket}  ·  {dagilim_adi} Dağılımı",
        fontsize=12,
        fontweight="bold",
        y=1.01,
    )
    gs = gridspec.GridSpec(1, 3, figure=fig, wspace=0.38)

    n = len(veri)
    sirali = np.sort(veri)

    ax1 = fig.add_subplot(gs[0, 0])
    teorik_q = teorik_kantiller(dagilim, params, n)
    ax1.scatter(teorik_q, sirali, s=20, color="#2E75B6", alpha=0.7, edgecolors="none")
    mn = min(teorik_q.min(), sirali.min())
    mx = max(teorik_q.max(), sirali.max())
    ax1.plot([mn, mx], [mn, mx], "r--", lw=1.5, label="Referans (y=x)")
    ax1.set_xlabel(f"Teorik Kantil ({birim})", fontsize=9)
    ax1.set_ylabel(f"Gözlenen Kantil ({birim})", fontsize=9)
    ax1.set_title("QQ Grafiği", fontsize=10, fontweight="bold")
    ax1.legend(fontsize=8)
    ax1.grid(True, alpha=0.3, linestyle="--")

    ax2 = fig.add_subplot(gs[0, 1])
    gozlenen_p = np.arange(1, n + 1) / (n + 1)
    teorik_p = dagilim.cdf(sirali, *params)
    ax2.scatter(
        teorik_p, gozlenen_p, s=20, color="#70AD47", alpha=0.7, edgecolors="none"
    )
    ax2.plot([0, 1], [0, 1], "r--", lw=1.5, label="Referans (y=x)")
    ax2.set_xlabel("Teorik Olasılık", fontsize=9)
    ax2.set_ylabel("Gözlenen Olasılık", fontsize=9)
    ax2.set_title("PP Grafiği", fontsize=10, fontweight="bold")
    ax2.legend(fontsize=8)
    ax2.grid(True, alpha=0.3, linestyle="--")

    ax3 = fig.add_subplot(gs[0, 2])
    ax3.hist(
        veri,
        bins=histogram_bin_ayari(n),
        density=True,
        color="#2E75B6",
        alpha=0.5,
        edgecolor="white",
        linewidth=0.5,
        label="Gözlem",
    )
    x_min = max(veri.min() - 0.05 * veri.std(), 0)
    x_max = veri.max() + 0.05 * veri.std()
    xs = np.linspace(x_min, x_max, 400)
    try:
        pdf_deger = dagilim.pdf(xs, *params)
        ax3.plot(xs, pdf_deger, "r-", lw=2, label="PDF")
    except Exception:
        pass
    ax3.set_xlabel(birim, fontsize=9)
    ax3.set_ylabel("Yoğunluk", fontsize=9)
    ax3.set_title("Histogram + PDF", fontsize=10, fontweight="bold")
    ax3.legend(fontsize=8)
    ax3.grid(True, alpha=0.3, linestyle="--")

    plt.tight_layout()
    plt.savefig(cikti_yolu, dpi=130, bbox_inches="tight")
    plt.close(fig)


# =============================================================================
# EXCEL CIKTI
# =============================================================================

R_BASLIK = "1F3864"
R_ALTBASLIK = "2E75B6"
R_EN_IYI = "E2EFDA"
R_CIFT = "DDEBF7"
R_BEYAZ = "FFFFFF"
R_YESIL_YZ = "375623"
R_KIRMIZI = "C00000"

INCE = Side(border_style="thin", color="BFBFBF")
KENAR = Border(left=INCE, right=INCE, top=INCE, bottom=INCE)


def stil_baslik(ws, satir, s_sutun, e_sutun, metin, bg=R_BASLIK, boyut=12):
    ws.merge_cells(
        start_row=satir,
        start_column=s_sutun,
        end_row=satir,
        end_column=e_sutun,
    )
    hucre = ws.cell(row=satir, column=s_sutun, value=metin)
    hucre.font = Font(name="Arial", bold=True, size=boyut, color="FFFFFF")
    hucre.fill = PatternFill("solid", fgColor=bg)
    hucre.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    hucre.border = KENAR
    ws.row_dimensions[satir].height = 24


def stil_sutun_basligi(ws, satir, sutunlar, metinler):
    for s, m in zip(sutunlar, metinler):
        hucre = ws.cell(row=satir, column=s, value=m)
        hucre.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        hucre.fill = PatternFill("solid", fgColor=R_ALTBASLIK)
        hucre.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        hucre.border = KENAR
    ws.row_dimensions[satir].height = 32


def doldur(
    ws,
    satir,
    sutun,
    deger,
    bg=None,
    kalin=False,
    yazi_rengi="000000",
    hizalama="center",
):
    hucre = ws.cell(row=satir, column=sutun, value=deger)
    hucre.font = Font(name="Arial", size=9, bold=kalin, color=yazi_rengi)
    hucre.alignment = Alignment(horizontal=hizalama, vertical="center")
    hucre.border = KENAR
    if bg:
        hucre.fill = PatternFill("solid", fgColor=bg)


def excel_sayfasi_olustur(wb, sayfa_adi, etiket, birim, sonuclar, gorsel_klasor, n):
    ws = wb.create_sheet(title=sayfa_adi[:31])

    sutunlar = [
        "Dağılım",
        "Statü",
        "Güven Seviyesi",
        "Skor Kovası",
        "AIC",
        "AICc",
        "BIC",
        "Wasserstein",
        "KS İstatistik",
        "KS Bootstrap p-değeri",
        "KS Karar",
        "AD İstatistik",
        "AD Bootstrap p-değeri",
        "AD Karar",
        "CvM İstatistik",
        "CvM Bootstrap p-değeri",
        "CvM Karar",
        "TT KS İstatistik",
        "TT KS p-değeri",
        "TT KS Karar",
        "TT LogLik",
        "TT Ortalama NLL",
        "TT Wasserstein",
        "TT CvM",
        "TT Quantile Hata",
        "Ham Skor",
        "Ağırlıklı Skor",
    ]
    genislikler = [
        18,
        12,
        14,
        12,
        11,
        11,
        11,
        13,
        14,
        16,
        10,
        14,
        16,
        12,
        14,
        16,
        12,
        14,
        12,
        12,
        12,
        13,
        12,
        12,
        14,
        14,
    ]

    stil_baslik(ws, 1, 1, len(sutunlar), f"İstatistiksel Dağılım Analizi  ·  {etiket}")
    stil_baslik(
        ws,
        2,
        1,
        len(sutunlar),
        f"Birim: {birim}  |  n = {n}  |  α = {ALPHA}  |  Bootstrap = {BOOTSTRAP_ITER}",
        bg=R_ALTBASLIK,
        boyut=10,
    )
    stil_sutun_basligi(ws, 3, list(range(1, len(sutunlar) + 1)), sutunlar)

    for i, genislik in enumerate(genislikler, 1):
        ws.column_dimensions[get_column_letter(i)].width = genislik

    gecerli = [
        r
        for r in sonuclar
        if r.get("_basarili")
        and r.get("Statü") == "Candidate"
        and r.get("_score_num") is not None
    ]
    en_iyi = (
        min(gecerli, key=lambda r: r.get("_score_num", np.inf)) if gecerli else None
    )

    def karar_rengi(karar):
        if karar == "Kabul":
            return R_YESIL_YZ
        if karar == "Red":
            return R_KIRMIZI
        return "000000"

    for idx, res in enumerate(sonuclar, 4):
        is_best = (
            en_iyi and res["Dağılım"] == en_iyi["Dağılım"] and res.get("_basarili")
        )
        satir_bg = R_EN_IYI if is_best else (R_CIFT if idx % 2 == 0 else R_BEYAZ)

        doldur(
            ws,
            idx,
            1,
            res.get("Dağılım", "?"),
            satir_bg,
            kalin=is_best,
            hizalama="left",
        )
        doldur(ws, idx, 2, res.get("Statü", "—"), satir_bg)
        doldur(ws, idx, 3, res.get("Güven Seviyesi", "—"), satir_bg)
        doldur(ws, idx, 4, res.get("Skor Kovası", "—"), satir_bg)
        doldur(ws, idx, 5, res.get("AIC", "—"), satir_bg)
        doldur(ws, idx, 6, res.get("AICc", "—"), satir_bg)
        doldur(ws, idx, 7, res.get("BIC", "—"), satir_bg)
        doldur(ws, idx, 8, res.get("Wasserstein", "—"), satir_bg)
        doldur(ws, idx, 9, res.get("KS İstatistik", "—"), satir_bg)
        doldur(ws, idx, 10, res.get("KS Bootstrap p-değeri", "—"), satir_bg)
        ks_karar = res.get("KS Karar", "—")
        doldur(ws, idx, 11, ks_karar, satir_bg, yazi_rengi=karar_rengi(ks_karar))
        doldur(ws, idx, 12, res.get("AD İstatistik", "—"), satir_bg)
        doldur(ws, idx, 13, res.get("AD Bootstrap p-değeri", "—"), satir_bg)
        ad_karar = res.get("AD Karar", "—")
        doldur(ws, idx, 14, ad_karar, satir_bg, yazi_rengi=karar_rengi(ad_karar))
        doldur(ws, idx, 15, res.get("CvM İstatistik", "—"), satir_bg)
        doldur(ws, idx, 16, res.get("CvM Bootstrap p-değeri", "—"), satir_bg)
        cvm_karar = res.get("CvM Karar", "—")
        doldur(ws, idx, 17, cvm_karar, satir_bg, yazi_rengi=karar_rengi(cvm_karar))
        doldur(ws, idx, 18, res.get("TT KS İstatistik", "—"), satir_bg)
        doldur(ws, idx, 19, res.get("TT KS p-değeri", "—"), satir_bg)
        tt_karar = res.get("TT KS Karar", "—")
        doldur(ws, idx, 20, tt_karar, satir_bg, yazi_rengi=karar_rengi(tt_karar))
        doldur(ws, idx, 21, res.get("TT LogLik", "—"), satir_bg)
        doldur(ws, idx, 22, res.get("TT Ortalama NLL", "—"), satir_bg)
        doldur(ws, idx, 23, res.get("TT Wasserstein", "—"), satir_bg)
        doldur(ws, idx, 24, res.get("TT CvM", "—"), satir_bg)
        doldur(ws, idx, 25, res.get("TT Quantile Hata", "—"), satir_bg)
        doldur(ws, idx, 26, res.get("Ham Skor", "—"), satir_bg, kalin=is_best)
        doldur(ws, idx, 27, res.get("Ağırlıklı Skor", "—"), satir_bg, kalin=is_best)

    alt_satir = 3 + len(sonuclar) + 1

    if en_iyi:
        stil_baslik(
            ws,
            alt_satir,
            1,
            len(sutunlar),
            f"✓  En İyi Dağılım (Çok Ölçütlü Skor): {en_iyi['Dağılım']}  "
            f"|  Ağırlıklı Skor = {en_iyi['Ağırlıklı Skor']}  |  Güven = {en_iyi.get('Güven Seviyesi', '—')}",
            bg="375623",
            boyut=11,
        )
        alt_satir += 1

    gorsel_dosyalari = sorted(
        [f for f in os.listdir(gorsel_klasor) if f.endswith(".png")]
    )
    if gorsel_dosyalari:
        alt_satir += 1
        stil_baslik(
            ws,
            alt_satir,
            1,
            len(sutunlar),
            "Görsel Analizler  (QQ · PP · Histogram + PDF)",
            bg=R_ALTBASLIK,
            boyut=11,
        )
        alt_satir += 1

        for dosya in gorsel_dosyalari:
            tam_yol = os.path.join(gorsel_klasor, dosya)
            gorsel = XLImage(tam_yol)
            gorsel.width = 920
            gorsel.height = 290
            ws.add_image(gorsel, f"A{alt_satir}")
            ws.row_dimensions[alt_satir].height = 218
            alt_satir += 16


def ozet_sayfasi(wb, veri_setleri, tum_sonuclar):
    ws = wb.create_sheet(title="ÖZET", index=0)

    sutunlar = [
        "Kategori",
        "Veri Seti",
        "Grup",
        "n",
        "En İyi Dağılım",
        "Statü",
        "Güven Seviyesi",
        "Skor Kovası",
        "Ağırlıklı Skor",
        "Ham Skor",
        "AICc",
        "BIC",
        "KS Kararı",
        "AD Kararı",
        "CvM Kararı",
        "TT KS Kararı",
        "TT Ortalama NLL",
    ]
    genislikler = [18, 34, 18, 8, 20, 12, 14, 12, 14, 12, 12, 12, 12, 12, 14, 14]

    stil_baslik(
        ws,
        1,
        1,
        len(sutunlar),
        "Dağılım Uyum Analizi – Genel Özet",
        bg=R_BASLIK,
        boyut=13,
    )
    stil_sutun_basligi(ws, 2, list(range(1, len(sutunlar) + 1)), sutunlar)
    for i, genislik in enumerate(genislikler, 1):
        ws.column_dimensions[get_column_letter(i)].width = genislik

    satir = 3

    def karar_rengi(karar):
        if karar == "Kabul":
            return R_YESIL_YZ
        if karar == "Red":
            return R_KIRMIZI
        return "000000"

    for key, bilgi in veri_setleri.items():
        sonuclar = tum_sonuclar[key]
        gecerli = [
            r
            for r in sonuclar
            if r.get("_basarili")
            and r.get("Statü") == "Candidate"
            and r.get("_score_num") is not None
        ]
        en_iyi = (
            min(gecerli, key=lambda r: r.get("_score_num", np.inf)) if gecerli else {}
        )
        bg = R_EN_IYI if satir % 2 == 0 else R_CIFT

        doldur(ws, satir, 1, bilgi["kategori"], bg)
        doldur(ws, satir, 2, bilgi["etiket"], bg, hizalama="left")
        doldur(ws, satir, 3, bilgi.get("grup", "—"), bg)
        doldur(ws, satir, 4, len(bilgi["veri"]), bg)
        doldur(ws, satir, 5, en_iyi.get("Dağılım", "—"), bg, kalin=True)
        doldur(ws, satir, 6, en_iyi.get("Statü", "—"), bg)
        doldur(ws, satir, 7, en_iyi.get("Güven Seviyesi", "—"), bg)
        doldur(ws, satir, 8, en_iyi.get("Skor Kovası", "—"), bg)
        doldur(ws, satir, 9, en_iyi.get("Ağırlıklı Skor", "—"), bg)
        doldur(ws, satir, 10, en_iyi.get("Ham Skor", "—"), bg)
        doldur(ws, satir, 11, en_iyi.get("AICc", "—"), bg)
        doldur(ws, satir, 12, en_iyi.get("BIC", "—"), bg)
        ks = en_iyi.get("KS Karar", "—")
        doldur(ws, satir, 13, ks, bg, yazi_rengi=karar_rengi(ks))
        ad = en_iyi.get("AD Karar", "—")
        doldur(ws, satir, 14, ad, bg, yazi_rengi=karar_rengi(ad))
        cvm = en_iyi.get("CvM Karar", "—")
        doldur(ws, satir, 15, cvm, bg, yazi_rengi=karar_rengi(cvm))
        tt = en_iyi.get("TT KS Karar", "—")
        doldur(ws, satir, 16, tt, bg, yazi_rengi=karar_rengi(tt))
        doldur(ws, satir, 17, en_iyi.get("TT Ortalama NLL", "—"), bg)
        satir += 1


# =============================================================================
# ANA FONKSIYON
# =============================================================================


def main():
    print("=" * 70)
    print("  Dağılım Uyum Analizi Başlatılıyor")
    print("=" * 70)

    os.makedirs(CIKTI_KLASOR, exist_ok=True)

    print(f"\n[1/4] Veri yükleniyor: {KAYNAK_DOSYA}")
    veri_setleri = veri_yukle(KAYNAK_DOSYA)
    for key, bilgi in veri_setleri.items():
        print(f"  - {key:<30} {bilgi['etiket']}  |  n = {len(bilgi['veri'])}")

    wb = Workbook()
    wb.remove(wb.active)
    tum_sonuclar = {}

    print("\n[2/4] Testler ve görseller oluşturuluyor...")
    toplam = len(veri_setleri)
    for sira, (key, bilgi) in enumerate(veri_setleri.items(), 1):
        veri = bilgi["veri"]
        etiket = bilgi["etiket"]
        birim = bilgi["birim"]
        klasor = os.path.join(
            CIKTI_KLASOR, bilgi["kategori"], slugify(bilgi.get("grup", key))
        )
        os.makedirs(klasor, exist_ok=True)

        print(f"\n  ({sira}/{toplam}) {etiket}  [n={len(veri)}]")
        sonuclar = []
        for dagilim_adi, dagilim in DAGITIMLAR.items():
            sonuc = tum_testler(dagilim_adi, dagilim, veri)
            sonuclar.append(sonuc)

            if sonuc["_basarili"]:
                gorsel_yol = os.path.join(
                    klasor, f"{dagilim_adi.lower().replace(' ', '_')}.png"
                )
                try:
                    gorsel_olustur(
                        dagilim_adi,
                        dagilim,
                        sonuc["_params"],
                        veri,
                        etiket,
                        birim,
                        gorsel_yol,
                    )
                    aic_text = (
                        f"{sonuc['_AIC_num']:.2f}"
                        if sonuc.get("_AIC_num") is not None
                        else str(sonuc["AIC"])
                    )
                    ks_boot_p = sonuc.get("KS Bootstrap p-değeri", "—")
                    durum = (
                        f"{sonuc.get('Statü', '—'):<10} "
                        f"AIC={aic_text:>10}  KSb-p={str(ks_boot_p):>8}"
                    )
                except Exception as hata:
                    durum = f"Görsel hata: {hata}"
            else:
                durum = f"HATA: {sonuc.get('_hata', '?')}"

            print(f"    {'✓' if sonuc['_basarili'] else '✗'} {dagilim_adi:<14} {durum}")

        sonuclari_skorla(sonuclar)
        tum_sonuclar[key] = sonuclar
        gecerli = [
            r
            for r in sonuclar
            if r.get("Statü") == "Candidate" and r.get("_score_num") is not None
        ]
        if gecerli:
            en_iyi = min(gecerli, key=lambda r: r["_score_num"])
            print(
                f"    -> En iyi (adaylar icinde): {en_iyi['Dağılım']}  "
                f"Agirlikli={en_iyi['Ağırlıklı Skor']}  Guven={en_iyi.get('Güven Seviyesi', '—')}  "
                f"KS={en_iyi.get('KS Karar', '—')}  AD={en_iyi.get('AD Karar', '—')}  CvM={en_iyi.get('CvM Karar', '—')}"
            )
        excel_sayfasi_olustur(
            wb,
            bilgi["sayfa_adi"],
            etiket,
            birim,
            sonuclar,
            klasor,
            len(veri),
        )

    print("\n[3/4] Özet sayfası oluşturuluyor...")
    ozet_sayfasi(wb, veri_setleri, tum_sonuclar)

    print("[4/4] Excel kaydediliyor...")
    cikti_dosya = os.path.join(CIKTI_KLASOR, "dagilim_analiz_sonuclari.xlsx")
    wb.save(cikti_dosya)

    print("\n" + "=" * 70)
    print(f"  Excel çıktı : {cikti_dosya}")
    print(f"  Görseller   : {CIKTI_KLASOR}/<kategori>/<grup>/*.png")
    print("=" * 70)


if __name__ == "__main__":
    main()