"""
KULLANIM ORNEKLERI
------------------
100 bara girisi yap ve tum sistem bosalana kadar calistir:
    .\\.venv312\\Scripts\\python.exe .\\anodizing_des.py --arrival-limit 100

8 saatlik simulasyon yap:
    .\\.venv312\\Scripts\\python.exe .\\anodizing_des.py --hours 8

1 gunluk simulasyon yap:
    .\\.venv312\\Scripts\\python.exe .\\anodizing_des.py --days 1

Hem bara limiti hem de sure koy:
    .\\.venv312\\Scripts\\python.exe .\\anodizing_des.py --arrival-limit 200 --hours 8

Ozel output dosyasi ver:
    .\\.venv312\\Scripts\\python.exe .\\anodizing_des.py --arrival-limit 50 --output-workbook .\\outputs\\rapor.xlsx

NOTLAR
------
- Temel zaman birimi saniyedir.
- --arrival-limit verilirse o kadar bara sisteme girer.
- --arrival-limit verilmezse giris limiti uygulanmaz.
- --until-seconds, --hours veya --days verilirse simulasyon o ana kadar kosar.
- Sure parametreleri birlikte verilirse toplam sure olarak toplanir.
"""

from __future__ import annotations

import argparse
import math
import re
import unicodedata
import warnings
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

import numpy as np
import pandas as pd
import simpy
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from scipy import stats


VALID_MICRONS = {7, 13, 25}
ACTIVE_PRIORITY_GROUPS = {"asitmat", "kostik", "notralizasyon", "eloksal", "tespit"}
RINSE_PROCESS_GROUPS = {
    "durulama_y",
    "durulama_a",
    "durulama_kostik",
    "durulama_1",
    "durulama_2",
    "durulama_post_eloksal",
    "durulama_post_renk",
    "durulama_post_tespit",
}
RINSE_STEP_KEYS = {
    "durulama_y",
    "durulama_a",
    "durulama_kostik",
    "pre_neutral_rinse",
    "post_eloksal_rinse",
    "post_color_rinse",
    "post_tespit_rinse",
}
PASSIVE_FORWARD_STEP_KEYS = set(RINSE_STEP_KEYS) | {"di"}
HAVUZ_SURELER_COLUMNS = [
    "Bara No.",
    "Çeşit",
    "Mikron",
    "YağAlma",
    "Durulama",
    "AsitMat",
    "Durulama.1",
    "Sökme",
    "Kostik",
    "Durulama V1",
    "Durulama V2",
    "Durulama V3",
    "Nötralizasyon",
    "Eloksal",
    "Durulama.2",
    "Kalay / Nikel (Renk)",
    "Durulama.3",
    "DI",
    "Tespit",
    "Durulama.4",
    "SıcakSu",
    "Süzme",
    "Fırın",
    "Başlangıç",
    "Bitiş",
    "Toplam Çevrim (sn)",
    "Açıklama",
]
ROUTE_LABELS = {
    "parlak": "Parlak",
    "sakem": "SAKEM",
    "pan": "PAN50/30",
    "natural": "Natural",
    "man": "MAN70/50/30-IRAK",
    "asitmat": "AsitMat",
}
SCIPY_DISTRIBUTIONS = {
    "Normal": stats.norm,
    "Lognormal": stats.lognorm,
    "Exponential": stats.expon,
    "Gamma": stats.gamma,
    "Weibull": stats.weibull_min,
    "LogLogistic": stats.fisk,
    "Cauchy": stats.cauchy,
    "Laplace": stats.laplace,
}


class DataContractError(RuntimeError):
    pass


def normalize_token(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = (
        text.replace("ı", "i")
        .replace("İ", "i")
        .replace("ş", "s")
        .replace("ğ", "g")
        .replace("ü", "u")
        .replace("ö", "o")
        .replace("ç", "c")
    )
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"[^a-z0-9]+", "", text)


def nearest_valid_micron(value: Any) -> int | None:
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return None
    if math.isnan(numeric):
        return None
    return min(sorted(VALID_MICRONS), key=lambda micron: (abs(float(micron) - numeric), -micron))


def to_seconds(value: Any) -> float | None:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return None
    if isinstance(value, pd.Timedelta):
        return float(value.total_seconds())
    if isinstance(value, pd.Timestamp):
        return float(
            value.hour * 3600 + value.minute * 60 + value.second + value.microsecond / 1_000_000
        )
    if hasattr(value, "hour") and hasattr(value, "minute") and hasattr(value, "second"):
        return float(value.hour * 3600 + value.minute * 60 + value.second)
    if isinstance(value, (int, float, np.integer, np.floating)):
        numeric = float(value)
        if math.isnan(numeric):
            return None
        if abs(numeric) <= 10:
            return numeric * 24.0 * 3600.0
        return numeric

    text = str(value).strip()
    if not text:
        return None
    if ":" in text:
        parts = text.split(":")
        try:
            parts = [float(piece.replace(",", ".")) for piece in parts]
        except ValueError:
            return None
        if len(parts) == 3:
            hours, minutes, seconds = parts
            return hours * 3600 + minutes * 60 + seconds
        if len(parts) == 2:
            minutes, seconds = parts
            return minutes * 60 + seconds
    try:
        numeric = float(text.replace(",", "."))
    except ValueError:
        return None
    if abs(numeric) <= 10:
        return numeric * 24.0 * 3600.0
    return numeric


def parse_param_string(text: str) -> tuple[list[float], float | None, float | None]:
    shape_params: list[float] = []
    loc = None
    scale = None
    for piece in str(text).split("|"):
        piece = piece.strip()
        if not piece or "=" not in piece:
            continue
        name, raw_value = piece.split("=", 1)
        name = normalize_token(name)
        try:
            value = float(raw_value.strip().replace(",", "."))
        except ValueError:
            continue
        if name == "loc":
            loc = value
        elif name == "scale":
            scale = value
        else:
            shape_params.append(value)
    return shape_params, loc, scale


def ensure_positive_seconds(value_seconds: float, label: str) -> float:
    if not math.isfinite(value_seconds) or value_seconds <= 0:
        raise DataContractError(f"{label} icin gecersiz sure uretiliyor: {value_seconds}")
    return value_seconds


def format_seconds_as_hms(seconds: float | None) -> str:
    if seconds is None or not math.isfinite(seconds):
        return ""
    total_seconds = max(0, int(round(seconds)))
    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    secs = total_seconds % 60
    return f"{hours:02d}:{minutes:02d}:{secs:02d}"


def route_family_label(route_family: str) -> str:
    return ROUTE_LABELS.get(route_family, route_family.title())


def crane_label(crane_id: str | None) -> str:
    if not crane_id:
        return ""
    match = re.search(r"(\d+)", crane_id)
    if match:
        return f"Vinç {match.group(1)}"
    return crane_id


def infer_valid_micron_from_seconds(duration_seconds: float | None) -> int | None:
    if duration_seconds is None or not math.isfinite(duration_seconds) or duration_seconds <= 0:
        return None
    target = min(VALID_MICRONS, key=lambda micron: abs(duration_seconds - micron * 60.0))
    return int(target)


@dataclass(slots=True)
class DistributionSpec:
    name: str
    params_text: str
    unit: str = "minutes"
    shapes: list[float] = field(init=False)
    loc: float | None = field(init=False)
    scale: float | None = field(init=False)
    raw_params: dict[str, float] = field(init=False, default_factory=dict)

    def __post_init__(self) -> None:
        if self.name == "Constant":
            self.shapes = []
            self.loc = None
            self.scale = None
            parsed: dict[str, float] = {}
            for piece in str(self.params_text).split("|"):
                piece = piece.strip()
                if not piece or "=" not in piece:
                    continue
                name, raw_value = piece.split("=", 1)
                try:
                    parsed[normalize_token(name)] = float(raw_value.strip().replace(",", "."))
                except ValueError:
                    continue
            required = {"value"}
            if not required.issubset(parsed):
                raise DataContractError(f"Constant dagilimi icin value gerekli: {self.params_text}")
            self.raw_params = parsed
            return

        if self.name == "Uniform":
            self.shapes = []
            self.loc = None
            self.scale = None
            parsed: dict[str, float] = {}
            for piece in str(self.params_text).split("|"):
                piece = piece.strip()
                if not piece or "=" not in piece:
                    continue
                name, raw_value = piece.split("=", 1)
                try:
                    parsed[normalize_token(name)] = float(raw_value.strip().replace(",", "."))
                except ValueError:
                    continue
            required = {"low", "high"}
            if not required.issubset(parsed):
                raise DataContractError(f"Uniform dagilimi icin low/high gerekli: {self.params_text}")
            self.raw_params = parsed
            return

        if self.name == "Triangular":
            self.shapes = []
            self.loc = None
            self.scale = None
            parsed: dict[str, float] = {}
            for piece in str(self.params_text).split("|"):
                piece = piece.strip()
                if not piece or "=" not in piece:
                    continue
                name, raw_value = piece.split("=", 1)
                try:
                    parsed[normalize_token(name)] = float(raw_value.strip().replace(",", "."))
                except ValueError:
                    continue
            required = {"min", "mode", "max"}
            if not required.issubset(parsed):
                raise DataContractError(f"Triangular dagilimi icin min/mode/max gerekli: {self.params_text}")
            self.raw_params = parsed
            return

        if self.name not in SCIPY_DISTRIBUTIONS:
            raise DataContractError(f"Desteklenmeyen dagilim: {self.name}")
        self.shapes, self.loc, self.scale = parse_param_string(self.params_text)

    @property
    def scipy_distribution(self):
        return SCIPY_DISTRIBUTIONS[self.name]

    def sample_seconds(self, rng: np.random.Generator, label: str, retries: int = 200) -> float:
        if self.name == "Constant":
            sample = self.raw_params["value"]
            if self.unit == "minutes":
                return ensure_positive_seconds(sample * 60.0, label)
            return ensure_positive_seconds(float(sample), label)

        if self.name == "Uniform":
            sample = rng.uniform(
                self.raw_params["low"],
                self.raw_params["high"],
            )
            if self.unit == "minutes":
                return ensure_positive_seconds(sample * 60.0, label)
            return ensure_positive_seconds(float(sample), label)

        if self.name == "Triangular":
            sample = rng.triangular(
                self.raw_params["min"],
                self.raw_params["mode"],
                self.raw_params["max"],
            )
            if self.unit == "minutes":
                return ensure_positive_seconds(sample * 60.0, label)
            return ensure_positive_seconds(float(sample), label)

        kwargs: dict[str, float] = {}
        if self.loc is not None:
            kwargs["loc"] = self.loc
        if self.scale is not None:
            kwargs["scale"] = self.scale

        for _ in range(retries):
            sample = self.scipy_distribution.rvs(*self.shapes, random_state=rng, **kwargs)
            if sample is None or not math.isfinite(sample):
                continue
            if sample <= 0:
                continue
            if self.unit == "minutes":
                return sample * 60.0
            return float(sample)

        raise DataContractError(f"{label} icin pozitif ornek uretilemedi: {self.name} {self.params_text}")


@dataclass(slots=True)
class CraneSpeedProfile:
    empty_mps: float
    loaded_mps: float


@dataclass(slots=True)
class StationDefinition:
    station_id: str
    display_name: str
    x_m: float
    process_group: str | None
    pickup_crane: str | None
    flex_cranes: tuple[str, ...] = ()
    is_virtual: bool = False
    is_sink: bool = False


@dataclass(slots=True)
class StepDefinition:
    key: str
    label: str


@dataclass(slots=True)
class Job:
    job_id: int
    route_family: str
    micron: int
    route: list[StepDefinition]
    created_at: float
    replay_step_seconds: dict[tuple[str, int], float] = field(default_factory=dict)
    source_bara_no: str | None = None
    source_date_text: str | None = None
    source_start_text: str | None = None
    next_step_index: int = 0
    current_station_id: str = "entry_conveyor"
    completed_at: float | None = None
    report_row: dict[str, Any] = field(default_factory=dict)
    step_occurrences: dict[str, int] = field(default_factory=dict)

    @property
    def next_step(self) -> StepDefinition | None:
        if self.next_step_index >= len(self.route):
            return None
        return self.route[self.next_step_index]


@dataclass(slots=True)
class StationVisitRecord:
    visit_id: int
    job_id: int
    route_family: str
    micron: int
    step_key: str
    step_label: str
    station_id: str
    station_name: str
    nominal_seconds: float
    entered_at: float
    exited_at: float | None = None


@dataclass(slots=True)
class ReplayArrivalRecord:
    row_index: int
    route_family: str
    micron: int
    source_bara_no: str | None
    source_date_text: str | None
    source_start_text: str | None
    source_finish_text: str | None
    start_offset_seconds: float | None
    fixed_step_seconds: dict[tuple[str, int], float]


@dataclass(slots=True)
class TransportTask:
    task_id: int
    job: Job
    source_station_id: str
    source_group: str
    pickup_crane: str
    eligible_cranes: tuple[str, ...]
    ready_time: float
    created_sequence: int
    priority: int
    timeout_seconds: float | None = None
    timeout_deadline: float | None = None
    timed_out: bool = False
    claimed: bool = False
    blocked_signature: str | None = None

    @property
    def effective_priority(self) -> int:
        return 0 if self.timed_out else self.priority


SHARED_FLEX_STATIONS = {
    "durulama_1",
    "durulama_2",
    "durulama_3",
    "durulama_4",
    "durulama_5",
}
SHARED_OVERLAP_STATIONS = set(SHARED_FLEX_STATIONS)


@dataclass(slots=True)
class DataBundle:
    source_workbook: Path
    analysis_workbook: Path
    station_positions: dict[str, float]
    station_definitions: dict[str, StationDefinition]
    crane_speeds: dict[str, CraneSpeedProfile]
    pool_distributions: dict[str, DistributionSpec]
    eloxal_distributions_by_micron: dict[int, DistributionSpec]
    interarrival_distribution: DistributionSpec
    arrival_mix: list[tuple[str, int, float]]
    havuz_replay_records: list[ReplayArrivalRecord]
    timeout_seconds_by_group: dict[str, float | None]
    warnings: list[str]


class ExcelDataLoader:
    def __init__(
        self,
        source_workbook: Path,
        analysis_workbook: Path,
        timeout_csv: Path | None = None,
        allow_missing_timeouts: bool = True,
        use_constant_durations: bool = True,
        use_q_arrival_mix: bool = False,
    ) -> None:
        self.source_workbook = Path(source_workbook)
        self.analysis_workbook = Path(analysis_workbook)
        self.timeout_csv = Path(timeout_csv) if timeout_csv else None
        self.allow_missing_timeouts = allow_missing_timeouts
        self.use_constant_durations = use_constant_durations
        self.use_q_arrival_mix = use_q_arrival_mix
        self.warnings: list[str] = []

    def load(self) -> DataBundle:
        station_positions = self._load_station_positions()
        station_definitions = build_station_definitions(station_positions)
        crane_speeds = self._load_crane_speeds()
        pool_distributions, eloxal_by_micron, interarrival = self._load_distribution_specs()
        arrival_mix = self._load_arrival_mix_from_q() if self.use_q_arrival_mix else self._load_arrival_mix()
        havuz_replay_records = self._load_havuz_replay_records()
        timeout_seconds = self._load_timeout_seconds()
        return DataBundle(
            source_workbook=self.source_workbook,
            analysis_workbook=self.analysis_workbook,
            station_positions=station_positions,
            station_definitions=station_definitions,
            crane_speeds=crane_speeds,
            pool_distributions=pool_distributions,
            eloxal_distributions_by_micron=eloxal_by_micron,
            interarrival_distribution=interarrival,
            arrival_mix=arrival_mix,
            havuz_replay_records=havuz_replay_records,
            timeout_seconds_by_group=timeout_seconds,
            warnings=list(self.warnings),
        )

    def _load_station_positions(self) -> dict[str, float]:
        df = pd.read_excel(self.source_workbook, sheet_name="Simülasyon için")
        if "Nereden - Nereye" not in df.columns or "Mesafe (cm)" not in df.columns:
            raise DataContractError("'Simülasyon için' sayfasinda mesafe tablosu bulunamadi.")

        positions: dict[str, float] = {"entry_conveyor": 0.0}
        cumulative = 0.0
        duplicate_counter: Counter[str] = Counter()

        for _, row in df[["Nereden - Nereye", "Mesafe (cm)"]].dropna(subset=["Nereden - Nereye"]).iterrows():
            label = str(row["Nereden - Nereye"]).strip()
            distance_cm = float(row["Mesafe (cm)"])
            parts = re.split(r"\s*-\s*", label, maxsplit=1)
            if len(parts) != 2:
                raise DataContractError(f"Mesafe etiketi cozulmedi: {label}")
            end_raw = parts[1]
            normalized = normalize_token(end_raw)
            duplicate_counter[normalized] += 1
            station_id = map_linear_station_id(normalized, duplicate_counter[normalized], end_raw)
            cumulative += distance_cm / 100.0
            positions[station_id] = cumulative

        required = {
            "yagalma_1",
            "yagalma_2",
            "durulama_y",
            "kostik",
            "durulama_kostik",
            "durulama_1",
            "durulama_2",
            "notralizasyon_1",
            "notralizasyon_2",
            "eloksal_1",
            "eloksal_5",
            "durulama_3",
            "durulama_5",
            "kalay_1",
            "kalay_2",
            "durulama_kalay",
            "nikel",
            "durulama_nikel",
            "di",
            "tespit_1",
            "tespit_3",
            "durulama_t",
            "sicaksu_1",
            "sicaksu_3",
            "suzme_1",
            "suzme_3",
            "firin_1",
            "firin_2",
            "exit_conveyor",
        }
        missing = sorted(required - positions.keys())
        if missing:
            raise DataContractError(f"Mesafe tablosunda beklenen istasyonlar eksik: {missing}")
        return positions

    def _load_crane_speeds(self) -> dict[str, CraneSpeedProfile]:
        df = pd.read_excel(self.source_workbook, sheet_name="Simülasyon için")
        if "Vinç Hızları (m/sn)" not in df.columns:
            raise DataContractError("'Simülasyon için' sayfasinda vinc hizlari bulunamadi.")

        speeds: dict[str, CraneSpeedProfile] = {}
        for _, row in df[["Vinç Hızları (m/sn)", "Unnamed: 4"]].dropna(subset=["Vinç Hızları (m/sn)"]).iterrows():
            crane_name = normalize_token(row["Vinç Hızları (m/sn)"])
            speed_value = row["Unnamed: 4"]
            if crane_name.startswith("vinc") and pd.notna(speed_value):
                speed = float(speed_value)
                speeds[crane_name] = CraneSpeedProfile(empty_mps=speed, loaded_mps=speed)

        if len(speeds) != 4:
            raise DataContractError(f"Beklenen 4 vinc hizi okunamadi: {speeds}")

        return speeds

    def _load_distribution_specs(
        self,
    ) -> tuple[dict[str, DistributionSpec], dict[int, DistributionSpec], DistributionSpec]:
        df = pd.read_excel(self.analysis_workbook, sheet_name="ÖZET", header=1)
        if "Kategori" not in df.columns:
            raise DataContractError("Analiz ozet sayfasinin basliklari okunamadi.")

        pool_distributions: dict[str, DistributionSpec] = {}
        eloxal_by_micron: dict[int, DistributionSpec] = {}
        interarrival: DistributionSpec | None = None

        for _, row in df.dropna(subset=["Kategori", "Grup", "En İyi Dağılım", "Parametreler"]).iterrows():
            category = normalize_token(row["Kategori"])
            group = normalize_token(row["Grup"])
            spec = DistributionSpec(
                name=str(row["En İyi Dağılım"]).strip(),
                params_text=str(row["Parametreler"]).strip(),
                unit="minutes",
            )

            if category == "havuzsureler":
                canonical = map_summary_group_to_process_key(group)
                if canonical:
                    pool_distributions[canonical] = spec
            elif category == "gelislerarasi":
                interarrival = spec
            elif category == "eloksaluretim":
                micron = parse_micron_group(group)
                if micron in VALID_MICRONS:
                    eloxal_by_micron[micron] = spec

        required_pools = {
            "yagalma",
            "durulama_y",
            "asitmat",
            "durulama_a",
            "sokme",
            "kostik",
            "durulama_kostik",
            "durulama_1",
            "durulama_2",
            "notralizasyon",
            "durulama_post_eloksal",
            "renk",
            "durulama_post_renk",
            "di",
            "durulama_post_tespit",
            "sicaksu",
            "suzme",
            "firin",
        }
        missing_pools = sorted(required_pools - pool_distributions.keys())
        if missing_pools:
            raise DataContractError(f"Analiz dosyasinda eksik havuz dagilimlari var: {missing_pools}")
        if interarrival is None:
            raise DataContractError("Gelisler arasi dagilimi bulunamadi.")
        if set(eloxal_by_micron.keys()) != VALID_MICRONS:
            raise DataContractError(f"Eloksal mikron dagilimlari eksik: {sorted(VALID_MICRONS - set(eloxal_by_micron))}")

        interarrival = self._build_fixed_interarrival_override()
        if self.use_constant_durations:
            pool_distributions.update(self._build_fixed_process_overrides())
            eloxal_by_micron.update(self._build_fixed_eloksal_overrides())
        else:
            pool_distributions.update(self._build_rinse_overrides())
            pool_distributions.update(self._load_triangular_overrides())

        return pool_distributions, eloxal_by_micron, interarrival

    def _load_triangular_overrides(self) -> dict[str, DistributionSpec]:
        df = pd.read_excel(self.source_workbook, sheet_name="Havuz Süreler")
        overrides: dict[str, DistributionSpec] = {}
        column_map = {
            "asitmat": "AsitMat",
            "sokme": "Sökme",
        }
        for process_group, column_name in column_map.items():
            if column_name not in df.columns:
                raise DataContractError(f"Havuz Süreler sayfasinda '{column_name}' kolonu bulunamadi.")

            values_seconds = [
                int(round(seconds))
                for seconds in (to_seconds(value) for value in df[column_name])
                if seconds is not None and seconds > 0
            ]
            if len(values_seconds) < 2:
                raise DataContractError(
                    f"{column_name} icin triangular dagilim kurmak uzere yeterli ham veri yok."
                )

            min_seconds = min(values_seconds)
            max_seconds = max(values_seconds)
            mode_seconds = int(pd.Series(values_seconds).mode().iloc[0])
            mode_seconds = min(max(mode_seconds, min_seconds), max_seconds)

            overrides[process_group] = DistributionSpec(
                name="Triangular",
                params_text=(
                    f"min={min_seconds / 60.0:.6f} | "
                    f"mode={mode_seconds / 60.0:.6f} | "
                    f"max={max_seconds / 60.0:.6f}"
                ),
                unit="minutes",
            )
        return overrides

    def _build_rinse_overrides(self) -> dict[str, DistributionSpec]:
        return {
            process_group: DistributionSpec(
                name="Uniform",
                params_text="low=20 | high=25",
                unit="seconds",
            )
            for process_group in sorted(RINSE_PROCESS_GROUPS)
        }

    def _build_fixed_process_overrides(self) -> dict[str, DistributionSpec]:
        fixed_minutes = {
            "yagalma": 5.0,
            "durulama_y": 0.5,
            "asitmat": 4.0,
            "durulama_a": 0.5,
            "sokme": 0.7,
            "kostik": 5.0,
            "durulama_kostik": 0.5,
            "durulama_1": 0.5,
            "durulama_2": 0.5,
            "notralizasyon": 2.0,
            "durulama_post_eloksal": 0.5,
            "renk": 10.0,
            "durulama_post_renk": 0.5,
            "di": 0.5,
            "durulama_post_tespit": 0.5,
            "sicaksu": 2.5,
            "suzme": 3.5,
            "firin": 8.0,
        }
        return {
            process_group: DistributionSpec(
                name="Constant",
                params_text=f"value={value_minutes}",
                unit="minutes",
            )
            for process_group, value_minutes in fixed_minutes.items()
        }

    def _build_fixed_eloksal_overrides(self) -> dict[int, DistributionSpec]:
        fixed_minutes = {
            7: 20.0,
            13: 30.0,
            25: 50.0,
        }
        return {
            micron: DistributionSpec(
                name="Constant",
                params_text=f"value={value_minutes}",
                unit="minutes",
            )
            for micron, value_minutes in fixed_minutes.items()
        }

    def _build_fixed_interarrival_override(self) -> DistributionSpec:
        return DistributionSpec(
            name="Gamma",
            params_text="shape=3.480 | scale=2.181",
            unit="minutes",
        )

    def _load_arrival_mix(self) -> list[tuple[str, int, float]]:
        df = pd.read_excel(self.source_workbook, sheet_name="Eloksal Üretim")
        if "Mikron" not in df.columns or "Renk" not in df.columns:
            raise DataContractError("'Eloksal Üretim' sayfasinda mikron/renk kolonlari okunamadi.")

        if "İptal" in df.columns:
            df = df[df["İptal"].isna()]

        df = df[df["Mikron"].isin(sorted(VALID_MICRONS))].copy()
        df["route_family"] = df["Renk"].map(self._classify_route_family)
        df = df.dropna(subset=["route_family"])

        if df.empty:
            raise DataContractError("Gecerli mikron ve rota ailesi eslesmesine sahip uretim kaydi bulunamadi.")

        counts = df.groupby(["route_family", "Mikron"]).size().reset_index(name="count")
        total = int(counts["count"].sum())
        if total <= 0:
            raise DataContractError("Gelis karmasi hesaplanamadi.")

        arrival_mix: list[tuple[str, int, float]] = []
        for _, row in counts.sort_values("count", ascending=False).iterrows():
            arrival_mix.append((str(row["route_family"]), int(row["Mikron"]), float(row["count"]) / float(total)))
        return arrival_mix

    def _load_arrival_mix_from_q(self) -> list[tuple[str, int, float]]:
        workbook = default_q_workbook()
        if not workbook.exists():
            raise DataContractError(
                f"Q arrival mix secildi ama dosya bulunamadi: {workbook}. "
                "Lutfen Q.xlsx dosyasini kodun bulundugu klasore koyun."
            )

        df = pd.read_excel(workbook, sheet_name=0)
        normalized_columns = {normalize_token(column): column for column in df.columns}
        color_col = (
            normalized_columns.get("renk")
            or normalized_columns.get("cesit")
            or normalized_columns.get("uruntipi")
            or normalized_columns.get("urun")
            or normalized_columns.get("rota")
        )
        micron_col = (
            normalized_columns.get("mikron")
            or normalized_columns.get("micron")
            or normalized_columns.get("kalinlik")
            or normalized_columns.get("um")
        )
        if not color_col or not micron_col:
            raise DataContractError(
                "Q.xlsx icinde arrival mix icin en az 'Renk/Cesit' ve 'Mikron' kolonlari bulunmali."
            )

        local_df = df[[color_col, micron_col]].copy()
        local_df["route_family"] = local_df[color_col].map(self._classify_route_family)
        local_df["resolved_mikron"] = local_df[micron_col].map(nearest_valid_micron)
        skipped_count = int(
            ((local_df["route_family"].isna()) | (local_df["resolved_mikron"].isna())).sum()
        )
        local_df = local_df.dropna(subset=["route_family", "resolved_mikron"])
        if skipped_count > 0:
            self.warnings.append(
                f"Q.xlsx arrival mix icinde rota/mikron cozumlenemeyen {skipped_count} satir atlandi."
            )

        if local_df.empty:
            raise DataContractError("Q.xlsx icinde gecerli rota ailesi ve mikron eslesmesine sahip kayit bulunamadi.")

        counts = (
            local_df.groupby(["route_family", "resolved_mikron"])
            .size()
            .reset_index(name="count")
        )
        total = int(counts["count"].sum())
        if total <= 0:
            raise DataContractError("Q.xlsx arrival mix olasiliklari hesaplanamadi.")

        arrival_mix: list[tuple[str, int, float]] = []
        for _, row in counts.sort_values("count", ascending=False).iterrows():
            arrival_mix.append(
                (
                    str(row["route_family"]),
                    int(row["resolved_mikron"]),
                    float(row["count"]) / float(total),
                )
            )
        return arrival_mix

    def _infer_route_family_from_havuz_row(self, row: pd.Series) -> str | None:
        explicit = self._classify_route_family(row.get("Renk"))
        if explicit is not None:
            return explicit
        has_asitmat = any(
            (to_seconds(row.get(column_name)) or 0) > 0
            for column_name in ["AsitMat", "Sökme"]
        )
        has_kostik = any(
            (to_seconds(row.get(column_name)) or 0) > 0
            for column_name in ["Kostik", "Durulama V1"]
        )
        has_color = any(
            (to_seconds(row.get(column_name)) or 0) > 0
            for column_name in ["Kalay / Nikel (Renk)", "Durulama.3"]
        )

        if has_asitmat:
            return "asitmat"
        if has_color and has_kostik:
            return "man"
        if has_color:
            return "pan"
        if has_kostik:
            return "natural"
        return "parlak"

    def _resolve_havuz_micron(
        self,
        row: pd.Series,
        eu_matches: pd.DataFrame,
    ) -> int | None:
        if not eu_matches.empty and "Mikron" in eu_matches.columns:
            mikron_values = [
                int(value)
                for value in eu_matches["Mikron"].dropna().tolist()
                if int(value) in VALID_MICRONS
            ]
            if mikron_values:
                return int(mikron_values[0])

        tespit_seconds = to_seconds(row.get("Tespit"))
        inferred = infer_valid_micron_from_seconds(tespit_seconds)
        if inferred in VALID_MICRONS:
            return inferred
        return None

    def _fixed_step_seconds_from_havuz_row(
        self,
        row: pd.Series,
        route_family: str,
    ) -> dict[tuple[str, int], float]:
        step_map: list[tuple[str, str, int]] = [
            ("YağAlma", "yagalma", 0),
            ("Durulama", "durulama_y", 0),
            ("AsitMat", "asitmat", 0),
            ("Durulama.1", "durulama_a", 0),
            ("Sökme", "sokme", 0),
            ("Kostik", "kostik", 0),
            ("Durulama V1", "durulama_kostik", 0),
            ("Durulama V2", "pre_neutral_rinse", 0),
            ("Durulama V3", "pre_neutral_rinse", 1),
            ("Nötralizasyon", "notralizasyon", 0),
            ("Eloksal", "eloksal", 0),
            ("Durulama.2", "post_eloksal_rinse", 0),
            ("Kalay / Nikel (Renk)", "kalay" if route_family == "man" else "nikel", 0),
            ("Durulama.3", "post_color_rinse", 0),
            ("DI", "di", 0),
            ("Tespit", "tespit", 0),
            ("Durulama.4", "post_tespit_rinse", 0),
            ("SıcakSu", "sicaksu", 0),
            ("Süzme", "drip", 0),
            ("Fırın", "firin", 0),
        ]

        durations: dict[tuple[str, int], float] = {}
        for column_name, step_key, occurrence in step_map:
            seconds = to_seconds(row.get(column_name))
            if seconds is None or seconds <= 0:
                continue
            durations[(step_key, occurrence)] = float(seconds)
        return durations

    def _load_havuz_replay_records(self) -> list[ReplayArrivalRecord]:
        hs = pd.read_excel(self.source_workbook, sheet_name="Havuz Süreler")
        eu = pd.read_excel(self.source_workbook, sheet_name="Eloksal Üretim")

        hs["__row_index__"] = range(len(hs))
        hs["__date__"] = pd.to_datetime(hs["Tarih"], errors="coerce").dt.date
        hs["__bara_key__"] = pd.to_numeric(hs["BaraNo."], errors="coerce").astype("Int64")
        hs["__route_family__"] = hs.apply(self._infer_route_family_from_havuz_row, axis=1)

        eu["__date__"] = pd.to_datetime(eu["Tarih"], errors="coerce").dt.date
        eu["__bara_key__"] = pd.to_numeric(eu["Bara No."], errors="coerce").astype("Int64")
        eu["__route_family__"] = eu["Renk"].map(self._classify_route_family)

        records: list[ReplayArrivalRecord] = []
        actual_datetimes: list[datetime] = []

        for _, row in hs.iterrows():
            route_family = row.get("__route_family__")
            if route_family is None:
                continue

            eu_matches = eu[
                (eu["__date__"] == row.get("__date__"))
                & (eu["__bara_key__"] == row.get("__bara_key__"))
            ].copy()
            if route_family is not None and not eu_matches.empty:
                routed = eu_matches[eu_matches["__route_family__"] == route_family]
                if not routed.empty:
                    eu_matches = routed

            micron = self._resolve_havuz_micron(row, eu_matches)
            if micron not in VALID_MICRONS:
                self.warnings.append(
                    f"Havuz Süreler replay kaydi atlandi: satir {int(row['__row_index__']) + 2} icin mikron cozumlenemedi."
                )
                continue

            start_clock = row.get("Başlangıç")
            if pd.isna(start_clock) and not eu_matches.empty and "Giriş Saati" in eu_matches.columns:
                start_clock = eu_matches.iloc[0].get("Giriş Saati")
            finish_clock = row.get("Bitiş")
            if pd.isna(finish_clock) and not eu_matches.empty and "Çıkış Saati" in eu_matches.columns:
                finish_clock = eu_matches.iloc[0].get("Çıkış Saati")

            start_offset_seconds: float | None = None
            if row.get("__date__") is not None and to_seconds(start_clock) is not None:
                if hasattr(start_clock, "hour") and hasattr(start_clock, "minute") and hasattr(start_clock, "second"):
                    start_time = start_clock
                else:
                    start_time = pd.to_datetime(start_clock).time()
                actual_dt = datetime.combine(row["__date__"], start_time)
                actual_datetimes.append(actual_dt)
                start_offset_seconds = float(actual_dt.timestamp())

            records.append(
                ReplayArrivalRecord(
                    row_index=int(row["__row_index__"]),
                    route_family=str(route_family),
                    micron=int(micron),
                    source_bara_no=None if pd.isna(row.get("BaraNo.")) else str(int(float(row.get("BaraNo.")))),
                    source_date_text="" if pd.isna(row.get("Tarih")) else str(pd.to_datetime(row.get("Tarih")).date()),
                    source_start_text="" if pd.isna(start_clock) else format_seconds_as_hms(to_seconds(start_clock)),
                    source_finish_text="" if pd.isna(finish_clock) else format_seconds_as_hms(to_seconds(finish_clock)),
                    start_offset_seconds=start_offset_seconds,
                    fixed_step_seconds=self._fixed_step_seconds_from_havuz_row(row, str(route_family)),
                )
            )

        if actual_datetimes:
            baseline = min(actual_datetimes).timestamp()
            for record in records:
                if record.start_offset_seconds is not None:
                    record.start_offset_seconds = float(record.start_offset_seconds - baseline)

        records.sort(
            key=lambda record: (
                "" if not record.source_date_text else record.source_date_text,
                float("inf") if to_seconds(record.source_start_text) is None else float(to_seconds(record.source_start_text)),
                record.row_index,
            )
        )
        return records

    def _classify_route_family(self, color_value: Any) -> str | None:
        color = normalize_token(color_value)
        if not color:
            return None
        if color in {"n", "natural", "naturel"}:
            return "natural"
        if "parlak" in color:
            return "parlak"
        if "sakem" in color or "saken" in color:
            return "sakem"
        if "asit" in color:
            return "asitmat"
        if color.startswith("man") or "irak" in color:
            return "man"
        if color.startswith("pan"):
            return "pan"
        if any(token in color for token in {"nazilli", "suriye", "sevinc", "onur"}):
            return "pan"
        return None

    def _load_timeout_seconds(self) -> dict[str, float | None]:
        if self.timeout_csv and self.timeout_csv.exists():
            df = pd.read_csv(self.timeout_csv)
            return parse_timeout_dataframe(df)

        if self.allow_missing_timeouts:
            return {}

        raise DataContractError(
            "Timeout toleranslari kaynak dosyalarda acik bir tablo olarak bulunamadi. "
            "Sahte veri kullanmamak icin simülasyon durduruldu. "
            "CSV formatinda 'process_group,timeout_seconds' kolonlari ile ek bir tolerans dosyasi verin "
            "veya '--allow-missing-timeouts' bayragi ile bu ozelligi gecici olarak kapatin."
        )


def parse_timeout_dataframe(df: pd.DataFrame) -> dict[str, float | None]:
    expected = {normalize_token(col): col for col in df.columns}
    process_col = expected.get("processgroup") or expected.get("station") or expected.get("group")
    timeout_col = expected.get("timeoutseconds") or expected.get("tolerance_seconds") or expected.get("timeout")
    if not process_col or not timeout_col:
        raise DataContractError("Timeout CSV kolonlari 'process_group' ve 'timeout_seconds' seklinde olmali.")

    result: dict[str, float | None] = {}
    for _, row in df.dropna(subset=[process_col]).iterrows():
        process_group = normalize_token(row[process_col])
        timeout_seconds = to_seconds(row[timeout_col])
        result[process_group] = timeout_seconds
    return result


def map_linear_station_id(normalized_end: str, occurrence: int, raw_label: str) -> str:
    mapping = {
        "buffer1": "buffer_1",
        "buffer2": "buffer_2",
        "yagalma1": "yagalma_1",
        "yagalma2": "yagalma_2",
        "durulamay": "durulama_y",
        "asitmat": "asitmat",
        "durulamaa": "durulama_a",
        "sokme": "sokme",
        "kostik": "kostik",
        "durulama1": "durulama_1",
        "durulama2": "durulama_2",
        "notralizasyon1": "notralizasyon_1",
        "notralizasyon2": "notralizasyon_2",
        "neutralizasyon1": "notralizasyon_1",
        "neutralizasyon2": "notralizasyon_2",
        "eloksal1": "eloksal_1",
        "eloksal2": "eloksal_2",
        "eloksal3": "eloksal_3",
        "durulamae": "durulama_e",
        "eloksal4": "eloksal_4",
        "eloksal5": "eloksal_5",
        "elosakl5": "eloksal_5",
        "eloksal6": "eloksal_6",
        "durulama3": "durulama_3",
        "durulama4": "durulama_4",
        "durulama5": "durulama_5",
        "kalay1": "kalay_1",
        "kalay2": "kalay_2",
        "nikel": "nikel",
        "durulaman": "durulama_nikel",
        "di": "di",
        "tespit1": "tespit_1",
        "tespit2": "tespit_2",
        "tespit3": "tespit_3",
        "durulamat": "durulama_t",
        "sicaksu1": "sicaksu_1",
        "sicaksu2": "sicaksu_2",
        "sicaksu3": "sicaksu_3",
        "suzme1": "suzme_1",
        "suzme2": "suzme_2",
        "suzme3": "suzme_3",
        "firin1": "firin_1",
        "firin2": "firin_2",
        "konveyor2": "exit_conveyor",
    }
    if normalized_end == "durulamak":
        return "durulama_kostik" if occurrence == 1 else "durulama_kalay"
    try:
        return mapping[normalized_end]
    except KeyError as exc:
        raise DataContractError(f"Mesafe tablosunda beklenmeyen istasyon sonu bulundu: {raw_label}") from exc


def map_summary_group_to_process_key(group: str) -> str | None:
    mapping = {
        "yagalma": "yagalma",
        "durulama": "durulama_y",
        "asitmat": "asitmat",
        "durulama2": "durulama_a",
        "sokme": "sokme",
        "kostik": "kostik",
        "durulamav1": "durulama_kostik",
        "durulamav2": "durulama_1",
        "durulamav3": "durulama_2",
        "notralizasyon": "notralizasyon",
        "eloksal": "eloksal",
        "durulama3": "durulama_post_eloksal",
        "kalaynikelrenk": "renk",
        "durulama4": "durulama_post_renk",
        "di": "di",
        "tespit": "tespit",
        "durulama5": "durulama_post_tespit",
        "sicaksu": "sicaksu",
        "suzme": "suzme",
        "firin": "firin",
    }
    return mapping.get(group)


def parse_micron_group(group: str) -> int | None:
    match = re.search(r"(\d+)", group)
    if not match:
        return None
    return int(match.group(1))


def build_station_definitions(positions: dict[str, float]) -> dict[str, StationDefinition]:
    station_rows = {
        "entry_conveyor": ("Giris Konveyor", None, "vinc1", (), True, False),
        "yagalma_1": ("Yag Alma 1", "yagalma", "vinc1", (), False, False),
        "yagalma_2": ("Yag Alma 2", "yagalma", "vinc1", (), False, False),
        "durulama_y": ("Durulama Y", "durulama_y", "vinc1", (), False, False),
        "asitmat": ("AsitMat", "asitmat", "vinc1", (), False, False),
        "durulama_a": ("Durulama A", "durulama_a", "vinc1", (), False, False),
        "sokme": ("Sokme", "sokme", "vinc1", (), False, False),
        "kostik": ("Kostik", "kostik", "vinc1", (), False, False),
        "durulama_kostik": ("Kostik Durulama", "durulama_kostik", "vinc1", (), False, False),
        "durulama_1": ("Durulama 1", "durulama_1", "vinc1", ("vinc2",), False, False),
        "durulama_2": ("Durulama 2", "durulama_2", "vinc1", ("vinc2",), False, False),
        "notralizasyon_1": ("Notralizasyon 1", "notralizasyon", "vinc2", (), False, False),
        "notralizasyon_2": ("Notralizasyon 2", "notralizasyon", "vinc2", (), False, False),
        "eloksal_1": ("Eloksal 1", "eloksal", "vinc2", (), False, False),
        "eloksal_2": ("Eloksal 2", "eloksal", "vinc2", (), False, False),
        "eloksal_3": ("Eloksal 3", "eloksal", "vinc2", (), False, False),
        "durulama_e": ("Durulama E", None, "vinc2", (), False, False),
        "eloksal_4": ("Eloksal 4", "eloksal", "vinc2", (), False, False),
        "eloksal_5": ("Eloksal 5", "eloksal", "vinc2", (), False, False),
        "durulama_3": ("Durulama 3", "durulama_post_eloksal", "vinc3", ("vinc2",), False, False),
        "durulama_4": ("Durulama 4", "durulama_post_eloksal", "vinc3", ("vinc2",), False, False),
        "durulama_5": ("Durulama 5", "durulama_post_eloksal", "vinc3", ("vinc2",), False, False),
        "kalay_1": ("Kalay 1", "renk", "vinc3", (), False, False),
        "kalay_2": ("Kalay 2", "renk", "vinc3", (), False, False),
        "durulama_kalay": ("Durulama Kalay", "durulama_post_renk", "vinc3", (), False, False),
        "nikel": ("Nikel", "renk", "vinc3", (), False, False),
        "durulama_nikel": ("Durulama Nikel", "durulama_post_renk", "vinc3", (), False, False),
        "di": ("DI", "di", "vinc3", (), False, False),
        "tespit_1": ("Tespit 1", "tespit", "vinc3", (), False, False),
        "tespit_2": ("Tespit 2", "tespit", "vinc3", (), False, False),
        "tespit_3": ("Tespit 3", "tespit", "vinc3", (), False, False),
        "durulama_t": ("Durulama T", "durulama_post_tespit", "vinc3", (), False, False),
        "sicaksu_1": ("Sicak Su 1", "sicaksu", "vinc4", (), False, False),
        "sicaksu_2": ("Sicak Su 2", "sicaksu", "vinc4", (), False, False),
        "sicaksu_3": ("Sicak Su 3", "sicaksu", "vinc4", (), False, False),
        "firin_1": ("Firin 1", "firin", "vinc4", (), False, False),
        "firin_2": ("Firin 2", "firin", "vinc4", (), False, False),
        "exit_conveyor": ("Cikis Konveyor", None, None, (), True, True),
    }
    definitions: dict[str, StationDefinition] = {}
    for station_id, (
        display_name,
        process_group,
        pickup_crane,
        flex_cranes,
        is_virtual,
        is_sink,
    ) in station_rows.items():
        definitions[station_id] = StationDefinition(
            station_id=station_id,
            display_name=display_name,
            x_m=positions[station_id],
            process_group=process_group,
            pickup_crane=pickup_crane,
            flex_cranes=tuple(flex_cranes),
            is_virtual=is_virtual,
            is_sink=is_sink,
        )
    return definitions


class Dispatcher:
    def __init__(self, env: simpy.Environment, station_definitions: dict[str, StationDefinition]) -> None:
        self.env = env
        self.station_definitions = station_definitions
        self.pending: dict[int, TransportTask] = {}
        self.waiters: dict[str, simpy.Event] = {}

    def publish(self, task: TransportTask) -> None:
        self.pending[task.task_id] = task
        for crane_id in task.eligible_cranes:
            self._wake(crane_id)

    def escalate_timeout(self, task_id: int) -> None:
        task = self.pending.get(task_id)
        if task is None or task.claimed:
            return
        task.timed_out = True
        for crane_id in task.eligible_cranes:
            self._wake(crane_id)

    def _wake(self, crane_id: str) -> None:
        waiter = self.waiters.pop(crane_id, None)
        if waiter is not None and not waiter.triggered:
            waiter.succeed(True)

    def wake_all(self) -> None:
        for crane_id in list(self.waiters):
            self._wake(crane_id)

    def requeue(self, task: TransportTask) -> None:
        task.claimed = False
        self.pending[task.task_id] = task
        for crane_id in task.eligible_cranes:
            self._wake(crane_id)

    def _choose_best_task(self, crane: "Crane") -> TransportTask | None:
        eligible = [
            task
            for task in self.pending.values()
            if crane.crane_id in task.eligible_cranes
            and crane.plant.is_task_dispatchable(task, crane)
            and crane.plant.can_crane_claim_task(crane, task)
        ]
        if not eligible:
            return None

        def rank(task: TransportTask) -> tuple[int, int, float, float, float, int]:
            source_x = crane.plant.station_definitions[task.source_station_id].x_m
            distance = abs(crane.current_x - source_x)
            flex_penalty = 0.0 if task.pickup_crane == crane.crane_id else 1000.0
            passive_subpriority = 0 if task.source_station_id == "entry_conveyor" else 1
            if task.effective_priority < 2:
                passive_subpriority = 0
            return (
                task.effective_priority,
                passive_subpriority,
                flex_penalty,
                distance,
                task.ready_time,
                task.created_sequence,
            )

        eligible.sort(key=rank)
        return eligible[0]

    def wait_for_next_task(self, crane: "Crane"):
        while True:
            best = self._choose_best_task(crane)
            if best is not None:
                best.claimed = True
                best.blocked_signature = None
                self.pending.pop(best.task_id, None)
                return best
            blocked_tasks = [
                task for task in self.pending.values() if crane.crane_id in task.eligible_cranes
            ]
            for task in blocked_tasks:
                crane.plant.log_task_blockage(task)
            waiter = self.env.event()
            self.waiters[crane.crane_id] = waiter
            yield waiter


class PhysicalStation:
    def __init__(self, env: simpy.Environment, definition: StationDefinition, plant: "AnodizingPlant") -> None:
        self.env = env
        self.definition = definition
        self.plant = plant
        self.resource = None if definition.is_virtual or definition.is_sink else simpy.Resource(env, capacity=1)
        self.current_job: Job | None = None
        self.current_request: simpy.ResourceRequest | None = None
        self.current_task_id: int | None = None
        self.current_visit_id: int | None = None
        self.current_process_complete_at: float | None = None

    def accept_job(self, job: Job, request: simpy.ResourceRequest | None, process_seconds: float) -> None:
        if self.definition.is_virtual or self.definition.is_sink:
            raise DataContractError(f"{self.definition.station_id} fiziksel havuz olarak kullanilamaz.")
        if self.current_job is not None:
            raise DataContractError(f"{self.definition.display_name} doluyken ikinci is kabul edildi.")
        self.current_job = job
        self.current_request = request
        self.current_process_complete_at = self.env.now + process_seconds
        job.current_station_id = self.definition.station_id
        self.env.process(self._process_job(job, process_seconds))

    def occupy_without_process(self, job: Job, request: simpy.ResourceRequest | None) -> None:
        if self.definition.is_virtual or self.definition.is_sink:
            raise DataContractError(f"{self.definition.station_id} fiziksel havuz olarak kullanilamaz.")
        if self.current_job is not None:
            raise DataContractError(f"{self.definition.display_name} doluyken ikinci is kabul edildi.")
        self.current_job = job
        self.current_request = request
        self.current_process_complete_at = None
        job.current_station_id = self.definition.station_id

    def release_job_for_pickup(self) -> Job:
        if self.current_job is None:
            raise DataContractError(f"{self.definition.display_name} icinde alinacak is yok.")
        job = self.current_job
        if self.current_visit_id is not None:
            self.plant.finish_station_visit(self.current_visit_id, self.env.now)
        self.current_job = None
        self.current_task_id = None
        self.current_visit_id = None
        self.current_process_complete_at = None
        if self.current_request is not None:
            self.resource.release(self.current_request)
            self.current_request = None
        self.plant.dispatcher.wake_all()
        return job

    def _process_job(self, job: Job, process_seconds: float):
        yield self.env.timeout(process_seconds)
        self.current_process_complete_at = None
        self.plant.handle_station_process_completion(self, job, process_seconds)


class RailController:
    def __init__(
        self,
        env: simpy.Environment,
        crane_order: list[str],
        initial_positions: dict[str, float],
        zone_bounds: dict[str, tuple[float, float]],
        safety_gap_m: float = 0.05,
    ) -> None:
        self.env = env
        self.crane_order = list(crane_order)
        self.positions = dict(initial_positions)
        self.zone_bounds = dict(zone_bounds)
        self.safety_gap_m = safety_gap_m
        self.waiters: list[simpy.Event] = []
        self.last_wait_signature: dict[str, tuple[str, str | None, float]] = {}
        self.wait_started_at: dict[str, float] = {}
        self.active_moves: dict[str, tuple[float, float, float]] = {}
        self.on_position_change = None

    def _projected_position(self, crane_id: str) -> float:
        active = self.active_moves.get(crane_id)
        if active is not None:
            return active[2]
        return self.positions[crane_id]

    def _path_interval(self, start_x: float, target_x: float) -> tuple[float, float]:
        return (min(start_x, target_x), max(start_x, target_x))

    def _validate_zone_target(self, crane_id: str, target_x: float) -> None:
        zone_min, zone_max = self.zone_bounds[crane_id]
        if zone_min - 1e-9 <= target_x <= zone_max + 1e-9:
            return
        raise DataContractError(
            f"{crane_id} x={target_x:.3f} noktasina gidemez. "
            f"Izinli fiziksel bolge [{zone_min:.3f}, {zone_max:.3f}]"
        )

    def _neighbor_bounds(self, crane_id: str) -> tuple[float, float]:
        idx = self.crane_order.index(crane_id)
        left_bound = -math.inf
        right_bound = math.inf
        if idx > 0:
            left_neighbor = self.crane_order[idx - 1]
            left_bound = self._projected_position(left_neighbor) + self.safety_gap_m
        if idx < len(self.crane_order) - 1:
            right_neighbor = self.crane_order[idx + 1]
            right_bound = self._projected_position(right_neighbor) - self.safety_gap_m
        return left_bound, right_bound

    def can_move_to(self, crane_id: str, target_x: float) -> bool:
        zone_min, zone_max = self.zone_bounds[crane_id]
        if not (zone_min - 1e-9 <= target_x <= zone_max + 1e-9):
            return False
        left_bound, right_bound = self._neighbor_bounds(crane_id)
        return left_bound <= target_x <= right_bound

    def blocking_crane_ids(self, crane_id: str, target_x: float) -> list[str]:
        idx = self.crane_order.index(crane_id)
        start_x = self.positions[crane_id]
        path_min, path_max = self._path_interval(start_x, target_x)
        blockers: list[str] = []

        if idx > 0:
            left_neighbor = self.crane_order[idx - 1]
            left_position = self._projected_position(left_neighbor)
            if target_x < left_position + self.safety_gap_m:
                blockers.append(left_neighbor)
            active = self.active_moves.get(left_neighbor)
            if active is not None:
                neighbor_min, neighbor_max, _ = active
                if path_min <= neighbor_max + self.safety_gap_m and neighbor_min <= path_max + self.safety_gap_m:
                    blockers.append(left_neighbor)
        if idx < len(self.crane_order) - 1:
            right_neighbor = self.crane_order[idx + 1]
            right_position = self._projected_position(right_neighbor)
            if target_x > right_position - self.safety_gap_m:
                blockers.append(right_neighbor)
            active = self.active_moves.get(right_neighbor)
            if active is not None:
                neighbor_min, neighbor_max, _ = active
                if path_min <= neighbor_max + self.safety_gap_m and neighbor_min <= path_max + self.safety_gap_m:
                    blockers.append(right_neighbor)

        unique_blockers: list[str] = []
        for blocker in blockers:
            if blocker not in unique_blockers:
                unique_blockers.append(blocker)
        return unique_blockers

    def notify_position_change(self) -> None:
        waiters = self.waiters
        self.waiters = []
        for waiter in waiters:
            if not waiter.triggered:
                waiter.succeed(True)
        if self.on_position_change is not None:
            self.on_position_change()

    def move(
        self,
        crane: "Crane",
        target_x: float,
        speed_mps: float,
        target_station_id: str | None = None,
    ):
        self._validate_zone_target(crane.crane_id, target_x)
        while True:
            blocking_crane_ids = self.blocking_crane_ids(crane.crane_id, target_x)
            if not blocking_crane_ids:
                wait_started_at = self.wait_started_at.pop(crane.crane_id, None)
                previous_signature = self.last_wait_signature.pop(crane.crane_id, None)
                if wait_started_at is not None:
                    wait_seconds = max(0.0, float(self.env.now) - float(wait_started_at))
                    blocker_labels = ""
                    if previous_signature is not None and previous_signature[0]:
                        blocker_labels = ", ".join(
                            crane_label(blocker_id)
                            for blocker_id in previous_signature[0].split("|")
                            if blocker_id
                        )
                    crane.plant.log_event(
                        event_type="crane_wait_done",
                        message=(
                            f"{crane_label(crane.crane_id)}, "
                            f"{crane.plant.describe_target_location(target_station_id, target_x)} yonune hareket etmeden once "
                            f"{wait_seconds:.1f} sn bekledi."
                        ),
                        crane_id=crane.crane_id,
                        destination_station_id=target_station_id,
                        duration_seconds=wait_seconds,
                        metadata={
                            "blockage_kind": "rail_neighbor",
                            "blocking_crane": blocker_labels,
                        },
                    )
                self.active_moves[crane.crane_id] = (*self._path_interval(crane.current_x, target_x), target_x)
                self.notify_position_change()
                yield self.env.timeout(abs(crane.current_x - target_x) / speed_mps)
                crane.current_x = target_x
                self.positions[crane.crane_id] = target_x
                self.active_moves.pop(crane.crane_id, None)
                self.notify_position_change()
                return

            signature = ("|".join(blocking_crane_ids), target_station_id, round(target_x, 3))
            if self.last_wait_signature.get(crane.crane_id) != signature:
                self.last_wait_signature[crane.crane_id] = signature
                self.wait_started_at.setdefault(crane.crane_id, float(self.env.now))
                blocker_labels = ", ".join(crane_label(blocker) for blocker in blocking_crane_ids)
                crane.plant.log_event(
                    event_type="crane_wait",
                    message=(
                        f"{crane_label(crane.crane_id)}, "
                        f"{crane.plant.describe_target_location(target_station_id, target_x)} yonune giderken "
                        f"{blocker_labels} yuzunden bekledi."
                    ),
                    crane_id=crane.crane_id,
                    destination_station_id=target_station_id,
                    metadata={
                        "blockage_kind": "rail_neighbor",
                        "blocking_crane": blocker_labels,
                    },
                )
            waiter = self.env.event()
            self.waiters.append(waiter)
            yield waiter


class Crane:
    def __init__(
        self,
        env: simpy.Environment,
        crane_id: str,
        speed: CraneSpeedProfile,
        plant: "AnodizingPlant",
        home_x: float,
    ) -> None:
        self.env = env
        self.crane_id = crane_id
        self.speed = speed
        self.plant = plant
        self.home_x = home_x
        self.current_x = home_x
        self.busy_seconds = 0.0
        self.completed_moves = 0
        self.current_task_priority: int | None = None

    def run(self):
        while True:
            task = yield self.env.process(self.plant.dispatcher.wait_for_next_task(self))
            start = self.env.now
            self.current_task_priority = task.effective_priority
            try:
                yield self.env.process(self.handle_task(task))
            finally:
                self.current_task_priority = None
            self.busy_seconds += self.env.now - start
            self.completed_moves += 1

    def handle_task(self, task: TransportTask):
        job = task.job
        source_definition = self.plant.station_definitions[task.source_station_id]
        source_x = source_definition.x_m
        next_step = job.next_step
        if next_step is None:
            raise DataContractError(f"{job.job_id} icin bir sonraki adim yok.")

        if task.pickup_crane and task.pickup_crane != self.crane_id:
            self.plant.log_event(
                event_type="flex_claim",
                message=(
                    f"{crane_label(self.crane_id)}, {job.job_id} numarali "
                    f"({route_family_label(job.route_family)}, {job.micron} mikron) barayi "
                    f"{source_definition.display_name} ortak bolgesinden esneyerek devraldi. "
                    f"Ana sorumlu {crane_label(task.pickup_crane)} daha oncelikli isle meşguldü."
                ),
                job=job,
                crane_id=self.crane_id,
                source_station_id=source_definition.station_id,
                metadata={"blocking_crane": crane_label(task.pickup_crane)},
            )

        if self.current_x != source_x:
            yield self.env.process(
                self.plant.rail_controller.move(
                    self,
                    source_x,
                    self.speed.empty_mps,
                    target_station_id=source_definition.station_id,
                )
            )

        reserved_destination: tuple[str, simpy.ResourceRequest] | None = None
        if next_step.key not in {"exit", "drip"}:
            candidate_ids = self.plant.resolve_crane_destination_candidates(
                self.crane_id,
                job,
                next_step.key,
                source_definition.station_id,
            )
            reserved_destination = self.plant.try_reserve_station_now(candidate_ids, source_definition.x_m)
            if reserved_destination is None:
                self.plant.dispatcher.requeue(task)
                self.plant.log_task_blockage(task)
                return

        yield self.env.timeout(self.plant.sample_lift_seconds())

        if source_definition.is_virtual:
            if source_definition.station_id == "entry_conveyor":
                self.plant.entry_conveyor_busy = False
                if not self.plant.entry_conveyor_available.triggered:
                    self.plant.entry_conveyor_available.succeed(True)
        else:
            self.plant.physical_stations[source_definition.station_id].release_job_for_pickup()

        self.plant.log_event(
            event_type="pickup",
            message=(
                f"{crane_label(self.crane_id)}, {job.job_id} numarali "
                f"({route_family_label(job.route_family)}, {job.micron} mikron) barayi "
                f"{next_step.label} adimina goturmek uzere {source_definition.display_name} noktasindan aldi."
            ),
            job=job,
            crane_id=self.crane_id,
            source_station_id=source_definition.station_id,
        )

        yield self.env.process(self._move_loaded_job(job, source_definition, reserved_destination))

    def _retreat_to_home(self):
        if abs(self.current_x - self.home_x) <= 1e-9:
            return
        yield self.env.process(
            self.plant.rail_controller.move(self, self.home_x, self.speed.empty_mps)
        )

    def _clear_handoff_path(self, standby_x: float):
        if abs(self.current_x - standby_x) <= 1e-9:
            return
        yield self.env.process(
            self.plant.rail_controller.move(self, standby_x, self.speed.empty_mps)
        )

    def _move_loaded_job(
        self,
        job: Job,
        source_definition: StationDefinition,
        reserved_destination: tuple[str, simpy.ResourceRequest] | None = None,
    ):
        next_step = job.next_step
        if next_step is None:
            raise DataContractError(f"{job.job_id} icin bir sonraki adim yok.")

        if next_step.key in RINSE_STEP_KEYS:
            yield self.env.process(self._handle_rinse_step(job, source_definition, reserved_destination))
            return

        if next_step.key == "drip":
            yield self.env.process(self._handle_drip_and_oven(job, source_definition))
            return

        if next_step.key == "exit":
            dest_id = "exit_conveyor"
            dest_x = self.plant.station_definitions[dest_id].x_m
            yield self.env.process(
                self.plant.rail_controller.move(
                    self,
                    dest_x,
                    self.speed.loaded_mps,
                    target_station_id=dest_id,
                )
            )
            yield self.env.timeout(self.plant.sample_drop_seconds())
            job.next_step_index += 1
            job.current_station_id = dest_id
            self.plant.log_event(
                event_type="drop",
                message=(
                    f"{crane_label(self.crane_id)}, {job.job_id} numarali "
                    f"({route_family_label(job.route_family)}, {job.micron} mikron) barayi cikis konveyorune birakti."
                ),
                job=job,
                crane_id=self.crane_id,
                source_station_id=source_definition.station_id,
                destination_station_id=dest_id,
            )
            self.plant.complete_job(job)
            return

        if reserved_destination is None:
            candidate_ids = self.plant.resolve_crane_destination_candidates(
                self.crane_id,
                job,
                next_step.key,
                source_definition.station_id,
            )
            destination_id, destination_request = yield self.env.process(
                self.plant.request_first_available_station(candidate_ids, source_definition.x_m)
            )
        else:
            destination_id, destination_request = reserved_destination
        dest_x = self.plant.station_definitions[destination_id].x_m
        process_seconds = self.plant.sample_process_seconds(
            self.plant.station_definitions[destination_id].process_group,
            job,
            destination_id,
            step_key=next_step.key,
        )
        yield self.env.process(
            self.plant.rail_controller.move(
                self,
                dest_x,
                self.speed.loaded_mps,
                target_station_id=destination_id,
            )
        )
        yield self.env.timeout(self.plant.sample_drop_seconds())

        self.plant.record_step_duration(job, next_step.key, process_seconds)
        job.next_step_index += 1
        self.plant.log_event(
            event_type="drop",
            message=(
                f"{crane_label(self.crane_id)}, {job.job_id} numarali "
                f"({route_family_label(job.route_family)}, {job.micron} mikron) barayi "
                f"{self.plant.station_definitions[destination_id].display_name} istasyonuna birakti. "
                f"Burada {format_seconds_as_hms(process_seconds)} ({process_seconds:.1f} sn) kalacak."
            ),
            job=job,
            crane_id=self.crane_id,
            source_station_id=source_definition.station_id,
            destination_station_id=destination_id,
            duration_seconds=process_seconds,
        )
        self.plant.start_station_visit(
            job=job,
            step_key=next_step.key,
            step_label=next_step.label,
            station_id=destination_id,
            nominal_seconds=process_seconds,
        )
        self.plant.physical_stations[destination_id].accept_job(job, destination_request, process_seconds)
        if self.plant.station_definitions[destination_id].pickup_crane not in {None, self.crane_id}:
            yield self.env.process(self._clear_handoff_path(source_definition.x_m))

    def _handle_rinse_step(
        self,
        job: Job,
        source_definition: StationDefinition,
        reserved_destination: tuple[str, simpy.ResourceRequest] | None = None,
    ):
        next_step = job.next_step
        if next_step is None:
            raise DataContractError(f"{job.job_id} icin durulama adimi bulunamadi.")

        if reserved_destination is None:
            candidate_ids = self.plant.resolve_crane_destination_candidates(
                self.crane_id,
                job,
                next_step.key,
                source_definition.station_id,
            )
            destination_id, destination_request = yield self.env.process(
                self.plant.request_first_available_station(candidate_ids, source_definition.x_m)
            )
        else:
            destination_id, destination_request = reserved_destination
        resolved_step_index, resolved_step = self.plant.resolve_rinse_destination_step(
            job,
            source_definition.station_id,
            destination_id,
        )
        immediate_candidates = self.plant.resolve_destination_candidates(
            job,
            next_step.key,
            source_definition.station_id,
        )
        if destination_id not in immediate_candidates and destination_id != "durulama_e":
            self.plant.log_event(
                event_type="forward_rinse_skip",
                message=(
                    f"{crane_label(self.crane_id)}, {job.job_id} numarali "
                    f"({route_family_label(job.route_family)}, {job.micron} mikron) barayi "
                    f"{source_definition.display_name} cikisinda, standart durulama dolu oldugu icin "
                    f"ilerideki {self.plant.station_definitions[destination_id].display_name} tamponuna yonlendirdi."
                ),
                job=job,
                crane_id=self.crane_id,
                source_station_id=source_definition.station_id,
                destination_station_id=destination_id,
            )
        if destination_id == "durulama_e":
            dest_x = self.plant.station_definitions[destination_id].x_m
            yield self.env.process(
                self.plant.rail_controller.move(
                    self,
                    dest_x,
                    self.speed.loaded_mps,
                    target_station_id=destination_id,
                )
            )
            yield self.env.timeout(self.plant.sample_drop_seconds())
            station = self.plant.physical_stations[destination_id]
            station.occupy_without_process(job, destination_request)
            self.plant.start_station_visit(
                job=job,
                step_key="aux_post_eloksal_buffer",
                step_label="Durulama E Tampon",
                station_id=destination_id,
                nominal_seconds=0.0,
            )
            self.plant.log_event(
                event_type="aux_buffer_drop",
                message=(
                    f"{crane_label(self.crane_id)}, {job.job_id} numarali "
                    f"({route_family_label(job.route_family)}, {job.micron} mikron) barayi "
                    f"ana durulamalar dolu oldugu icin {self.plant.station_definitions[destination_id].display_name} "
                    f"yardimci tamponuna birakti."
                ),
                job=job,
                crane_id=self.crane_id,
                source_station_id=source_definition.station_id,
                destination_station_id=destination_id,
            )
            self.plant.publish_transport_task(job=job, source_station_id=destination_id, priority=2)
            return
        dest_x = self.plant.station_definitions[destination_id].x_m
        rinse_seconds = self.plant.sample_process_seconds(
            self.plant.station_definitions[destination_id].process_group,
            job,
            destination_id,
            step_key=resolved_step.key,
        )
        yield self.env.process(
            self.plant.rail_controller.move(
                self,
                dest_x,
                self.speed.loaded_mps,
                target_station_id=destination_id,
            )
        )
        yield self.env.timeout(self.plant.sample_drop_seconds())

        station = self.plant.physical_stations[destination_id]
        station.occupy_without_process(job, destination_request)
        self.plant.record_step_duration(job, resolved_step.key, rinse_seconds)
        self.plant.log_event(
            event_type="drop",
            message=(
                f"{crane_label(self.crane_id)}, {job.job_id} numarali "
                f"({route_family_label(job.route_family)}, {job.micron} mikron) barayi "
                f"{self.plant.station_definitions[destination_id].display_name} durulamasina indirdi. "
                f"Normal daldir-cikar bekleme suresi {format_seconds_as_hms(rinse_seconds)} ({rinse_seconds:.1f} sn)."
            ),
            job=job,
            crane_id=self.crane_id,
            source_station_id=source_definition.station_id,
            destination_station_id=destination_id,
            duration_seconds=rinse_seconds,
        )
        self.plant.start_station_visit(
            job=job,
            step_key=resolved_step.key,
            step_label=resolved_step.label,
            station_id=destination_id,
            nominal_seconds=rinse_seconds,
        )
        yield self.env.timeout(rinse_seconds)

        job.next_step_index = resolved_step_index + 1
        if (
            self.plant.station_definitions[destination_id].pickup_crane != self.crane_id
            or self.plant.has_higher_priority_waiting_task(self.crane_id, threshold_priority=2)
            or not self.plant.can_place_next_step(job, destination_id)
        ):
            self.plant.publish_transport_task(
                job=job,
                source_station_id=destination_id,
                priority=2,
            )
            self.plant.log_event(
                event_type="rinse_buffered",
                message=(
                    f"{crane_label(self.crane_id)}, daha onemli bir is ciktigi icin "
                    f"{job.job_id} numarali ({route_family_label(job.route_family)}, {job.micron} mikron) barayi "
                    f"{self.plant.station_definitions[destination_id].display_name} durulamasinda tampon olarak birakti."
                ),
                job=job,
                crane_id=self.crane_id,
                source_station_id=destination_id,
            )
            if self.plant.station_definitions[destination_id].pickup_crane not in {None, self.crane_id}:
                yield self.env.process(self._clear_handoff_path(source_definition.x_m))
            return

        next_reservation: tuple[str, simpy.ResourceRequest] | None = None
        upcoming_step = job.next_step
        if upcoming_step is not None and upcoming_step.key not in {"exit", "drip"}:
            candidate_ids = self.plant.resolve_crane_destination_candidates(
                self.crane_id,
                job,
                upcoming_step.key,
                destination_id,
            )
            next_reservation = self.plant.try_reserve_station_now(candidate_ids, dest_x)
            if next_reservation is None:
                self.plant.publish_transport_task(
                    job=job,
                    source_station_id=destination_id,
                    priority=2,
                )
                self.plant.log_event(
                    event_type="rinse_buffered",
                    message=(
                        f"{crane_label(self.crane_id)}, sonraki havuz simdilik musait olmadigi icin "
                        f"{job.job_id} numarali ({route_family_label(job.route_family)}, {job.micron} mikron) barayi "
                        f"{self.plant.station_definitions[destination_id].display_name} durulamasinda tampon olarak birakti."
                    ),
                    job=job,
                    crane_id=self.crane_id,
                    source_station_id=destination_id,
                )
                if self.plant.station_definitions[destination_id].pickup_crane not in {None, self.crane_id}:
                    yield self.env.process(self._clear_handoff_path(source_definition.x_m))
                return

        yield self.env.timeout(self.plant.sample_lift_seconds())
        station.release_job_for_pickup()
        self.plant.log_event(
            event_type="rinse_lift",
            message=(
                f"{crane_label(self.crane_id)}, {job.job_id} numarali "
                f"({route_family_label(job.route_family)}, {job.micron} mikron) barayi "
                f"{self.plant.station_definitions[destination_id].display_name} durulamasindan kaldirdi."
            ),
            job=job,
            crane_id=self.crane_id,
            source_station_id=destination_id,
        )
        yield self.env.process(
            self._move_loaded_job(
                job,
                self.plant.station_definitions[destination_id],
                next_reservation,
            )
        )

    def _handle_drip_and_oven(self, job: Job, source_definition: StationDefinition):
        source_x = source_definition.x_m
        drip_x = self.plant.drip_hold_x_m
        drip_seconds = self.plant.sample_drip_seconds(job)

        yield self.env.process(self.plant.rail_controller.move(self, drip_x, self.speed.loaded_mps))
        self.plant.record_step_duration(job, "drip", drip_seconds)
        self.plant.log_event(
            event_type="drip_start",
            message=(
                f"{crane_label(self.crane_id)}, {job.job_id} numarali "
                f"({route_family_label(job.route_family)}, {job.micron} mikron) barayi "
                f"havadaki suzme adiminda tutmaya basladi. Suzme suresi {format_seconds_as_hms(drip_seconds)} "
                f"({drip_seconds:.1f} sn)."
            ),
            job=job,
            crane_id=self.crane_id,
            source_station_id=source_definition.station_id,
            duration_seconds=drip_seconds,
        )
        yield self.env.timeout(drip_seconds)

        oven_candidates = self.plant.resolve_destination_candidates(job, "firin", source_definition.station_id)
        oven_id, oven_request = yield self.env.process(
            self.plant.request_first_available_station(oven_candidates, self.current_x)
        )
        oven_x = self.plant.station_definitions[oven_id].x_m
        oven_process_seconds = self.plant.sample_process_seconds(
            self.plant.station_definitions[oven_id].process_group,
            job,
            oven_id,
            step_key="firin",
        )
        yield self.env.process(
            self.plant.rail_controller.move(
                self,
                oven_x,
                self.speed.loaded_mps,
                target_station_id=oven_id,
            )
        )
        yield self.env.timeout(self.plant.sample_drop_seconds())

        self.plant.record_step_duration(job, "firin", oven_process_seconds)
        job.next_step_index += 2
        self.plant.log_event(
            event_type="drop",
            message=(
                f"{crane_label(self.crane_id)}, {job.job_id} numarali "
                f"({route_family_label(job.route_family)}, {job.micron} mikron) barayi "
                f"{self.plant.station_definitions[oven_id].display_name} istasyonuna birakti. "
                f"Burada {format_seconds_as_hms(oven_process_seconds)} ({oven_process_seconds:.1f} sn) kalacak."
            ),
            job=job,
            crane_id=self.crane_id,
            source_station_id=source_definition.station_id,
            destination_station_id=oven_id,
            duration_seconds=oven_process_seconds,
        )
        self.plant.start_station_visit(
            job=job,
            step_key="firin",
            step_label="Firin",
            station_id=oven_id,
            nominal_seconds=oven_process_seconds,
        )
        self.plant.physical_stations[oven_id].accept_job(job, oven_request, oven_process_seconds)


class AnodizingPlant:
    def __init__(self, env: simpy.Environment, data: DataBundle, seed: int = 42) -> None:
        self.env = env
        self.data = data
        self.rng = np.random.default_rng(seed)
        self.dispatcher = Dispatcher(env, data.station_definitions)
        self.station_definitions = data.station_definitions
        self.physical_stations = {
            station_id: PhysicalStation(env, definition, self)
            for station_id, definition in data.station_definitions.items()
            if not definition.is_virtual and not definition.is_sink
        }
        self.drip_hold_x_m = data.station_positions["suzme_2"]
        crane_home_positions = {
            "vinc1": data.station_positions["buffer_2"],
            "vinc2": data.station_positions["eloksal_4"],
            "vinc3": data.station_positions["kalay_1"],
            "vinc4": data.station_positions["firin_2"],
        }
        crane_zone_bounds = {
            "vinc1": (data.station_positions["entry_conveyor"], data.station_positions["notralizasyon_2"]),
            "vinc2": (data.station_positions["durulama_1"], data.station_positions["durulama_5"]),
            "vinc3": (data.station_positions["durulama_3"], data.station_positions["sicaksu_3"]),
            "vinc4": (data.station_positions["sicaksu_1"], data.station_positions["exit_conveyor"]),
        }
        self.rail_controller = RailController(
            env=env,
            crane_order=["vinc1", "vinc2", "vinc3", "vinc4"],
            initial_positions=crane_home_positions,
            zone_bounds=crane_zone_bounds,
        )
        self.rail_controller.on_position_change = self.dispatcher.wake_all
        self.cranes = {
            crane_id: Crane(env, crane_id, speed_profile, self, crane_home_positions[crane_id])
            for crane_id, speed_profile in data.crane_speeds.items()
        }
        self.task_sequence = 0
        self.job_sequence = 0
        self.active_jobs = 0
        self.generator_finished = False
        self.all_done = env.event()
        self.system_idle_event = env.event()
        self.system_idle_event.succeed(True)
        self.completed_jobs: list[Job] = []
        self.arrived_jobs: list[Job] = []
        self.event_log: list[dict[str, Any]] = []
        self.entry_conveyor_busy = False
        self.entry_conveyor_available = env.event()
        self.entry_conveyor_available.succeed(True)
        self.station_visit_sequence = 0
        self.station_visit_records: list[StationVisitRecord] = []
        self.station_visit_lookup: dict[int, StationVisitRecord] = {}

        for crane in self.cranes.values():
            env.process(crane.run())

    def log_event(
        self,
        event_type: str,
        message: str,
        job: Job | None = None,
        crane_id: str | None = None,
        source_station_id: str | None = None,
        destination_station_id: str | None = None,
        duration_seconds: float | None = None,
        metadata: dict[str, Any] | None = None,
    ) -> None:
        row = {
            "time_seconds": float(self.env.now),
            "time_hms": format_seconds_as_hms(float(self.env.now)),
            "event_type": event_type,
            "message": message,
            "bara_no": None if job is None else job.job_id,
            "route_family": None if job is None else job.route_family,
            "cesit": None if job is None else route_family_label(job.route_family),
            "mikron": None if job is None else job.micron,
            "crane_id": crane_id,
            "vinc": crane_label(crane_id),
            "source_station_id": source_station_id,
            "kaynak": "" if source_station_id is None else self.station_definitions[source_station_id].display_name,
            "destination_station_id": destination_station_id,
            "hedef": "" if destination_station_id is None else self.station_definitions[destination_station_id].display_name,
            "sure_sn": None if duration_seconds is None else float(duration_seconds),
            "sure_hms": format_seconds_as_hms(duration_seconds),
            "blockage_kind": "",
            "blocking_crane": "",
            "next_step_label": "",
            "blocking_station_names": "",
            "blocking_station_ids": "",
            "blocking_job_ids": "",
        }
        if metadata:
            row.update(metadata)
        self.event_log.append(row)

    def initialize_job_report(self, job: Job) -> None:
        source_parts = []
        if job.source_date_text:
            source_parts.append(job.source_date_text)
        if job.source_start_text:
            source_parts.append(job.source_start_text)
        if job.source_bara_no:
            source_parts.append(f"Kaynak Bara {job.source_bara_no}")
        job.report_row = {
            "Bara No.": job.job_id,
            "Çeşit": route_family_label(job.route_family),
            "Mikron": job.micron,
            "Başlangıç": format_seconds_as_hms(job.created_at),
            "Açıklama": " | ".join(source_parts) if source_parts else "Simülasyon çıktısı",
        }

    def _column_for_step_occurrence(self, step_key: str, occurrence: int) -> str | None:
        direct = {
            "yagalma": "YağAlma",
            "durulama_y": "Durulama",
            "asitmat": "AsitMat",
            "durulama_a": "Durulama.1",
            "sokme": "Sökme",
            "kostik": "Kostik",
            "durulama_kostik": "Durulama V1",
            "notralizasyon": "Nötralizasyon",
            "eloksal": "Eloksal",
            "post_eloksal_rinse": "Durulama.2",
            "kalay": "Kalay / Nikel (Renk)",
            "nikel": "Kalay / Nikel (Renk)",
            "post_color_rinse": "Durulama.3",
            "di": "DI",
            "tespit": "Tespit",
            "post_tespit_rinse": "Durulama.4",
            "sicaksu": "SıcakSu",
            "drip": "Süzme",
            "firin": "Fırın",
        }
        if step_key in direct:
            return direct[step_key]
        if step_key == "pre_neutral_rinse":
            return "Durulama V2" if occurrence == 0 else "Durulama V3"
        return None

    def _column_for_step(self, job: Job, step_key: str) -> str | None:
        occurrence = job.step_occurrences.get(step_key, 0)
        return self._column_for_step_occurrence(step_key, occurrence)

    def record_step_duration(self, job: Job, step_key: str, process_seconds: float) -> None:
        column = self._column_for_step(job, step_key)
        job.step_occurrences[step_key] = job.step_occurrences.get(step_key, 0) + 1
        if column:
            job.report_row[column] = format_seconds_as_hms(process_seconds)

    def start_station_visit(
        self,
        job: Job,
        step_key: str,
        step_label: str,
        station_id: str,
        nominal_seconds: float,
    ) -> int:
        self.station_visit_sequence += 1
        record = StationVisitRecord(
            visit_id=self.station_visit_sequence,
            job_id=job.job_id,
            route_family=job.route_family,
            micron=job.micron,
            step_key=step_key,
            step_label=step_label,
            station_id=station_id,
            station_name=self.station_definitions[station_id].display_name,
            nominal_seconds=float(nominal_seconds),
            entered_at=float(self.env.now),
        )
        self.station_visit_records.append(record)
        self.station_visit_lookup[record.visit_id] = record
        self.physical_stations[station_id].current_visit_id = record.visit_id
        return record.visit_id

    def finish_station_visit(self, visit_id: int, exited_at: float) -> None:
        record = self.station_visit_lookup.get(visit_id)
        if record is None:
            return
        if record.exited_at is None:
            record.exited_at = float(exited_at)

    def sample_lift_seconds(self) -> float:
        return float(self.rng.uniform(20.0, 30.0))

    def sample_drop_seconds(self) -> float:
        return float(self.rng.uniform(10.0, 15.0))

    def sample_drip_seconds(self, job: Job | None = None) -> float:
        if job is not None:
            replay_value = job.replay_step_seconds.get(("drip", 0))
            if replay_value is not None:
                return ensure_positive_seconds(replay_value, "suzme-replay")
        spec = self.data.pool_distributions["suzme"]
        return ensure_positive_seconds(spec.sample_seconds(self.rng, "suzme"), "suzme")

    def sample_process_seconds(
        self,
        process_group: str | None,
        job: Job,
        station_id: str,
        step_key: str | None = None,
    ) -> float:
        logical_process_group = process_group
        if step_key is not None:
            occurrence = job.step_occurrences.get(step_key, 0)
            logical_process_group = {
                "yagalma": "yagalma",
                "durulama_y": "durulama_y",
                "asitmat": "asitmat",
                "durulama_a": "durulama_a",
                "sokme": "sokme",
                "kostik": "kostik",
                "durulama_kostik": "durulama_kostik",
                "notralizasyon": "notralizasyon",
                "eloksal": "eloksal",
                "post_eloksal_rinse": "durulama_post_eloksal",
                "kalay": "renk",
                "nikel": "renk",
                "post_color_rinse": "durulama_post_renk",
                "di": "di",
                "tespit": "tespit",
                "post_tespit_rinse": "durulama_post_tespit",
                "sicaksu": "sicaksu",
                "drip": "suzme",
                "firin": "firin",
            }.get(step_key, process_group)
            if step_key == "pre_neutral_rinse":
                logical_process_group = "durulama_1" if occurrence <= 0 else "durulama_2"
        if logical_process_group is None:
            raise DataContractError(f"{station_id} icin islem grubu tanimsiz.")
        if step_key is not None:
            occurrence = job.step_occurrences.get(step_key, 0)
            replay_value = job.replay_step_seconds.get((step_key, occurrence))
            if replay_value is not None:
                return ensure_positive_seconds(replay_value, f"{step_key}-replay")
        if logical_process_group == "eloksal":
            spec = self.data.eloxal_distributions_by_micron[job.micron]
            return ensure_positive_seconds(spec.sample_seconds(self.rng, f"eloksal-{job.micron}"), "eloksal")
        if logical_process_group == "tespit":
            return float(job.micron * 60)
        spec = self.data.pool_distributions[logical_process_group]
        return ensure_positive_seconds(
            spec.sample_seconds(self.rng, logical_process_group),
            logical_process_group,
        )

    def publish_transport_task(self, job: Job, source_station_id: str, priority: int) -> TransportTask:
        station_definition = self.station_definitions[source_station_id]
        eligible_cranes = tuple(
            crane_id
            for crane_id in (station_definition.pickup_crane, *station_definition.flex_cranes)
            if crane_id
        )
        self.task_sequence += 1
        task = TransportTask(
            task_id=self.task_sequence,
            job=job,
            source_station_id=source_station_id,
            source_group=station_definition.process_group or "unknown",
            pickup_crane=station_definition.pickup_crane or "",
            eligible_cranes=eligible_cranes,
            ready_time=self.env.now,
            created_sequence=self.task_sequence,
            priority=priority,
            timeout_seconds=None,
            timeout_deadline=None,
        )
        self.dispatcher.publish(task)
        if not self.is_task_dispatchable(task):
            self.log_task_blockage(task)
        return task

    def has_higher_priority_waiting_task(self, crane_id: str, threshold_priority: int) -> bool:
        return any(
            task.pickup_crane == crane_id and task.effective_priority < threshold_priority
            for task in self.dispatcher.pending.values()
        )

    def crane_zone_allows_station(self, crane_id: str, station_id: str) -> bool:
        zone_min, zone_max = self.rail_controller.zone_bounds[crane_id]
        station_x = self.station_definitions[station_id].x_m
        return zone_min - 1e-9 <= station_x <= zone_max + 1e-9

    def station_is_immediately_available(self, station_id: str) -> bool:
        station = self.physical_stations[station_id]
        return (
            station.current_job is None
            and station.resource.count == 0
            and len(station.resource.queue) == 0
        )

    def try_reserve_station_now(
        self,
        candidate_ids: Iterable[str],
        reference_x: float,
    ) -> tuple[str, simpy.ResourceRequest] | None:
        immediate = [
            station_id
            for station_id in candidate_ids
            if self.station_is_immediately_available(station_id)
        ]
        if not immediate:
            return None

        chosen = min(
            immediate,
            key=lambda station_id: abs(self.station_definitions[station_id].x_m - reference_x),
        )
        request = self.physical_stations[chosen].resource.request()
        if not request.triggered:
            request.cancel()
            return None
        return chosen, request

    def resolve_effective_destination_candidates(
        self,
        job: Job,
        step_key: str,
        current_station_id: str,
    ) -> list[str]:
        candidates = self.resolve_destination_candidates(job, step_key, current_station_id)
        if step_key in RINSE_STEP_KEYS:
            candidates = self._dedupe_station_ids(
                self._resolve_forward_rinse_skip_candidates(job, current_station_id) + candidates
            )
        if step_key != "post_eloksal_rinse":
            return candidates

        if current_station_id == "durulama_e":
            return candidates

        if any(self.station_is_immediately_available(station_id) for station_id in candidates):
            return candidates

        if self.station_is_immediately_available("durulama_e"):
            return ["durulama_e"]

        return candidates + ["durulama_e"]

    def resolve_crane_destination_candidates(
        self,
        crane_id: str,
        job: Job,
        step_key: str,
        current_station_id: str,
    ) -> list[str]:
        return [
            station_id
            for station_id in self.resolve_effective_destination_candidates(job, step_key, current_station_id)
            if self.crane_zone_allows_station(crane_id, station_id)
        ]

    def estimate_task_service_seconds(self, crane: "Crane", task: TransportTask) -> float:
        source_x = self.station_definitions[task.source_station_id].x_m
        next_step = task.job.next_step
        if next_step is None:
            return 0.0
        empty_travel = abs(crane.current_x - source_x) / max(crane.speed.empty_mps, 1e-9)
        if next_step.key == "exit":
            destination_ids = ["exit_conveyor"]
        elif next_step.key == "drip":
            destination_ids = ["firin_1", "firin_2"]
        else:
            destination_ids = self.resolve_crane_destination_candidates(
                crane.crane_id,
                task.job,
                next_step.key,
                task.source_station_id,
            )
        if not destination_ids:
            return float("inf")
        destination_x = min(
            (self.station_definitions[station_id].x_m for station_id in destination_ids),
            key=lambda x_value: abs(x_value - source_x),
        )
        loaded_travel = abs(destination_x - source_x) / max(crane.speed.loaded_mps, 1e-9)
        return empty_travel + 30.0 + loaded_travel + 15.0

    def has_pending_primary_zone_task(self, crane_id: str) -> bool:
        return any(
            self.station_definitions[task.source_station_id].pickup_crane == crane_id
            for task in self.dispatcher.pending.values()
            if not task.claimed
        )

    def has_imminent_primary_zone_signal(self, crane_id: str, horizon_seconds: float) -> bool:
        deadline = self.env.now + horizon_seconds
        for station in self.physical_stations.values():
            if station.definition.pickup_crane != crane_id:
                continue
            if station.current_job is None or station.current_task_id is not None:
                continue
            complete_at = station.current_process_complete_at
            if complete_at is not None and complete_at <= deadline + 1e-9:
                return True
        return False

    def should_crane_clear_overlap(self, crane: "Crane") -> bool:
        return any(
            abs(self.station_definitions[station_id].x_m - crane.current_x) <= 1e-9
            for station_id in SHARED_OVERLAP_STATIONS
        )

    def _destination_candidates_for_step(
        self,
        job: Job,
        step_key: str,
        current_station_id: str,
        occurrence_override: int | None = None,
    ) -> list[str]:
        single = {
            "durulama_y": ["durulama_y"],
            "asitmat": ["asitmat"],
            "durulama_a": ["durulama_a"],
            "sokme": ["sokme"],
            "kostik": ["kostik"],
            "durulama_kostik": ["durulama_kostik"],
            "notralizasyon": ["notralizasyon_1", "notralizasyon_2"],
            "eloksal": ["eloksal_1", "eloksal_2", "eloksal_3", "eloksal_4", "eloksal_5"],
            "post_eloksal_rinse": ["durulama_3", "durulama_4", "durulama_5"],
            "di": ["di"],
            "tespit": ["tespit_1", "tespit_2", "tespit_3"],
            "post_tespit_rinse": ["durulama_t"],
            "sicaksu": ["sicaksu_1", "sicaksu_2", "sicaksu_3"],
            "firin": ["firin_1", "firin_2"],
            "yagalma": ["yagalma_1", "yagalma_2"],
        }
        if step_key in single:
            return list(single[step_key])
        if step_key == "pre_neutral_rinse":
            occurrence = job.step_occurrences.get(step_key, 0) if occurrence_override is None else occurrence_override
            if occurrence <= 0:
                return ["durulama_1"]
            return ["durulama_2"]
        if step_key == "kalay":
            return ["kalay_1", "kalay_2"]
        if step_key == "nikel":
            return ["nikel"]
        if step_key == "post_color_rinse":
            return ["durulama_kalay"] if job.route_family == "man" else ["durulama_nikel"]
        if step_key == "exit":
            return ["exit_conveyor"]
        raise DataContractError(f"Destinasyon cozumlenemedi: {step_key}")

    def _dedupe_station_ids(self, station_ids: Iterable[str]) -> list[str]:
        ordered: list[str] = []
        seen: set[str] = set()
        for station_id in station_ids:
            if station_id in seen:
                continue
            seen.add(station_id)
            ordered.append(station_id)
        return ordered

    def _resolve_forward_rinse_skip_candidates(
        self,
        job: Job,
        current_station_id: str,
    ) -> list[str]:
        source_group = self.station_definitions[current_station_id].process_group
        if source_group not in ACTIVE_PRIORITY_GROUPS:
            return []

        extra_candidates: list[str] = []
        virtual_occurrences: defaultdict[str, int] = defaultdict(int)
        upcoming_steps = job.route[job.next_step_index :]
        for index, step in enumerate(upcoming_steps):
            if index > 0 and step.key not in PASSIVE_FORWARD_STEP_KEYS:
                break
            if step.key in RINSE_STEP_KEYS:
                occurrence = job.step_occurrences.get(step.key, 0) + virtual_occurrences[step.key]
                extra_candidates.extend(
                    self._destination_candidates_for_step(
                        job,
                        step.key,
                        current_station_id,
                        occurrence_override=occurrence,
                    )
                )
            virtual_occurrences[step.key] += 1

        if (
            job.next_step is not None
            and job.next_step.key == "post_eloksal_rinse"
            and job.route_family in {"parlak", "natural", "asitmat"}
        ):
            extra_candidates.extend(["durulama_kalay", "durulama_nikel"])

        return self._dedupe_station_ids(extra_candidates)

    def resolve_rinse_destination_step(
        self,
        job: Job,
        current_station_id: str,
        destination_id: str,
    ) -> tuple[int, StepDefinition]:
        next_step = job.next_step
        if next_step is None:
            raise DataContractError("Durulama destinasyonu cozumlenirken sonraki adim bulunamadi.")

        matched_index = job.next_step_index
        virtual_occurrences: defaultdict[str, int] = defaultdict(int)
        upcoming_steps = job.route[job.next_step_index :]
        for relative_index, step in enumerate(upcoming_steps):
            if relative_index > 0 and step.key not in PASSIVE_FORWARD_STEP_KEYS:
                break

            if step.key == "di":
                if destination_id == "di":
                    matched_index = job.next_step_index + relative_index
            elif step.key in RINSE_STEP_KEYS:
                occurrence = job.step_occurrences.get(step.key, 0) + virtual_occurrences[step.key]
                candidate_ids = self._destination_candidates_for_step(
                    job,
                    step.key,
                    current_station_id,
                    occurrence_override=occurrence,
                )
                if destination_id in candidate_ids:
                    matched_index = job.next_step_index + relative_index

            virtual_occurrences[step.key] += 1

        return matched_index, job.route[matched_index]

    def flex_claim_would_delay_primary(self, helper_crane_id: str, task: TransportTask) -> bool:
        station_definition = self.station_definitions[task.source_station_id]
        primary_crane_id = station_definition.pickup_crane
        if primary_crane_id is None or primary_crane_id == helper_crane_id:
            return False

        order = self.rail_controller.crane_order
        helper_idx = order.index(helper_crane_id)
        primary_idx = order.index(primary_crane_id)
        source_x = station_definition.x_m
        primary_x = self.rail_controller._projected_position(primary_crane_id)
        gap = self.rail_controller.safety_gap_m

        # Esneme sadece ana vincin overlap agzini kapatmayacaksa serbest.
        # Yardimci vinç primary'nin sagindaysa, primary overlap'e soldan yaklasir.
        if helper_idx > primary_idx:
            return primary_x <= source_x + gap
        # Yardimci vinç primary'nin solundaysa, primary overlap'e sagdan yaklasir.
        if helper_idx < primary_idx:
            return primary_x >= source_x - gap
        return False

    def can_crane_claim_task(self, crane: "Crane", task: TransportTask) -> bool:
        station_definition = self.station_definitions[task.source_station_id]
        if crane.crane_id == station_definition.pickup_crane:
            return True
        if crane.crane_id not in station_definition.flex_cranes:
            return False
        if task.source_station_id not in SHARED_FLEX_STATIONS:
            return False

        next_step = task.job.next_step
        if next_step is None:
            return False
        reachable_candidates = self.resolve_crane_destination_candidates(
            crane.crane_id,
            task.job,
            next_step.key,
            task.source_station_id,
        )
        if not reachable_candidates:
            return False
        if not any(self.station_is_immediately_available(station_id) for station_id in reachable_candidates):
            return False

        primary_crane_id = station_definition.pickup_crane
        if not primary_crane_id:
            return False
        primary_crane = self.cranes[primary_crane_id]
        primary_has_more_urgent_work = (
            primary_crane.current_task_priority is not None
            and primary_crane.current_task_priority < task.effective_priority
        ) or self.has_higher_priority_waiting_task(primary_crane_id, task.effective_priority)
        if not primary_has_more_urgent_work:
            return False
        if self.flex_claim_would_delay_primary(crane.crane_id, task):
            return False

        horizon_seconds = self.estimate_task_service_seconds(crane, task)
        if not math.isfinite(horizon_seconds):
            return False
        if self.has_pending_primary_zone_task(crane.crane_id):
            return False
        if self.has_imminent_primary_zone_signal(crane.crane_id, horizon_seconds):
            return False
        return True

    def describe_target_location(self, station_id: str | None, target_x: float) -> str:
        if station_id is not None and station_id in self.station_definitions:
            return self.station_definitions[station_id].display_name
        return f"x={target_x:.2f} m"

    def describe_task_blockage(self, task: TransportTask) -> tuple[str, str, dict[str, Any]] | None:
        next_step = task.job.next_step
        if next_step is None or next_step.key == "exit":
            return None

        if next_step.key == "drip":
            candidate_ids = ["firin_1", "firin_2"]
        else:
            candidate_ids = self.resolve_effective_destination_candidates(
                task.job,
                next_step.key,
                task.source_station_id,
            )

        blocking_parts: list[str] = []
        signature_parts: list[str] = []
        blocking_station_names: list[str] = []
        blocking_job_ids: list[str] = []
        for station_id in candidate_ids:
            station = self.physical_stations[station_id]
            blocker_job = station.current_job
            if blocker_job is None:
                return None
            signature_parts.append(f"{station_id}:{blocker_job.job_id}")
            blocking_station_names.append(self.station_definitions[station_id].display_name)
            blocking_job_ids.append(str(blocker_job.job_id))
            blocking_parts.append(
                f"{self.station_definitions[station_id].display_name} icinde "
                f"{blocker_job.job_id} numarali ({route_family_label(blocker_job.route_family)}, "
                f"{blocker_job.micron} mikron) bara var"
            )

        source_name = self.station_definitions[task.source_station_id].display_name
        message = (
            f"{task.job.job_id} numarali ({route_family_label(task.job.route_family)}, {task.job.micron} mikron) "
            f"bara {source_name} havuzundan cikamadi. "
            f"Siradaki {next_step.label} adimi icin "
            + "; ".join(blocking_parts)
            + "."
        )
        metadata = {
            "blockage_kind": "pool",
            "next_step_label": next_step.label,
            "blocking_station_names": "; ".join(blocking_station_names),
            "blocking_station_ids": "|".join(candidate_ids),
            "blocking_job_ids": "|".join(blocking_job_ids),
        }
        return "|".join(signature_parts), message, metadata

    def log_task_blockage(self, task: TransportTask) -> None:
        blockage = self.describe_task_blockage(task)
        if blockage is None:
            task.blocked_signature = None
            return
        signature, message, metadata = blockage
        if task.blocked_signature == signature:
            return
        task.blocked_signature = signature
        self.log_event(
            event_type="task_blocked",
            message=message,
            job=task.job,
            crane_id=task.pickup_crane,
            source_station_id=task.source_station_id,
            metadata=metadata,
        )

    def can_place_next_step(self, job: Job, source_station_id: str) -> bool:
        next_step = job.next_step
        if next_step is None:
            return False
        if next_step.key == "exit":
            return True
        if next_step.key == "drip":
            return any(self.station_is_immediately_available(station_id) for station_id in ["firin_1", "firin_2"])
        candidates = self.resolve_effective_destination_candidates(job, next_step.key, source_station_id)
        return any(self.station_is_immediately_available(station_id) for station_id in candidates)

    def resolve_destination_candidates(self, job: Job, step_key: str, current_station_id: str) -> list[str]:
        return self._destination_candidates_for_step(job, step_key, current_station_id)

    def is_task_dispatchable(self, task: TransportTask, crane: "Crane" | None = None) -> bool:
        next_step = task.job.next_step
        if next_step is None:
            return False
        if next_step.key == "exit":
            if crane is None:
                return True
            source_x = self.station_definitions[task.source_station_id].x_m
            target_x = self.station_definitions["exit_conveyor"].x_m
            return self.rail_controller.can_move_to(crane.crane_id, source_x) and self.rail_controller.can_move_to(
                crane.crane_id, target_x
            )
        if next_step.key == "drip":
            candidates = ["firin_1", "firin_2"]
        elif crane is None:
            candidates = self.resolve_effective_destination_candidates(
                task.job,
                next_step.key,
                task.source_station_id,
            )
        else:
            candidates = self.resolve_crane_destination_candidates(
                crane.crane_id,
                task.job,
                next_step.key,
                task.source_station_id,
            )

        available = [station_id for station_id in candidates if self.station_is_immediately_available(station_id)]
        if not available:
            return False
        if crane is None:
            return True

        source_x = self.station_definitions[task.source_station_id].x_m
        if not self.rail_controller.can_move_to(crane.crane_id, source_x):
            return False

        return any(
            self.rail_controller.can_move_to(crane.crane_id, self.station_definitions[station_id].x_m)
            for station_id in available
        )

    def request_first_available_station(self, candidate_ids: Iterable[str], reference_x: float):
        candidate_ids = list(candidate_ids)
        if not candidate_ids:
            raise DataContractError("Vinç için erişilebilir boş hedef havuz bulunamadi.")
        requests = {
            station_id: self.physical_stations[station_id].resource.request()
            for station_id in candidate_ids
        }
        result = yield self.env.any_of(list(requests.values()))
        triggered = [station_id for station_id, request in requests.items() if request in result]
        chosen = min(triggered, key=lambda station_id: abs(self.station_definitions[station_id].x_m - reference_x))

        for station_id, request in requests.items():
            if station_id == chosen:
                continue
            if request.triggered:
                self.physical_stations[station_id].resource.release(request)
            else:
                request.cancel()

        return chosen, requests[chosen]

    def handle_station_process_completion(self, station: PhysicalStation, job: Job, process_seconds: float) -> None:
        priority = 1 if station.definition.process_group in ACTIVE_PRIORITY_GROUPS else 2
        timeout_seconds = self.data.timeout_seconds_by_group.get(station.definition.process_group)
        task = self.publish_transport_task(job=job, source_station_id=station.definition.station_id, priority=priority)
        task.timeout_seconds = timeout_seconds
        task.timeout_deadline = None if timeout_seconds is None else self.env.now + timeout_seconds
        station.current_task_id = task.task_id
        self.log_event(
            event_type="process_completed",
            message=(
                f"{job.job_id} numarali ({route_family_label(job.route_family)}, {job.micron} mikron) baranin "
                f"{station.definition.display_name} istasyonundaki islemi tamamlandi."
            ),
            job=job,
            source_station_id=station.definition.station_id,
            duration_seconds=process_seconds,
        )
        if timeout_seconds is not None:
            self.env.process(self._timeout_watch(task.task_id, timeout_seconds))

    def _timeout_watch(self, task_id: int, timeout_seconds: float):
        yield self.env.timeout(timeout_seconds)
        self.dispatcher.escalate_timeout(task_id)

    def build_route(self, route_family: str) -> list[StepDefinition]:
        routes = {
            "parlak": [
                StepDefinition("yagalma", "Yag Alma"),
                StepDefinition("durulama_y", "Durulama"),
                StepDefinition("notralizasyon", "Notralizasyon"),
                StepDefinition("eloksal", "Eloksal"),
                StepDefinition("post_eloksal_rinse", "Eloksal Sonrasi Durulama"),
                StepDefinition("di", "DI"),
                StepDefinition("tespit", "Tespit"),
                StepDefinition("post_tespit_rinse", "Tespit Sonrasi Durulama"),
                StepDefinition("sicaksu", "Sicak Su"),
                StepDefinition("drip", "Suzme"),
                StepDefinition("firin", "Firin"),
                StepDefinition("exit", "Cikis Konveyor"),
            ],
            "sakem": [
                StepDefinition("yagalma", "Yag Alma"),
                StepDefinition("durulama_y", "Durulama"),
                StepDefinition("kostik", "Kostik"),
                StepDefinition("durulama_kostik", "Kostik Durulama"),
                StepDefinition("pre_neutral_rinse", "Genel Durulama"),
                StepDefinition("notralizasyon", "Notralizasyon"),
                StepDefinition("eloksal", "Eloksal"),
                StepDefinition("post_eloksal_rinse", "Eloksal Sonrasi Durulama"),
                StepDefinition("nikel", "Nikel"),
                StepDefinition("post_color_rinse", "Renk Sonrasi Durulama"),
                StepDefinition("tespit", "Tespit"),
                StepDefinition("post_tespit_rinse", "Tespit Sonrasi Durulama"),
                StepDefinition("sicaksu", "Sicak Su"),
                StepDefinition("drip", "Suzme"),
                StepDefinition("firin", "Firin"),
                StepDefinition("exit", "Cikis Konveyor"),
            ],
            "pan": [
                StepDefinition("yagalma", "Yag Alma"),
                StepDefinition("durulama_y", "Durulama"),
                StepDefinition("notralizasyon", "Notralizasyon"),
                StepDefinition("eloksal", "Eloksal"),
                StepDefinition("post_eloksal_rinse", "Eloksal Sonrasi Durulama"),
                StepDefinition("nikel", "Nikel"),
                StepDefinition("post_color_rinse", "Renk Sonrasi Durulama"),
                StepDefinition("tespit", "Tespit"),
                StepDefinition("post_tespit_rinse", "Tespit Sonrasi Durulama"),
                StepDefinition("sicaksu", "Sicak Su"),
                StepDefinition("drip", "Suzme"),
                StepDefinition("firin", "Firin"),
                StepDefinition("exit", "Cikis Konveyor"),
            ],
            "natural": [
                StepDefinition("yagalma", "Yag Alma"),
                StepDefinition("durulama_y", "Durulama"),
                StepDefinition("kostik", "Kostik"),
                StepDefinition("durulama_kostik", "Kostik Durulama"),
                StepDefinition("pre_neutral_rinse", "Genel Durulama"),
                StepDefinition("notralizasyon", "Notralizasyon"),
                StepDefinition("eloksal", "Eloksal"),
                StepDefinition("post_eloksal_rinse", "Eloksal Sonrasi Durulama"),
                StepDefinition("di", "DI"),
                StepDefinition("tespit", "Tespit"),
                StepDefinition("post_tespit_rinse", "Tespit Sonrasi Durulama"),
                StepDefinition("sicaksu", "Sicak Su"),
                StepDefinition("drip", "Suzme"),
                StepDefinition("firin", "Firin"),
                StepDefinition("exit", "Cikis Konveyor"),
            ],
            "man": [
                StepDefinition("yagalma", "Yag Alma"),
                StepDefinition("durulama_y", "Durulama"),
                StepDefinition("kostik", "Kostik"),
                StepDefinition("durulama_kostik", "Kostik Durulama"),
                StepDefinition("pre_neutral_rinse", "Genel Durulama"),
                StepDefinition("notralizasyon", "Notralizasyon"),
                StepDefinition("eloksal", "Eloksal"),
                StepDefinition("post_eloksal_rinse", "Eloksal Sonrasi Durulama"),
                StepDefinition("kalay", "Kalay"),
                StepDefinition("post_color_rinse", "Renk Sonrasi Durulama"),
                StepDefinition("tespit", "Tespit"),
                StepDefinition("post_tespit_rinse", "Tespit Sonrasi Durulama"),
                StepDefinition("sicaksu", "Sicak Su"),
                StepDefinition("drip", "Suzme"),
                StepDefinition("firin", "Firin"),
                StepDefinition("exit", "Cikis Konveyor"),
            ],
            "asitmat": [
                StepDefinition("yagalma", "Yag Alma"),
                StepDefinition("durulama_y", "Durulama"),
                StepDefinition("asitmat", "AsitMat"),
                StepDefinition("durulama_a", "AsitMat Sonrasi Durulama"),
                StepDefinition("sokme", "Sokme"),
                StepDefinition("pre_neutral_rinse", "Durulama 1"),
                StepDefinition("pre_neutral_rinse", "Durulama 2"),
                StepDefinition("notralizasyon", "Notralizasyon"),
                StepDefinition("eloksal", "Eloksal"),
                StepDefinition("post_eloksal_rinse", "Eloksal Sonrasi Durulama"),
                StepDefinition("di", "DI"),
                StepDefinition("tespit", "Tespit"),
                StepDefinition("post_tespit_rinse", "Tespit Sonrasi Durulama"),
                StepDefinition("sicaksu", "Sicak Su"),
                StepDefinition("drip", "Suzme"),
                StepDefinition("firin", "Firin"),
                StepDefinition("exit", "Cikis Konveyor"),
            ],
        }
        try:
            return list(routes[route_family])
        except KeyError as exc:
            raise DataContractError(f"Desteklenmeyen rota ailesi: {route_family}") from exc

    def sample_arrival_family_and_micron(self) -> tuple[str, int]:
        probabilities = [item[2] for item in self.data.arrival_mix]
        choice_index = int(self.rng.choice(len(self.data.arrival_mix), p=probabilities))
        route_family, micron, _ = self.data.arrival_mix[choice_index]
        return route_family, micron

    def create_arrival_job(
        self,
        route_family: str | None = None,
        micron: int | None = None,
        replay_step_seconds: dict[tuple[str, int], float] | None = None,
        source_bara_no: str | None = None,
        source_date_text: str | None = None,
        source_start_text: str | None = None,
    ) -> Job:
        self.job_sequence += 1
        if route_family is None or micron is None:
            route_family, micron = self.sample_arrival_family_and_micron()
        job = Job(
            job_id=self.job_sequence,
            route_family=route_family,
            micron=micron,
            route=self.build_route(route_family),
            created_at=self.env.now,
            replay_step_seconds={} if replay_step_seconds is None else dict(replay_step_seconds),
            source_bara_no=source_bara_no,
            source_date_text=source_date_text,
            source_start_text=source_start_text,
        )
        self.initialize_job_report(job)
        self.arrived_jobs.append(job)
        if self.active_jobs == 0:
            self.system_idle_event = self.env.event()
        self.active_jobs += 1
        self.entry_conveyor_busy = True
        self.entry_conveyor_available = self.env.event()
        self.task_sequence += 1
        task = TransportTask(
            task_id=self.task_sequence,
            job=job,
            source_station_id="entry_conveyor",
            source_group="entry_conveyor",
            pickup_crane="vinc1",
            eligible_cranes=("vinc1",),
            ready_time=self.env.now,
            created_sequence=self.task_sequence,
            priority=2,
            timeout_seconds=None,
            timeout_deadline=None,
        )
        self.dispatcher.publish(task)
        self.log_event(
            event_type="arrival",
            message=(
                f"{job.job_id} numarali ({route_family_label(route_family)}, {micron} mikron) bara sisteme geldi."
            ),
            job=job,
            source_station_id="entry_conveyor",
        )
        return job

    def arrival_generator(self, arrival_limit: int | None = None):
        count = 0
        while arrival_limit is None or count < arrival_limit:
            if self.entry_conveyor_busy:
                yield self.entry_conveyor_available
            self.create_arrival_job()
            count += 1
            delay = self.data.interarrival_distribution.sample_seconds(self.rng, "gelislerarasi")
            yield self.env.timeout(delay)
        self.generator_finished = True
        self._check_all_done()

    def replay_arrival_generator(
        self,
        replay_records: list[ReplayArrivalRecord],
        arrival_limit: int | None = None,
        use_fixed_step_seconds: bool = False,
    ):
        selected_records = list(replay_records[:arrival_limit]) if arrival_limit is not None else list(replay_records)
        for index, record in enumerate(selected_records):
            if index > 0:
                delay = self.data.interarrival_distribution.sample_seconds(self.rng, "gelislerarasi")
                yield self.env.timeout(delay)

            if self.entry_conveyor_busy:
                yield self.entry_conveyor_available

            self.create_arrival_job(
                route_family=record.route_family,
                micron=record.micron,
                replay_step_seconds=record.fixed_step_seconds if use_fixed_step_seconds else None,
                source_bara_no=record.source_bara_no,
                source_date_text=record.source_date_text,
                source_start_text=record.source_start_text,
            )

        self.generator_finished = True
        self._check_all_done()

    def complete_job(self, job: Job) -> None:
        job.completed_at = self.env.now
        job.report_row["Bitiş"] = format_seconds_as_hms(job.completed_at)
        job.report_row["Toplam Çevrim (sn)"] = round(job.completed_at - job.created_at, 2)
        self.completed_jobs.append(job)
        self.active_jobs -= 1
        if self.active_jobs == 0 and not self.system_idle_event.triggered:
            self.system_idle_event.succeed(True)
        self.log_event(
            event_type="departure",
            message=(
                f"{job.job_id} numarali ({route_family_label(job.route_family)}, {job.micron} mikron) bara "
                f"sistemden cikti. Toplam cevrim suresi {job.completed_at - job.created_at:.1f} sn."
            ),
            job=job,
            source_station_id=job.current_station_id,
            duration_seconds=job.completed_at - job.created_at,
        )
        self._check_all_done()

    def _check_all_done(self) -> None:
        if self.generator_finished and self.active_jobs == 0 and not self.all_done.triggered:
            self.all_done.succeed(True)

    def summary(self) -> dict[str, Any]:
        completed_cycle_times = [job.completed_at - job.created_at for job in self.completed_jobs if job.completed_at is not None]
        by_family = Counter(job.route_family for job in self.completed_jobs)
        return {
            "sim_time_seconds": float(self.env.now),
            "arrivals": len(self.arrived_jobs),
            "completed": len(self.completed_jobs),
            "wip": self.active_jobs,
            "avg_cycle_seconds": float(np.mean(completed_cycle_times)) if completed_cycle_times else None,
            "completed_by_family": dict(by_family),
            "cranes": {
                crane_id: {
                    "busy_seconds": float(crane.busy_seconds),
                    "completed_moves": crane.completed_moves,
                    "utilization": float(crane.busy_seconds / self.env.now) if self.env.now > 0 else 0.0,
                }
                for crane_id, crane in self.cranes.items()
            },
        }

    def build_wip_trace(self) -> list[tuple[float, int]]:
        changes: list[tuple[float, int]] = []
        for row in self.event_log:
            event_type = row.get("event_type")
            if event_type == "arrival":
                changes.append((float(row["time_seconds"]), 1))
            elif event_type == "departure":
                changes.append((float(row["time_seconds"]), -1))
        changes.sort(key=lambda item: (item[0], 0 if item[1] > 0 else 1))
        trace: list[tuple[float, int]] = [(0.0, 0)]
        current_wip = 0
        for time_seconds, delta in changes:
            current_wip += int(delta)
            trace.append((float(time_seconds), int(current_wip)))
        if not trace or trace[-1][0] < float(self.env.now):
            trace.append((float(self.env.now), int(current_wip)))
        return trace

    def _job_lookup(self) -> dict[int, Job]:
        return {job.job_id: job for job in self.arrived_jobs}

    def _route_text_for_job(self, job: Job) -> str:
        return " -> ".join(step.label for step in job.route)

    def _route_text_for_job_id(self, job_id: int) -> str:
        job = self._job_lookup().get(int(job_id))
        return "" if job is None else self._route_text_for_job(job)

    def _build_havuz_zaman_detay_rows(self) -> list[list[Any]]:
        job_lookup = self._job_lookup()
        rows: list[list[Any]] = []
        for record in self.station_visit_records:
            actual_seconds = self._actual_seconds_for_visit(record)
            waiting_seconds = (
                None if actual_seconds is None else max(0.0, float(actual_seconds) - float(record.nominal_seconds))
            )
            route_text = self._route_text_for_job_id(record.job_id)
            rows.append(
                [
                    record.job_id,
                    route_family_label(record.route_family),
                    record.micron,
                    route_text,
                    record.step_label,
                    record.station_name,
                    record.station_id,
                    format_seconds_as_hms(record.entered_at),
                    format_seconds_as_hms(record.exited_at),
                    round(float(record.nominal_seconds), 2),
                    format_seconds_as_hms(record.nominal_seconds),
                    None if waiting_seconds is None else round(float(waiting_seconds), 2),
                    format_seconds_as_hms(waiting_seconds),
                ]
            )
        return rows

    def _build_crane_transport_rows(self) -> list[list[Any]]:
        rows: list[list[Any]] = []
        open_move_by_crane: dict[str, dict[str, Any]] = {}

        for row in self.event_log:
            crane_id = row.get("crane_id")
            if not crane_id:
                continue
            event_type = str(row.get("event_type") or "")
            if event_type == "pickup":
                open_move_by_crane[crane_id] = {
                    "job_id": row.get("bara_no"),
                    "route_family": row.get("route_family"),
                    "cesit": row.get("cesit"),
                    "mikron": row.get("mikron"),
                    "source_name": row.get("kaynak", ""),
                    "source_id": row.get("source_station_id", ""),
                    "start_time": float(row.get("time_seconds", 0.0)),
                }
                continue

            if event_type == "drip_start":
                open_move = open_move_by_crane.pop(crane_id, None)
                if open_move is not None:
                    elapsed = max(0.0, float(row.get("time_seconds", 0.0)) - float(open_move["start_time"]))
                    rows.append(
                        [
                            crane_label(crane_id),
                            open_move["job_id"],
                            open_move["cesit"],
                            open_move["mikron"],
                            self._route_text_for_job_id(int(open_move["job_id"])),
                            open_move["source_name"],
                            open_move["source_id"],
                            "Süzme",
                            "drip_hold",
                            format_seconds_as_hms(open_move["start_time"]),
                            row.get("time_hms", ""),
                            round(elapsed, 2),
                            format_seconds_as_hms(elapsed),
                        ]
                    )
                open_move_by_crane[crane_id] = {
                    "job_id": row.get("bara_no"),
                    "route_family": row.get("route_family"),
                    "cesit": row.get("cesit"),
                    "mikron": row.get("mikron"),
                    "source_name": "Süzme",
                    "source_id": "drip_hold",
                    "start_time": float(row.get("time_seconds", 0.0)),
                }
                continue

            if event_type not in {"drop", "aux_buffer_drop"}:
                continue

            open_move = open_move_by_crane.pop(crane_id, None)
            if open_move is None:
                continue
            elapsed = max(0.0, float(row.get("time_seconds", 0.0)) - float(open_move["start_time"]))
            rows.append(
                [
                    crane_label(crane_id),
                    open_move["job_id"],
                    open_move["cesit"],
                    open_move["mikron"],
                    self._route_text_for_job_id(int(open_move["job_id"])),
                    open_move["source_name"],
                    open_move["source_id"],
                    row.get("hedef", ""),
                    row.get("destination_station_id", ""),
                    format_seconds_as_hms(open_move["start_time"]),
                    row.get("time_hms", ""),
                    round(elapsed, 2),
                    format_seconds_as_hms(elapsed),
                ]
            )

        return rows

    def _build_crane_wait_rows(self) -> list[list[Any]]:
        rows: list[list[Any]] = []
        for row in self.event_log:
            if row.get("event_type") != "crane_wait_done":
                continue
            duration_seconds = row.get("sure_sn")
            if duration_seconds is None:
                continue
            end_time = float(row.get("time_seconds", 0.0))
            start_time = max(0.0, end_time - float(duration_seconds))
            rows.append(
                [
                    row.get("vinc", ""),
                    row.get("bara_no", ""),
                    row.get("cesit", ""),
                    row.get("mikron", ""),
                    self._route_text_for_job_id(int(row["bara_no"])) if row.get("bara_no") else "",
                    row.get("hedef", ""),
                    format_seconds_as_hms(start_time),
                    row.get("time_hms", ""),
                    round(float(duration_seconds), 2),
                    format_seconds_as_hms(float(duration_seconds)),
                    row.get("blocking_crane", ""),
                    row.get("message", ""),
                ]
            )
        return rows

    def _build_resource_busy_idle_rows(self) -> list[list[Any]]:
        rows: list[list[Any]] = []
        sim_seconds = float(self.env.now)
        visits_by_station: dict[str, list[StationVisitRecord]] = defaultdict(list)
        for record in self.station_visit_records:
            visits_by_station[record.station_id].append(record)

        for station_id, definition in sorted(self.station_definitions.items(), key=lambda item: (item[1].x_m, item[0])):
            if definition.is_virtual or definition.is_sink:
                continue
            visits = visits_by_station.get(station_id, [])
            busy_seconds = sum(self._actual_seconds_for_visit(record) or 0.0 for record in visits)
            idle_seconds = max(0.0, sim_seconds - busy_seconds)
            rows.append(
                [
                    station_id,
                    definition.display_name,
                    definition.process_group or "",
                    len(visits),
                    round(float(busy_seconds), 2),
                    format_seconds_as_hms(busy_seconds),
                    round(float(idle_seconds), 2),
                    format_seconds_as_hms(idle_seconds),
                    round((float(busy_seconds) / sim_seconds * 100.0), 2) if sim_seconds > 0 else 0.0,
                ]
            )

        drip_events = [row for row in self.event_log if row.get("event_type") == "drip_start" and row.get("sure_sn") is not None]
        if drip_events:
            busy_seconds = sum(float(row["sure_sn"]) for row in drip_events)
            idle_seconds = max(0.0, sim_seconds - busy_seconds)
            rows.append(
                [
                    "drip_hold",
                    "Süzme",
                    "suzme",
                    len(drip_events),
                    round(float(busy_seconds), 2),
                    format_seconds_as_hms(busy_seconds),
                    round(float(idle_seconds), 2),
                    format_seconds_as_hms(idle_seconds),
                    round((float(busy_seconds) / sim_seconds * 100.0), 2) if sim_seconds > 0 else 0.0,
                ]
            )

        return rows

    def _build_wip_rows(self) -> list[list[Any]]:
        return [
            [round(float(time_seconds), 2), format_seconds_as_hms(float(time_seconds)), int(wip)]
            for time_seconds, wip in self.build_wip_trace()
        ]

    def _build_cycle_rows(self) -> list[list[Any]]:
        rows: list[list[Any]] = []
        completed_cycle_seconds: list[float] = []
        for job in self.arrived_jobs:
            cycle_seconds = None if job.completed_at is None else max(0.0, float(job.completed_at) - float(job.created_at))
            if cycle_seconds is not None:
                completed_cycle_seconds.append(cycle_seconds)
            rows.append(
                [
                    job.job_id,
                    route_family_label(job.route_family),
                    job.micron,
                    self._route_text_for_job(job),
                    format_seconds_as_hms(job.created_at),
                    format_seconds_as_hms(job.completed_at),
                    None if cycle_seconds is None else round(float(cycle_seconds), 2),
                    format_seconds_as_hms(cycle_seconds),
                ]
            )
        if completed_cycle_seconds:
            rows.append([])
            rows.append(
                [
                    "Ortalama",
                    "",
                    "",
                    "",
                    "",
                    "",
                    round(float(np.mean(completed_cycle_seconds)), 2),
                    format_seconds_as_hms(float(np.mean(completed_cycle_seconds))),
                ]
            )
        return rows

    def export_excel_report(
        self,
        output_path: Path,
        replication_results: list[dict[str, Any]] | None = None,
    ) -> Path:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        event_sheet = workbook.active
        event_sheet.title = "Olay Logu"
        self._write_sheet(
            event_sheet,
            [
                "Sim Zamanı (sn)",
                "Zaman",
                "Olay Tipi",
                "Mesaj",
                "Bara No",
                "Çeşit",
                "Mikron",
                "Vinç",
                "Kaynak",
                "Hedef",
                "Süre (sn)",
                "Süre (HH:MM:SS)",
            ],
            [
                [
                    row["time_seconds"],
                    row["time_hms"],
                    row["event_type"],
                    row["message"],
                    row["bara_no"],
                    row["cesit"],
                    row["mikron"],
                    row["vinc"],
                    row["kaynak"],
                    row["hedef"],
                    row["sure_sn"],
                    row["sure_hms"],
                ]
                for row in self.event_log
            ],
        )

        havuz_sheet = workbook.create_sheet("Havuz Süreler Sim")
        self._write_sheet(
            havuz_sheet,
            HAVUZ_SURELER_COLUMNS,
            [[job.report_row.get(column, "") for column in HAVUZ_SURELER_COLUMNS] for job in self.arrived_jobs],
        )

        actual_summary_sheet = workbook.create_sheet("Fiili Havuz Süreler Sim")
        self._write_sheet(
            actual_summary_sheet,
            HAVUZ_SURELER_COLUMNS,
            self._build_actual_havuz_report_rows(),
        )

        actual_sheet = workbook.create_sheet("Fiili Havuz Kalış Detay")
        self._write_sheet(
            actual_sheet,
            [
                "Bara No",
                "Çeşit",
                "Mikron",
                "Adım",
                "Havuz",
                "Havuz ID",
                "Giriş",
                "Çıkış",
                "Nominal Süre (sn)",
                "Nominal Süre (HH:MM:SS)",
                "Fiili Kalış (sn)",
                "Fiili Kalış (HH:MM:SS)",
                "Proses Sonrası Bekleme (sn)",
                "Proses Sonrası Bekleme (HH:MM:SS)",
                "Bekleme Nedeni",
                "Gideceği Adım",
                "Engelleyen Havuz(lar)",
                "Engelleyen Bara No(lar)",
                "İlk Blokaj",
                "Son Blokaj",
                "Açık Kayıt mı",
            ],
            [self._station_visit_row(record) for record in self.station_visit_records],
        )

        havuz_timeline_sheet = workbook.create_sheet("Havuz Zaman Detay")
        self._write_sheet(
            havuz_timeline_sheet,
            [
                "Askı/Bara ID",
                "Çeşit",
                "Mikron",
                "Rota Bilgisi",
                "Adım",
                "Havuz Adı",
                "Havuz ID",
                "Giriş Zamanı",
                "Çıkış Zamanı",
                "İşlem Süresi (sn)",
                "İşlem Süresi (HH:MM:SS)",
                "Bekleme Süresi (sn)",
                "Bekleme Süresi (HH:MM:SS)",
            ],
            self._build_havuz_zaman_detay_rows(),
        )

        crane_transport_sheet = workbook.create_sheet("Vinc Tasima Sureleri")
        self._write_sheet(
            crane_transport_sheet,
            [
                "Vinç",
                "Askı/Bara ID",
                "Çeşit",
                "Mikron",
                "Rota Bilgisi",
                "Kaynak",
                "Kaynak ID",
                "Hedef",
                "Hedef ID",
                "Başlangıç Zamanı",
                "Bitiş Zamanı",
                "Taşıma Süresi (sn)",
                "Taşıma Süresi (HH:MM:SS)",
            ],
            self._build_crane_transport_rows(),
        )

        crane_wait_sheet = workbook.create_sheet("Vinc Bekleme Sureleri")
        self._write_sheet(
            crane_wait_sheet,
            [
                "Vinç",
                "Askı/Bara ID",
                "Çeşit",
                "Mikron",
                "Rota Bilgisi",
                "Beklenen Hedef",
                "Bekleme Başlangıcı",
                "Bekleme Bitişi",
                "Bekleme Süresi (sn)",
                "Bekleme Süresi (HH:MM:SS)",
                "Engelleyen Vinç(ler)",
                "Mesaj",
            ],
            self._build_crane_wait_rows(),
        )

        resource_sheet = workbook.create_sheet("Kaynak Meshgul Bos")
        self._write_sheet(
            resource_sheet,
            [
                "Kaynak ID",
                "Kaynak Adı",
                "İşlem Grubu",
                "Ziyaret Sayısı",
                "Meşgul Süre (sn)",
                "Meşgul Süre (HH:MM:SS)",
                "Boş Süre (sn)",
                "Boş Süre (HH:MM:SS)",
                "Kullanım (%)",
            ],
            self._build_resource_busy_idle_rows(),
        )

        wip_sheet = workbook.create_sheet("WIP Iz")
        self._write_sheet(
            wip_sheet,
            ["Sim Zamanı (sn)", "Zaman", "WIP Miktarı"],
            self._build_wip_rows(),
        )

        cycle_sheet = workbook.create_sheet("Cevrim Sureleri")
        self._write_sheet(
            cycle_sheet,
            [
                "Askı/Bara ID",
                "Çeşit",
                "Mikron",
                "Rota Bilgisi",
                "Giriş Zamanı",
                "Çıkış Zamanı",
                "Çevrim Süresi (sn)",
                "Çevrim Süresi (HH:MM:SS)",
            ],
            self._build_cycle_rows(),
        )

        probability_sheet = workbook.create_sheet("Geliş Olasılıkları")
        route_probs = defaultdict(float)
        micron_probs = defaultdict(float)
        joint_rows = []
        for route_family, micron, probability in self.data.arrival_mix:
            route_probs[route_family_label(route_family)] += probability
            micron_probs[micron] += probability
            joint_rows.append(
                [route_family_label(route_family), micron, round(probability, 6), round(probability * 100.0, 3)]
            )
        probability_sheet.append(["Rota Ailesi", "Mikron", "Birleşik Olasılık", "Birleşik Olasılık (%)"])
        for row in joint_rows:
            probability_sheet.append(row)
        probability_sheet.append([])
        probability_sheet.append(["Rota Ailesi", "Marjinal Olasılık", "Marjinal Olasılık (%)"])
        for key, value in sorted(route_probs.items(), key=lambda item: item[1], reverse=True):
            probability_sheet.append([key, round(value, 6), round(value * 100.0, 3)])
        probability_sheet.append([])
        probability_sheet.append(["Mikron", "Marjinal Olasılık", "Marjinal Olasılık (%)"])
        for key, value in sorted(micron_probs.items(), key=lambda item: int(item[0])):
            probability_sheet.append([key, round(value, 6), round(value * 100.0, 3)])
        self._style_sheet(probability_sheet)

        distribution_sheet = workbook.create_sheet("Dağılım Parametreleri")
        distribution_rows = []
        for process_group, spec in self.data.pool_distributions.items():
            distribution_rows.append([process_group, spec.name, spec.params_text, "havuz"])
        for micron, spec in sorted(self.data.eloxal_distributions_by_micron.items()):
            distribution_rows.append([f"eloksal_{micron}um", spec.name, spec.params_text, "mikron"])
        distribution_rows.append(
            ["interarrival", self.data.interarrival_distribution.name, self.data.interarrival_distribution.params_text, "gelis"]
        )
        self._write_sheet(
            distribution_sheet,
            ["İşlem Grubu", "Dağılım", "Parametreler", "Kategori"],
            distribution_rows,
        )

        capacity_sheet = workbook.create_sheet("Istasyon Efektif Kapasite")
        self._write_sheet(
            capacity_sheet,
            [
                "İşlem Grubu",
                "İstasyon/Adım",
                "Paralel Birim",
                "Ortalama Proses (sn)",
                "Ortalama Proses (HH:MM:SS)",
                "Teorik Kapasite (iş/saat)",
                "Gerçekleşen Akış (iş/saat)",
                "Akış / Kapasite (%)",
                "Ortalama Fiili Kalış (sn)",
                "Ortalama Fiili Kalış (HH:MM:SS)",
                "Not",
            ],
            self._build_effective_capacity_rows(),
        )

        crane_sheet = workbook.create_sheet("Vinc Kullanim Orani")
        self._write_crane_utilization_sheet(
            crane_sheet,
            replication_results if replication_results else [{"replication": 1, "seed": None, "summary": self.summary()}],
        )

        if replication_results:
            replication_sheet = workbook.create_sheet("Replikasyon Ozeti")
            self._write_replication_summary_sheet(replication_sheet, replication_results)

        welch_sheet = workbook.create_sheet("Welch WIP")
        self._write_welch_sheet(
            welch_sheet,
            replication_results if replication_results else [{"replication": 1, "seed": None, "summary": self.summary(), "wip_trace": self.build_wip_trace()}],
        )

        bottleneck_sheet = workbook.create_sheet("Darbogaz Ozeti")
        self._write_bottleneck_sheet(bottleneck_sheet)

        try:
            workbook.save(output_path)
            return output_path
        except PermissionError:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            fallback = output_path.with_name(f"{output_path.stem}_{timestamp}{output_path.suffix}")
            workbook.save(fallback)
            return fallback

    def _write_sheet(self, sheet, headers: list[str], rows: list[list[Any]]) -> None:
        sheet.append(headers)
        for row in rows:
            sheet.append(row)
        self._style_sheet(sheet)

    def _actual_seconds_for_visit(self, record: StationVisitRecord) -> float | None:
        if record.exited_at is None:
            return None
        return max(0.0, float(record.exited_at) - float(record.entered_at))

    def _estimate_process_group_mean_seconds(self, process_group: str, sample_count: int = 4000) -> float:
        rng = np.random.default_rng(20260424)
        if process_group == "eloksal":
            micron_probs: dict[int, float] = defaultdict(float)
            for _, micron, probability in self.data.arrival_mix:
                micron_probs[int(micron)] += float(probability)
            weighted = 0.0
            for micron, probability in micron_probs.items():
                spec = self.data.eloxal_distributions_by_micron[micron]
                samples = [spec.sample_seconds(rng, f"eloksal-{micron}") for _ in range(sample_count)]
                weighted += float(np.mean(samples)) * probability
            return weighted
        if process_group == "tespit":
            return sum(float(probability) * float(micron * 60) for _, micron, probability in self.data.arrival_mix)
        if process_group == "suzme":
            spec = self.data.pool_distributions["suzme"]
            samples = [spec.sample_seconds(rng, "suzme") for _ in range(sample_count)]
            return float(np.mean(samples))
        spec = self.data.pool_distributions[process_group]
        samples = [spec.sample_seconds(rng, process_group) for _ in range(sample_count)]
        return float(np.mean(samples))

    def _build_effective_capacity_rows(self) -> list[list[Any]]:
        label_map = {
            "yagalma": "Yağ Alma",
            "durulama_y": "Durulama Y",
            "asitmat": "AsitMat",
            "durulama_a": "Durulama A",
            "sokme": "Sökme",
            "kostik": "Kostik",
            "durulama_kostik": "Kostik Durulama",
            "durulama_1": "Durulama 1",
            "durulama_2": "Durulama 2",
            "notralizasyon": "Nötralizasyon",
            "eloksal": "Eloksal",
            "durulama_post_eloksal": "Eloksal Sonrası Durulama",
            "renk": "Kalay / Nikel (Renk)",
            "durulama_post_renk": "Renk Sonrası Durulama",
            "di": "DI",
            "tespit": "Tespit",
            "durulama_post_tespit": "Tespit Sonrası Durulama",
            "sicaksu": "Sıcak Su",
            "suzme": "Süzme",
            "firin": "Fırın",
        }
        ordered_groups = [
            "yagalma",
            "durulama_y",
            "asitmat",
            "durulama_a",
            "sokme",
            "kostik",
            "durulama_kostik",
            "durulama_1",
            "durulama_2",
            "notralizasyon",
            "eloksal",
            "durulama_post_eloksal",
            "renk",
            "durulama_post_renk",
            "di",
            "tespit",
            "durulama_post_tespit",
            "sicaksu",
            "suzme",
            "firin",
        ]
        station_counts: dict[str, int] = defaultdict(int)
        for definition in self.station_definitions.values():
            if definition.process_group:
                station_counts[definition.process_group] += 1
        station_counts["suzme"] = 1

        station_actuals: dict[str, list[float]] = defaultdict(list)
        station_visits: dict[str, int] = defaultdict(int)
        for record in self.station_visit_records:
            actual_seconds = self._actual_seconds_for_visit(record)
            if actual_seconds is None:
                continue
            group = self.station_definitions[record.station_id].process_group
            if not group:
                continue
            station_actuals[group].append(float(actual_seconds))
            station_visits[group] += 1

        drip_events = [row for row in self.event_log if row.get("event_type") == "drip_start" and row.get("sure_sn")]
        if drip_events:
            station_actuals["suzme"] = [float(row["sure_sn"]) for row in drip_events]
            station_visits["suzme"] = len(drip_events)

        rows: list[list[Any]] = []
        sim_hours = float(self.env.now) / 3600.0 if self.env.now > 0 else 0.0
        for process_group in ordered_groups:
            if process_group != "suzme" and process_group not in self.data.pool_distributions and process_group != "tespit":
                continue
            mean_seconds = self._estimate_process_group_mean_seconds(process_group)
            parallel_units = max(1, int(station_counts.get(process_group, 1)))
            theoretical_capacity = parallel_units * 3600.0 / max(mean_seconds, 1e-9)
            actual_rate = (
                float(station_visits.get(process_group, 0)) / sim_hours if sim_hours > 0 else 0.0
            )
            load_ratio = (actual_rate / theoretical_capacity * 100.0) if theoretical_capacity > 0 else 0.0
            actual_mean = (
                float(np.mean(station_actuals[process_group])) if station_actuals.get(process_group) else None
            )
            note = ""
            if process_group == "suzme":
                note = "Ayrı havuz değil; Vinç 4 süzme boyunca kilitlenir."
            elif process_group == "firin":
                note = "Son hat kapasitesi süzme ve Vinç 4 ile birlikte değerlendirilmelidir."
            elif process_group == "eloksal":
                note = "Mikron karışımına göre ağırlıklı ortalama süre kullanıldı."
            elif process_group == "tespit":
                note = "Mikron * 60 sn kuralına göre ağırlıklı ortalama kullanıldı."
            rows.append(
                [
                    process_group,
                    label_map.get(process_group, process_group),
                    parallel_units,
                    round(mean_seconds, 2),
                    format_seconds_as_hms(mean_seconds),
                    round(theoretical_capacity, 2),
                    round(actual_rate, 2),
                    round(load_ratio, 2),
                    None if actual_mean is None else round(actual_mean, 2),
                    format_seconds_as_hms(actual_mean),
                    note,
                ]
            )
        return rows

    def _build_actual_havuz_report_rows(self) -> list[list[Any]]:
        visits_by_job: dict[int, list[StationVisitRecord]] = defaultdict(list)
        for record in self.station_visit_records:
            visits_by_job[record.job_id].append(record)
        drip_events_by_job: dict[int, list[dict[str, Any]]] = defaultdict(list)
        for row in self.event_log:
            if row.get("event_type") != "drip_start":
                continue
            job_id = row.get("bara_no")
            drip_seconds = row.get("sure_sn")
            if job_id is None or drip_seconds is None:
                continue
            drip_events_by_job[int(job_id)].append(row)

        rows: list[list[Any]] = []
        for job in self.arrived_jobs:
            row = {column: "" for column in HAVUZ_SURELER_COLUMNS}
            row["Bara No."] = job.job_id
            row["Çeşit"] = route_family_label(job.route_family)
            row["Mikron"] = job.micron
            row["Başlangıç"] = format_seconds_as_hms(job.created_at)
            row["Bitiş"] = format_seconds_as_hms(job.completed_at)
            row["Toplam Çevrim (sn)"] = (
                "" if job.completed_at is None else round(float(job.completed_at) - float(job.created_at), 2)
            )
            row["Açıklama"] = "Fiili havuz kalış süresi"

            occurrences: dict[str, int] = defaultdict(int)
            for record in sorted(visits_by_job.get(job.job_id, []), key=lambda item: item.visit_id):
                column = self._column_for_step_occurrence(record.step_key, occurrences[record.step_key])
                occurrences[record.step_key] += 1
                if not column:
                    continue
                actual_seconds = self._actual_seconds_for_visit(record)
                row[column] = format_seconds_as_hms(actual_seconds)

            for drip_row in drip_events_by_job.get(job.job_id, []):
                column = self._column_for_step_occurrence("drip", occurrences["drip"])
                occurrences["drip"] += 1
                if not column:
                    continue
                row[column] = format_seconds_as_hms(float(drip_row["sure_sn"]))

            rows.append([row.get(column, "") for column in HAVUZ_SURELER_COLUMNS])
        return rows

    def _build_pool_bottleneck_rows(self) -> list[list[Any]]:
        event_df = pd.DataFrame(self.event_log)
        if event_df.empty:
            return []
        task_blocked_df = event_df[event_df["event_type"] == "task_blocked"].copy()
        if task_blocked_df.empty:
            return []

        rows: list[list[Any]] = []
        for record in self.station_visit_records:
            actual_seconds = self._actual_seconds_for_visit(record)
            if actual_seconds is None:
                continue
            extra_seconds = max(0.0, float(actual_seconds) - float(record.nominal_seconds))
            if extra_seconds <= 1e-9:
                continue

            blocked_events = task_blocked_df[
                (task_blocked_df["bara_no"] == record.job_id)
                & (task_blocked_df["kaynak"] == record.station_name)
                & (task_blocked_df["time_seconds"] >= float(record.entered_at) + float(record.nominal_seconds) - 1e-9)
                & (task_blocked_df["time_seconds"] <= float(record.exited_at) + 1e-9)
            ].sort_values("time_seconds")
            if blocked_events.empty:
                continue

            first_block = blocked_events.iloc[0]
            rows.append(
                [
                    record.job_id,
                    route_family_label(record.route_family),
                    record.micron,
                    record.station_name,
                    first_block.get("next_step_label", ""),
                    first_block.get("blocking_station_names", ""),
                    first_block.get("blocking_job_ids", ""),
                    round(float(record.nominal_seconds), 2),
                    format_seconds_as_hms(record.nominal_seconds),
                    round(float(actual_seconds), 2),
                    format_seconds_as_hms(actual_seconds),
                    round(float(extra_seconds), 2),
                    format_seconds_as_hms(extra_seconds),
                    format_seconds_as_hms(float(record.entered_at) + float(record.nominal_seconds)),
                    format_seconds_as_hms(record.exited_at),
                ]
            )

        rows.sort(key=lambda item: (-float(item[11]), item[0], item[3], item[4]))
        return rows

    def _blockage_summary_for_visit(self, record: StationVisitRecord) -> dict[str, str]:
        actual_seconds = self._actual_seconds_for_visit(record)
        if actual_seconds is None:
            return {
                "wait_reason": "",
                "next_step_label": "",
                "blocking_station_names": "",
                "blocking_job_ids": "",
                "first_block_hms": "",
                "last_block_hms": "",
            }

        wait_start = float(record.entered_at) + float(record.nominal_seconds)
        if float(record.exited_at) <= wait_start + 1e-9:
            return {
                "wait_reason": "",
                "next_step_label": "",
                "blocking_station_names": "",
                "blocking_job_ids": "",
                "first_block_hms": "",
                "last_block_hms": "",
            }

        event_df = pd.DataFrame(self.event_log)
        if event_df.empty:
            return {
                "wait_reason": "Vinç/pickup beklemesi",
                "next_step_label": "",
                "blocking_station_names": "",
                "blocking_job_ids": "",
                "first_block_hms": "",
                "last_block_hms": "",
            }

        blocked_events = event_df[
            (event_df["event_type"] == "task_blocked")
            & (event_df["bara_no"] == record.job_id)
            & (event_df["kaynak"] == record.station_name)
            & (event_df["time_seconds"] >= wait_start - 1e-9)
            & (event_df["time_seconds"] <= float(record.exited_at) + 1e-9)
        ].sort_values("time_seconds")
        if blocked_events.empty:
            return {
                "wait_reason": "Vinç/pickup beklemesi",
                "next_step_label": "",
                "blocking_station_names": "",
                "blocking_job_ids": "",
                "first_block_hms": "",
                "last_block_hms": "",
            }

        next_steps: list[str] = []
        blocking_stations: list[str] = []
        blocking_jobs: list[str] = []
        for _, row in blocked_events.iterrows():
            next_step = str(row.get("next_step_label", "") or "").strip()
            if next_step and next_step not in next_steps:
                next_steps.append(next_step)
            for station_name in [item for item in str(row.get("blocking_station_names", "")).split("; ") if item]:
                if station_name not in blocking_stations:
                    blocking_stations.append(station_name)
            for job_id in [item for item in str(row.get("blocking_job_ids", "")).split("|") if item]:
                if job_id not in blocking_jobs:
                    blocking_jobs.append(job_id)

        return {
            "wait_reason": "Havuz blokaji",
            "next_step_label": "; ".join(next_steps),
            "blocking_station_names": "; ".join(blocking_stations),
            "blocking_job_ids": "|".join(blocking_jobs),
            "first_block_hms": format_seconds_as_hms(float(blocked_events.iloc[0]["time_seconds"])),
            "last_block_hms": format_seconds_as_hms(float(blocked_events.iloc[-1]["time_seconds"])),
        }

    def _station_visit_row(self, record: StationVisitRecord) -> list[Any]:
        actual_seconds = self._actual_seconds_for_visit(record)
        post_process_wait = (
            None
            if actual_seconds is None
            else max(0.0, float(actual_seconds) - float(record.nominal_seconds))
        )
        blockage = self._blockage_summary_for_visit(record)
        return [
            record.job_id,
            route_family_label(record.route_family),
            record.micron,
            record.step_label,
            record.station_name,
            record.station_id,
            format_seconds_as_hms(record.entered_at),
            format_seconds_as_hms(record.exited_at),
            round(record.nominal_seconds, 2),
            format_seconds_as_hms(record.nominal_seconds),
            None if actual_seconds is None else round(actual_seconds, 2),
            format_seconds_as_hms(actual_seconds),
            None if post_process_wait is None else round(post_process_wait, 2),
            format_seconds_as_hms(post_process_wait),
            blockage["wait_reason"],
            blockage["next_step_label"],
            blockage["blocking_station_names"],
            blockage["blocking_job_ids"],
            blockage["first_block_hms"],
            blockage["last_block_hms"],
            "Evet" if record.exited_at is None else "Hayır",
        ]

    def _write_bottleneck_sheet(self, sheet) -> None:
        df = pd.DataFrame(self.event_log)

        def append_section(title: str, headers: list[str], rows: list[list[Any]]) -> None:
            if sheet.max_row > 1 or sheet["A1"].value is not None:
                sheet.append([])
            title_row = sheet.max_row + 1
            sheet.append([title])
            sheet.cell(title_row, 1).font = Font(bold=True)
            header_row = sheet.max_row + 1
            sheet.append(headers)
            for cell in sheet[header_row]:
                cell.font = Font(bold=True)
            for row in rows:
                sheet.append(row)

        if df.empty:
            append_section("Genel", ["Metrik", "Deger"], [["Darbogaz kaydi yok", 0]])
            self._autosize_sheet(sheet)
            return

        crane_wait_df = df[df["event_type"] == "crane_wait"].copy()
        task_blocked_df = df[df["event_type"] == "task_blocked"].copy()

        append_section(
            "Genel",
            ["Metrik", "Deger"],
            [
                ["Toplam vinç bekleme olayi", int(len(crane_wait_df))],
                ["Toplam havuz blokaj olayi", int(len(task_blocked_df))],
                [
                    "Havuz blokajindan etkilenen farkli bara sayisi",
                    int(task_blocked_df["bara_no"].dropna().nunique()) if not task_blocked_df.empty else 0,
                ],
                [
                    "Vinç beklemesinde gorulen farkli vinç cifti",
                    int(crane_wait_df[["vinc", "blocking_crane"]].drop_duplicates().shape[0])
                    if not crane_wait_df.empty
                    else 0,
                ],
            ],
        )

        if not crane_wait_df.empty:
            crane_wait_summary = (
                crane_wait_df.groupby(["vinc", "blocking_crane", "hedef"], dropna=False)
                .agg(
                    olay_sayisi=("event_type", "size"),
                    ilk_zaman_sn=("time_seconds", "min"),
                    son_zaman_sn=("time_seconds", "max"),
                )
                .reset_index()
                .sort_values(["olay_sayisi", "ilk_zaman_sn"], ascending=[False, True])
            )
            crane_wait_rows = [
                [
                    row["vinc"],
                    row["blocking_crane"],
                    row["hedef"],
                    int(row["olay_sayisi"]),
                    format_seconds_as_hms(row["ilk_zaman_sn"]),
                    format_seconds_as_hms(row["son_zaman_sn"]),
                ]
                for _, row in crane_wait_summary.iterrows()
            ]
        else:
            crane_wait_rows = [["", "", "", 0, "", ""]]

        append_section(
            "Vinç Bekleme Ozeti",
            ["Bekleyen Vinç", "Engelleyen Vinç", "Hedef", "Olay Sayisi", "Ilk Zaman", "Son Zaman"],
            crane_wait_rows,
        )

        if not task_blocked_df.empty:
            expanded_rows: list[dict[str, Any]] = []
            for _, row in task_blocked_df.iterrows():
                station_names = [item for item in str(row.get("blocking_station_names", "")).split("; ") if item]
                station_ids = [item for item in str(row.get("blocking_station_ids", "")).split("|") if item]
                job_ids = [item for item in str(row.get("blocking_job_ids", "")).split("|") if item]
                count = max(len(station_names), len(station_ids), len(job_ids))
                for idx in range(count):
                    expanded_rows.append(
                        {
                            "kaynak": row["kaynak"],
                            "sonraki_adim": row.get("next_step_label", ""),
                            "engelleyen_havuz": station_names[idx] if idx < len(station_names) else "",
                            "engelleyen_bara_no": job_ids[idx] if idx < len(job_ids) else "",
                            "blocked_bara_no": row["bara_no"],
                            "time_seconds": row["time_seconds"],
                        }
                    )
            expanded_df = pd.DataFrame(expanded_rows)
            pool_summary = (
                expanded_df.groupby(["kaynak", "sonraki_adim", "engelleyen_havuz"], dropna=False)
                .agg(
                    blokaj_sayisi=("engelleyen_havuz", "size"),
                    etkilenen_farkli_bara=("blocked_bara_no", pd.Series.nunique),
                    engelleyen_farkli_bara=("engelleyen_bara_no", pd.Series.nunique),
                    ilk_zaman_sn=("time_seconds", "min"),
                    son_zaman_sn=("time_seconds", "max"),
                )
                .reset_index()
                .sort_values(["blokaj_sayisi", "ilk_zaman_sn"], ascending=[False, True])
            )
            pool_rows = [
                [
                    row["kaynak"],
                    row["sonraki_adim"],
                    row["engelleyen_havuz"],
                    int(row["blokaj_sayisi"]),
                    int(row["etkilenen_farkli_bara"]),
                    int(row["engelleyen_farkli_bara"]),
                    format_seconds_as_hms(row["ilk_zaman_sn"]),
                    format_seconds_as_hms(row["son_zaman_sn"]),
                ]
                for _, row in pool_summary.iterrows()
            ]

            combo_summary = (
                task_blocked_df.groupby(["kaynak", "next_step_label", "blocking_station_names"], dropna=False)
                .agg(
                    olay_sayisi=("event_type", "size"),
                    farkli_bara=("bara_no", pd.Series.nunique),
                    ilk_zaman_sn=("time_seconds", "min"),
                    son_zaman_sn=("time_seconds", "max"),
                )
                .reset_index()
                .sort_values(["olay_sayisi", "ilk_zaman_sn"], ascending=[False, True])
            )
            combo_rows = [
                [
                    row["kaynak"],
                    row["next_step_label"],
                    row["blocking_station_names"],
                    int(row["olay_sayisi"]),
                    int(row["farkli_bara"]),
                    format_seconds_as_hms(row["ilk_zaman_sn"]),
                    format_seconds_as_hms(row["son_zaman_sn"]),
                ]
                for _, row in combo_summary.iterrows()
            ]
        else:
            pool_rows = [["", "", "", 0, 0, 0, "", ""]]
            combo_rows = [["", "", "", 0, 0, "", ""]]
        pool_bottleneck_rows = self._build_pool_bottleneck_rows()
        if not pool_bottleneck_rows:
            pool_bottleneck_rows = [["", "", "", "", "", "", "", 0, "", 0, "", 0, "", "", ""]]

        append_section(
            "Havuz Blokaj Ozeti",
            [
                "Kaynak Havuz",
                "Sonraki Adim",
                "Engelleyen Havuz",
                "Blokaj Sayisi",
                "Etkilenen Farkli Bara",
                "Engelleyen Farkli Bara",
                "Ilk Zaman",
                "Son Zaman",
            ],
            pool_rows,
        )
        append_section(
            "En Sik Darbogaz Kombinasyonlari",
            [
                "Kaynak Havuz",
                "Sonraki Adim",
                "Engelleyen Havuz(lar)",
                "Olay Sayisi",
                "Etkilenen Farkli Bara",
                "Ilk Zaman",
                "Son Zaman",
            ],
            combo_rows,
        )
        append_section(
            "Bara Bazli Havuz Darbogazi",
            [
                "Bara No",
                "Çeşit",
                "Mikron",
                "Bulundugu Havuz",
                "Gidecegi Havuz/Adim",
                "Engelleyen Havuz(lar)",
                "Engelleyen Bara No(lar)",
                "Nominal Süre (sn)",
                "Nominal Süre (HH:MM:SS)",
                "Fiili Kalış (sn)",
                "Fiili Kalış (HH:MM:SS)",
                "Fazladan Durus (sn)",
                "Fazladan Durus (HH:MM:SS)",
                "Bekleme Baslangici",
                "Havuzdan Cikis",
            ],
            pool_bottleneck_rows,
        )
        self._autosize_sheet(sheet)

    def _write_crane_utilization_sheet(self, sheet, replication_results: list[dict[str, Any]]) -> None:
        headers = [
            "Replikasyon",
            "Seed",
            "Vinç",
            "Sim Süresi (sn)",
            "Yoğun Süre (sn)",
            "Kullanım Oranı (%)",
            "Tamamlanan Hareket",
        ]
        rows: list[list[Any]] = []
        for item in replication_results:
            summary = item["summary"]
            cranes = summary.get("cranes", {})
            for crane_id, crane_data in cranes.items():
                rows.append(
                    [
                        item.get("replication"),
                        item.get("seed"),
                        crane_label(crane_id),
                        round(float(summary.get("sim_time_seconds", 0.0)), 2),
                        round(float(crane_data.get("busy_seconds", 0.0)), 2),
                        round(float(crane_data.get("utilization", 0.0)) * 100.0, 2),
                        int(crane_data.get("completed_moves", 0)),
                    ]
                )
        self._write_sheet(sheet, headers, rows)

        if len(replication_results) > 1:
            sheet.append([])
            title_row = sheet.max_row + 1
            sheet.append(["Ortalama Kullanım"])
            sheet.cell(title_row, 1).font = Font(bold=True)
            header_row = sheet.max_row + 1
            avg_headers = ["Vinç", "Ortalama Kullanım (%)", "Min Kullanım (%)", "Max Kullanım (%)"]
            sheet.append(avg_headers)
            for cell in sheet[header_row]:
                cell.font = Font(bold=True)
            summary_df = pd.DataFrame(rows, columns=headers)
            avg_df = (
                summary_df.groupby("Vinç", dropna=False)["Kullanım Oranı (%)"]
                .agg(["mean", "min", "max"])
                .reset_index()
                .sort_values("Vinç")
            )
            for _, row in avg_df.iterrows():
                sheet.append(
                    [
                        row["Vinç"],
                        round(float(row["mean"]), 2),
                        round(float(row["min"]), 2),
                        round(float(row["max"]), 2),
                    ]
                )
            self._autosize_sheet(sheet)

    def _write_replication_summary_sheet(self, sheet, replication_results: list[dict[str, Any]]) -> None:
        headers = [
            "Replikasyon",
            "Seed",
            "Sim Süresi (sn)",
            "Giren Bara",
            "Çıkan Bara",
            "WIP",
            "Ortalama Çevrim (sn)",
        ]
        rows = [
            [
                item.get("replication"),
                item.get("seed"),
                round(float(item["summary"].get("sim_time_seconds", 0.0)), 2),
                int(item["summary"].get("arrivals", 0)),
                int(item["summary"].get("completed", 0)),
                int(item["summary"].get("wip", 0)),
                None
                if item["summary"].get("avg_cycle_seconds") is None
                else round(float(item["summary"]["avg_cycle_seconds"]), 2),
            ]
            for item in replication_results
        ]
        self._write_sheet(sheet, headers, rows)
        if len(replication_results) > 1:
            sheet.append([])
            title_row = sheet.max_row + 1
            sheet.append(["Toplu Özet"])
            sheet.cell(title_row, 1).font = Font(bold=True)
            header_row = sheet.max_row + 1
            sheet.append(["Metrik", "Ortalama", "Min", "Max"])
            for cell in sheet[header_row]:
                cell.font = Font(bold=True)
            df = pd.DataFrame(rows, columns=headers)
            for column in ["Sim Süresi (sn)", "Giren Bara", "Çıkan Bara", "WIP", "Ortalama Çevrim (sn)"]:
                values = pd.to_numeric(df[column], errors="coerce").dropna()
                if values.empty:
                    continue
                sheet.append(
                    [
                        column,
                        round(float(values.mean()), 2),
                        round(float(values.min()), 2),
                        round(float(values.max()), 2),
                    ]
                )
            self._autosize_sheet(sheet)

    def _write_welch_sheet(self, sheet, replication_results: list[dict[str, Any]]) -> None:
        valid_results = [item for item in replication_results if item.get("wip_trace")]
        if not valid_results:
            self._write_sheet(sheet, ["Not"], [["WIP izi bulunamadi."]])
            return

        max_time = max(float(item["summary"].get("sim_time_seconds", 0.0)) for item in valid_results)
        if max_time <= 0:
            self._write_sheet(sheet, ["Not"], [["Simulasyon suresi sifir gorunuyor."]])
            return

        target_points = 240.0
        step_seconds = max(30.0, math.ceil((max_time / target_points) / 30.0) * 30.0)
        sample_times: list[float] = []
        current = 0.0
        while current <= max_time + 1e-9:
            sample_times.append(round(current, 6))
            current += step_seconds
        if sample_times[-1] < max_time:
            sample_times.append(float(max_time))

        headers = ["Zaman (sn)", "Zaman (HH:MM:SS)"]
        replication_columns: list[list[float]] = []

        def sample_trace(trace: list[tuple[float, int]], target_time: float) -> int:
            current_wip = 0
            trace_idx = 0
            while trace_idx < len(trace) and float(trace[trace_idx][0]) <= target_time + 1e-9:
                current_wip = int(trace[trace_idx][1])
                trace_idx += 1
            return current_wip

        for item in valid_results:
            headers.append(f"Rep {item.get('replication')} WIP")
            trace = list(item["wip_trace"])
            column_values = [sample_trace(trace, t) for t in sample_times]
            replication_columns.append(column_values)

        headers.extend(["Ortalama WIP", "Welch Ortalama (MA-5)"])
        rows: list[list[Any]] = []
        average_series: list[float] = []
        moving_average_series: list[float] = []
        for idx, sample_time in enumerate(sample_times):
            rep_values = [column[idx] for column in replication_columns]
            average_wip = float(np.mean(rep_values)) if rep_values else 0.0
            average_series.append(average_wip)
            left = max(0, idx - 4)
            moving_average = float(np.mean(average_series[left : idx + 1]))
            moving_average_series.append(moving_average)
            rows.append(
                [
                    round(sample_time, 2),
                    format_seconds_as_hms(sample_time),
                    *rep_values,
                    round(average_wip, 4),
                    round(moving_average, 4),
                ]
            )

        self._write_sheet(sheet, headers, rows)

        chart = LineChart()
        chart.title = "Welch Grafigi - WIP"
        chart.y_axis.title = "WIP"
        chart.x_axis.title = "Simulasyon Zamani (sn)"
        chart.height = 10
        chart.width = 22
        avg_col = len(headers) - 1
        ma_col = len(headers)
        data = Reference(sheet, min_col=avg_col, max_col=ma_col, min_row=1, max_row=len(rows) + 1)
        cats = Reference(sheet, min_col=1, min_row=2, max_row=len(rows) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.legend.position = "r"
        anchor_col = get_column_letter(min(len(headers) + 2, 30))
        sheet.add_chart(chart, f"{anchor_col}2")
        self._autosize_sheet(sheet)

    def _style_sheet(self, sheet) -> None:
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        sheet.freeze_panes = "A2"
        self._autosize_sheet(sheet)

    def _autosize_sheet(self, sheet) -> None:
        for column_cells in sheet.columns:
            values = ["" if cell.value is None else str(cell.value) for cell in column_cells]
            width = min(max(len(value) for value in values) + 2, 60)
            sheet.column_dimensions[column_cells[0].column_letter].width = width


def default_source_workbook() -> Path:
    external = Path(
        r"C:\Users\taham\Documents\UNI\EUME\Zahit\Yüzey İşleme\Eloksal Veri\Eloksal Simsim\DBV2-2.xlsm"
    )
    if external.exists():
        return external
    local = Path.cwd() / "DBV2-2.xlsm"
    return local


def default_analysis_workbook() -> Path:
    candidate = Path(r"C:\Users\taham\Documents\UNI\EUME\Zahit\Yüzey İşleme\Eloksal Veri\Eloksal Simsim\analiz\dagilim_analiz_sonuclari.xlsx")
    return candidate


def default_q_workbook() -> Path:
    return Path.cwd() / "Q.xlsx"


def default_output_workbook() -> Path:
    return Path.cwd() / "outputs" / "anodizing_simulasyon_raporu.xlsx"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Event-driven anodizing DES")
    parser.add_argument("--source-workbook", type=Path, default=default_source_workbook())
    parser.add_argument("--analysis-workbook", type=Path, default=default_analysis_workbook())
    parser.add_argument("--output-workbook", type=Path, default=default_output_workbook())
    parser.add_argument("--timeout-csv", type=Path, default=None)
    parser.add_argument(
        "--require-timeouts",
        action="store_true",
        help="Timeout tolerans tablosu zorunlu olsun. Verilmezse timeout mantigi devre disi kalir.",
    )
    parser.add_argument("--arrival-limit", type=int, default=None)
    parser.add_argument("--until-seconds", type=float, default=None)
    parser.add_argument("--hours", type=float, default=0.0, help="Simulasyon suresine saat ekler.")
    parser.add_argument("--days", type=float, default=0.0, help="Simulasyon suresine gun ekler.")
    parser.add_argument("--replications", type=int, default=1, help="Ayni senaryoyu farkli seedlerle kac kez kosturacagi.")
    parser.add_argument(
        "--duration-mode",
        choices=["constant", "distribution"],
        default="constant",
        help="Proses sureleri sabit mi yoksa dagilim bazli mi calissin.",
    )
    parser.add_argument(
        "--use-q-arrival-mix",
        action="store_true",
        help="Gelis karmasini kod klasorundeki Q.xlsx dosyasindan oku.",
    )
    parser.add_argument("--seed", type=int, default=42)
    return parser.parse_args()


def resolve_until_seconds(args: argparse.Namespace) -> float | None:
    total = 0.0
    if args.until_seconds is not None:
        total += float(args.until_seconds)
    if args.hours:
        total += float(args.hours) * 3600.0
    if args.days:
        total += float(args.days) * 86400.0
    return total if total > 0 else None


def main() -> None:
    args = parse_args()
    loader = ExcelDataLoader(
        source_workbook=args.source_workbook,
        analysis_workbook=args.analysis_workbook,
        timeout_csv=args.timeout_csv,
        allow_missing_timeouts=not args.require_timeouts,
        use_constant_durations=args.duration_mode == "constant",
        use_q_arrival_mix=args.use_q_arrival_mix,
    )
    data = loader.load()
    for message in data.warnings:
        warnings.warn(message)

    until_seconds = resolve_until_seconds(args)

    def simulate_once(seed: int) -> AnodizingPlant:
        env = simpy.Environment()
        plant = AnodizingPlant(env, data, seed=seed)
        env.process(plant.arrival_generator(arrival_limit=args.arrival_limit))
        if until_seconds is not None:
            env.run(until=until_seconds)
        else:
            env.run(until=plant.all_done)
        return plant

    plant = simulate_once(args.seed)
    summary = plant.summary()
    replication_results = [
        {
            "replication": 1,
            "seed": args.seed,
            "summary": summary,
            "wip_trace": plant.build_wip_trace(),
        }
    ]
    for replication_index in range(2, max(1, int(args.replications)) + 1):
        replica_seed = int(args.seed) + replication_index - 1
        replica_plant = simulate_once(replica_seed)
        replica_summary = replica_plant.summary()
        replication_results.append(
            {
                "replication": replication_index,
                "seed": replica_seed,
                "summary": replica_summary,
                "wip_trace": replica_plant.build_wip_trace(),
            }
        )

    output_path = plant.export_excel_report(args.output_workbook, replication_results=replication_results)
    print("SIMULATION SUMMARY")
    for key, value in summary.items():
        print(f"{key}: {value}")
    print(f"output_workbook: {output_path}")


if __name__ == "__main__":
    main()
